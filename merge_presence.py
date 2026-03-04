#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
Fusion de fichiers Enregistrement des arrivées/départs + Transactions
→ Génère le fichier Présence au format attendu par rapport_core.py
"""

import openpyxl
from collections import defaultdict
from datetime import datetime, timedelta
import os


def parse_time_str(val):
    """Convertit une valeur en string HH:MM."""
    if val is None or str(val).strip() in ('', '-', 'None'):
        return None
    s = str(val).strip()
    # Format datetime
    if hasattr(val, 'strftime'):
        return val.strftime('%H:%M')
    # Format "HH:MM"
    if ':' in s and len(s) <= 5:
        return s
    return s


def time_to_minutes(t_str):
    """Convertit HH:MM en minutes depuis minuit."""
    if not t_str or t_str == '-':
        return None
    parts = t_str.split(':')
    return int(parts[0]) * 60 + int(parts[1])


def minutes_to_hhmm(mins):
    """Convertit des minutes en HH:MM."""
    if mins is None or mins < 0:
        return '00:00'
    h = int(mins) // 60
    m = int(mins) % 60
    return f"{h:02d}:{m:02d}"


def parse_enregistrement(filepath):
    """
    Parse le fichier Enregistrement des arrivées et départs.
    Retourne: dict[employee_id] = {
        'prenom': str, 'nom': str, 'service': str,
        'dates': dict[date_str] = {
            'sched_start': 'HH:MM', 'sched_end': 'HH:MM',
            'arrival': 'HH:MM' or None, 'departure': 'HH:MM' or None,
            'duration': 'HH:MM' or None
        }
    }
    """
    wb = openpyxl.load_workbook(filepath)
    ws = wb[wb.sheetnames[0]]
    
    # Trouver la ligne d'en-tête
    header_row = None
    for i, row in enumerate(ws.iter_rows(min_row=1, max_row=15, values_only=True), start=1):
        vals = [str(v).strip().lower() if v else '' for v in row]
        if 'prénom' in vals or 'prenom' in vals:
            header_row = i
            break
    
    if not header_row:
        return {}
    
    employees = {}
    
    for row in ws.iter_rows(min_row=header_row + 1, values_only=True):
        prenom = str(row[0]).strip() if row[0] else None
        nom = str(row[1]).strip() if row[1] else '-'
        emp_id = str(row[2]).strip() if row[2] else None
        service = str(row[3]).strip() if row[3] else ''
        date_val = str(row[4]).strip() if row[4] else None
        
        if not prenom or not emp_id or not date_val:
            continue
        
        # Horaire obligatoire
        sched_start = parse_time_str(row[6])  # Heure d'arrivée obligatoire
        sched_end = parse_time_str(row[8])     # Heure de départ obligatoire
        arrival = parse_time_str(row[9])       # Heure de contrôle d'arrivée
        departure = parse_time_str(row[10])    # Sortie à
        
        # Durée
        dur_raw = str(row[11]).strip() if row[11] else '00:00'
        dur_raw = dur_raw.replace(' : ', ':').replace(' :', ':').replace(': ', ':')
        
        if emp_id not in employees:
            employees[emp_id] = {
                'prenom': prenom,
                'nom': nom,
                'service': service,
                'dates': {},
                'schedules': []
            }
        
        # Normaliser date
        date_str = str(date_val)[:10]
        
        employees[emp_id]['dates'][date_str] = {
            'sched_start': sched_start,
            'sched_end': sched_end,
            'arrival': arrival if arrival != '-' else None,
            'departure': departure if departure != '-' else None,
            'duration': dur_raw
        }
        
        # Stocker le planning pour déduire plus tard
        if sched_start and sched_end:
            employees[emp_id]['schedules'].append((sched_start, sched_end))
    
    return employees


def parse_transactions(filepath):
    """
    Parse le fichier Transactions.
    Retourne: dict[employee_id] = {
        'prenom': str, 'nom': str, 'service': str,
        'dates': dict[date_str] = [list de 'HH:MM' triées]
    }
    """
    wb = openpyxl.load_workbook(filepath)
    ws = wb[wb.sheetnames[0]]
    
    header_row = None
    for i, row in enumerate(ws.iter_rows(min_row=1, max_row=15, values_only=True), start=1):
        vals = [str(v).strip().lower() if v else '' for v in row]
        if 'prénom' in vals or 'prenom' in vals:
            header_row = i
            break
    
    if not header_row:
        return {}
    
    employees = {}
    
    for row in ws.iter_rows(min_row=header_row + 1, values_only=True):
        prenom = str(row[0]).strip() if row[0] else None
        nom = str(row[1]).strip() if row[1] else '-'
        emp_id = str(row[2]).strip() if row[2] else None
        service = str(row[3]).strip() if row[3] else ''
        date_val = str(row[4]).strip() if row[4] else None
        heure = parse_time_str(row[5])
        
        if not prenom or not emp_id or not date_val or not heure:
            continue
        
        if emp_id not in employees:
            employees[emp_id] = {
                'prenom': prenom,
                'nom': nom,
                'service': service,
                'dates': defaultdict(list)
            }
        
        date_str = str(date_val)[:10]
        employees[emp_id]['dates'][date_str].append(heure)
    
    # Trier les heures par date
    for emp_id in employees:
        for date_str in employees[emp_id]['dates']:
            employees[emp_id]['dates'][date_str].sort()
    
    return employees


def get_typical_schedule(enr_emp):
    """Détermine le planning typique d'un employé depuis l'Enregistrement."""
    if not enr_emp or not enr_emp.get('schedules'):
        return ('07:00', '17:00')  # Défaut
    
    # Prendre le planning le plus fréquent
    from collections import Counter
    counter = Counter(enr_emp['schedules'])
    return counter.most_common(1)[0][0]


def merge_files(enr_path, trans_path):
    """
    Fusionne Enregistrement + Transactions → données Présence.
    Retourne une liste de lignes au format Présence.
    """
    enr_data = parse_enregistrement(enr_path)
    trans_data = parse_transactions(trans_path)
    
    # Collecter tous les IDs d'employés
    all_ids = set(list(enr_data.keys()) + list(trans_data.keys()))
    
    rows = []
    
    for emp_id in sorted(all_ids):
        enr_emp = enr_data.get(emp_id, {})
        trans_emp = trans_data.get(emp_id, {})
        
        prenom = enr_emp.get('prenom') or trans_emp.get('prenom', '')
        nom = enr_emp.get('nom') or trans_emp.get('nom', '-')
        service = enr_emp.get('service') or trans_emp.get('service', '')
        
        # Plannings typiques de cet employé
        typical_start, typical_end = get_typical_schedule(enr_emp)
        is_night_typical = time_to_minutes(typical_start) > time_to_minutes(typical_end)
        
        # Dédupliquer les transactions
        trans_dates = {}
        if trans_emp and 'dates' in trans_emp:
            for d, times in trans_emp['dates'].items():
                trans_dates[d] = sorted(set(times))
        
        # Collecter toutes les dates
        all_dates = set()
        if enr_emp and 'dates' in enr_emp:
            all_dates.update(enr_emp['dates'].keys())
        all_dates.update(trans_dates.keys())
        
        for date_str in sorted(all_dates, reverse=True):
            enr_day = enr_emp.get('dates', {}).get(date_str, {}) if enr_emp else {}
            times = trans_dates.get(date_str, [])
            
            # --- Planning ---
            sched_start = enr_day.get('sched_start') or typical_start
            sched_end = enr_day.get('sched_end') or typical_end
            
            ss_m = time_to_minutes(sched_start)
            se_m = time_to_minutes(sched_end)
            is_night = ss_m > se_m  # ex: 19:00 > 07:00
            
            # --- Arrivée & Départ depuis Transactions ---
            arrival = None
            departure = None
            
            if times:
                if is_night:
                    # Poste de nuit : arrivée = badge >= 14h, départ = badge < 14h
                    evening_badges = [t for t in times if time_to_minutes(t) >= 840]  # >= 14h
                    morning_badges = [t for t in times if time_to_minutes(t) < 840]   # < 14h
                    
                    arrival = evening_badges[0] if evening_badges else None
                    
                    # Départ = premier badge du lendemain matin
                    next_date = None
                    try:
                        from datetime import datetime as dt, timedelta
                        d = dt.strptime(date_str, '%Y-%m-%d')
                        next_date = (d + timedelta(days=1)).strftime('%Y-%m-%d')
                    except:
                        pass
                    
                    if next_date and next_date in trans_dates:
                        next_morning = [t for t in trans_dates[next_date] if time_to_minutes(t) < 840]
                        if next_morning:
                            departure = next_morning[0]
                    
                    # Fallback sur les badges matin du même jour
                    if not departure and morning_badges:
                        departure = morning_badges[0]
                    
                    # Si pas d'arrivée soir mais des badges matin → c'est le départ d'un poste précédent, ignorer
                    if not arrival and morning_badges:
                        continue  # Ce jour est juste le départ d'une nuit précédente
                        
                else:
                    # Poste de jour : premier badge = arrivée, dernier = départ
                    arrival = times[0]
                    departure = times[-1] if len(times) > 1 else times[0]
            
            # --- Fallback sur Enregistrement si Transactions incomplet ---
            if not arrival and enr_day:
                arr_val = enr_day.get('arrival')
                if arr_val and arr_val != '-':
                    arrival = arr_val
            
            if not departure and enr_day:
                dep_val = enr_day.get('departure')
                if dep_val and dep_val != '-':
                    departure = dep_val
            
            if not arrival:
                arrival = '-'
            if not departure:
                departure = '-'
            
            # --- Calculer la durée ---
            if arrival != '-' and departure != '-':
                arr_m = time_to_minutes(arrival)
                dep_m = time_to_minutes(departure)
                if arr_m is not None and dep_m is not None:
                    if is_night or dep_m < arr_m:
                        dur_m = (24 * 60 - arr_m) + dep_m
                    else:
                        dur_m = dep_m - arr_m
                    # Éviter les durées absurdes (> 18h)
                    if dur_m > 18 * 60:
                        dur_m = 0
                    duration = minutes_to_hhmm(dur_m)
                else:
                    duration = '00:00'
            else:
                duration = '00:00'
            
            rows.append([
                prenom, nom, emp_id, service, date_str,
                sched_start, sched_end, arrival, departure, duration
            ])
    
    return rows


def generate_presence_xlsx(enr_path, trans_path, output_path):
    """Génère le fichier Présence .xlsx à partir des 2 fichiers source."""
    rows = merge_files(enr_path, trans_path)
    
    if not rows:
        return None
    
    # Détecter le nom du service/client
    services = set()
    for r in rows:
        if r[3]:
            parts = r[3].split('>')
            if len(parts) >= 2:
                services.add(parts[1].strip())
    client_name = list(services)[0] if services else "CLIENT"
    
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Présence"
    
    # Titre
    ws.append([f"Présence - {client_name}"])
    
    # En-têtes
    ws.append([
        'Prénom', 'Nom de famille', 'ID', 'Service', 'Date',
        "Heure d'arrivée obligatoire", 'Heure de départ obligatoire',
        "Heure de contrôle d'arrivée", 'Sortie à', 'Durée'
    ])
    
    # Données
    for row in rows:
        ws.append(row)
    
    # Ajuster largeurs
    col_widths = [15, 18, 10, 45, 12, 12, 12, 12, 10, 10]
    for i, w in enumerate(col_widths):
        ws.column_dimensions[chr(65 + i)].width = w
    
    wb.save(output_path)
    
    emp_count = len(set(r[2] for r in rows))
    return {
        'path': output_path,
        'client': client_name,
        'employees': emp_count,
        'rows': len(rows)
    }


# ======================== TEST ========================
if __name__ == '__main__':
    result = generate_presence_xlsx(
        '/mnt/user-data/uploads/Enregistrement_des_arrivées_et_départs_2026-02-01_2026-02-28.xlsx',
        '/mnt/user-data/uploads/Transactions_2026-02-01_2026-02-28.xlsx',
        '/home/claude/test_presence.xlsx'
    )
    print(f"Résultat: {result}")
    
    # Vérifier
    wb = openpyxl.load_workbook('/home/claude/test_presence.xlsx')
    ws = wb.active
    for i, row in enumerate(ws.iter_rows(min_row=1, max_row=10, values_only=True)):
        print(f"  Row {i+1}: {list(row)}")
    print(f"  ... Total: {ws.max_row} lignes")
