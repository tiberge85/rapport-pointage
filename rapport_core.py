#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
=============================================================================
  PROGRAMME DE TRAITEMENT DES RAPPORTS DE POINTAGE
  Générateur de rapport enrichi (heures sup / respect horaire)
=============================================================================
  Entrée  : Fichier Excel (.xlsx) de présence
  Sortie  : PDF enrichi (portrait) avec :
            - Rapports individuels par employé
            - Rapport de présence global
            - Classement retards & absences
            - Graphique d'assiduité
=============================================================================
  Usage : python3 rapport_heures.py [chemin_du_xlsx]
=============================================================================
"""

import sys, os, re, math
import openpyxl
from reportlab.lib.pagesizes import A4
from reportlab.lib.units import mm
from reportlab.lib.colors import HexColor, white, black
from reportlab.lib import colors
from reportlab.platypus import (
    SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer, PageBreak,
    KeepTogether
)
from reportlab.lib.styles import ParagraphStyle
from reportlab.lib.enums import TA_CENTER, TA_LEFT, TA_RIGHT
from reportlab.graphics.shapes import Drawing, Wedge, String, Circle, Rect
from reportlab.graphics import renderPDF
from datetime import datetime
from collections import defaultdict, OrderedDict

# ======================== COULEURS ========================
TEAL = HexColor("#1A7A6D")
DARK_TEAL = HexColor("#0D6B5E")
ORANGE = HexColor("#E8672A")
GREEN = HexColor("#008000")
RED = HexColor("#CC0000")
BLUE = HexColor("#0000CC")
LGRAY = HexColor("#F5F5F5")
MGRAY = HexColor("#DDDDDD")

# ======================== UTILITAIRES ========================

def t2m(t):
    """Convertit HH:MM en minutes."""
    if not t or str(t).strip() in ['-','nan','','None']: return 0
    s = str(t).strip().replace('\n','')
    p = s.split(':')
    try: return int(p[0])*60+int(p[1]) if len(p)==2 else 0
    except: return 0

def m2h(m):
    """Convertit minutes en HH:MM."""
    if m <= 0: return "00:00"
    return f"{int(m)//60:02d}:{int(m)%60:02d}"

def safe(s):
    """Échappe pour XML ReportLab."""
    return str(s).replace('&','&amp;').replace('<','&lt;').replace('>','&gt;')

# ======================== EXTRACTION EXCEL ========================

def extract_from_excel(xlsx_path):
    """Extrait les données depuis le fichier Excel."""
    wb = openpyxl.load_workbook(xlsx_path, data_only=True)
    ws = wb[wb.sheetnames[0]]
    
    # Détection du nom du client
    title = str(ws.cell(1,1).value or "")
    client_name = ""
    for r in range(3, min(6, ws.max_row+1)):
        svc = str(ws.cell(r, 4).value or "")
        if '>' in svc:
            client_name = svc.split('>')[-1].strip()
            break
    if not client_name:
        client_name = "ENTREPRISE"
    
    # Extraction des données par employé
    employees = OrderedDict()
    
    for r in range(3, ws.max_row + 1):
        prenom = str(ws.cell(r, 1).value or "").strip()
        nom = str(ws.cell(r, 2).value or "").strip()
        eid = str(ws.cell(r, 3).value or "").strip()
        
        if not prenom and not nom:
            continue
        
        full_name = f"{prenom} {nom}".strip()
        date_val = str(ws.cell(r, 5).value or "").strip()
        sched_start = str(ws.cell(r, 6).value or "").strip()
        sched_end = str(ws.cell(r, 7).value or "").strip()
        actual_arr = str(ws.cell(r, 8).value or "").strip()
        actual_dep = str(ws.cell(r, 9).value or "").strip()
        duration = str(ws.cell(r, 10).value or "").strip()
        
        # Nettoyer les dates datetime
        if 'datetime' in str(type(ws.cell(r, 5).value)):
            date_val = ws.cell(r, 5).value.strftime('%Y-%m-%d')
        
        # Nettoyer les heures datetime  
        for col_idx, col_name in [(6,'sched_start'),(7,'sched_end'),(8,'actual_arr'),(9,'actual_dep'),(10,'duration')]:
            val = ws.cell(r, col_idx).value
            if val and 'datetime' in str(type(val)):
                locals()[col_name] = val.strftime('%H:%M')
            elif val and 'time' in str(type(val)):
                locals()[col_name] = val.strftime('%H:%M')
        
        # Re-read after potential conversion
        sched_start = str(ws.cell(r, 6).value or "").strip()
        sched_end = str(ws.cell(r, 7).value or "").strip()
        actual_arr = str(ws.cell(r, 8).value or "").strip()
        actual_dep = str(ws.cell(r, 9).value or "").strip()
        duration = str(ws.cell(r, 10).value or "").strip()
        
        # Gérer les formats datetime
        for field in [sched_start, sched_end, actual_arr, actual_dep, duration]:
            pass  # Already strings
        
        key = (full_name, eid)
        if key not in employees:
            employees[key] = {
                'name': full_name,
                'ref': eid,
                'records': []
            }
        
        employees[key]['records'].append({
            'date': date_val[:10] if len(date_val) >= 10 else date_val,
            'sched_start': sched_start[:5] if len(sched_start) >= 5 else sched_start,
            'sched_end': sched_end[:5] if len(sched_end) >= 5 else sched_end,
            'arrival': actual_arr[:5] if len(actual_arr) >= 5 else actual_arr,
            'departure': actual_dep[:5] if len(actual_dep) >= 5 else actual_dep,
            'duration': duration[:5] if len(duration) >= 5 else duration,
        })
    
    # Trier les records par date pour chaque employé
    for key in employees:
        employees[key]['records'].sort(key=lambda x: x['date'])
    
    return list(employees.values()), client_name

# ======================== CALCULS ========================

def calc_employee_stats(emp, hp=0, hp_weekend=0, hourly_cost=0, rest_days=None):
    """Calcule les statistiques complètes d'un employé. rest_days=liste des jours de repos (0=lundi..6=dimanche)."""
    if rest_days is None: rest_days = []
    records = emp['records']
    total_required = 0
    total_worked = 0
    total_overtime = 0
    total_deficit = 0
    total_late_mins = 0
    days_present = 0
    days_late = 0
    days_punctual = 0
    days_absent = 0
    days_badge_error = 0
    days_rest = 0
    hm = hp * 60
    hm_we = hp_weekend * 60
    
    enriched = []
    
    for rec in records:
        ss = t2m(rec['sched_start'])
        se = t2m(rec['sched_end'])
        aa = t2m(rec['arrival'])
        ad = t2m(rec['departure'])
        dur = t2m(rec['duration'])
        
        # Déterminer le jour de la semaine
        is_weekend = False
        is_rest_day = False
        try:
            from datetime import datetime as _dt
            d = _dt.strptime(rec['date'][:10], '%Y-%m-%d')
            is_weekend = d.weekday() >= 5
            if d.weekday() in rest_days:
                is_rest_day = True
        except:
            pass
        
        # Sélectionner les heures obligatoires selon le jour
        if is_rest_day:
            required = 0  # Jour de repos — pas d'heures obligatoires
        elif is_weekend and hp_weekend > 0:
            required = hm_we
        elif not is_weekend and hp > 0:
            required = hm
        elif hp > 0 and hp_weekend == 0:
            required = hm
        else:
            required = se - ss if se > ss else 0
        
        if not is_rest_day:
            total_required += required
        
        schedule_str = f"({rec['sched_start']}_{rec['sched_end']})"
        
        # Déterminer l'état
        if is_rest_day:
            state = "Repos"
            days_rest += 1
            worked = t2m(rec['duration']) if t2m(rec['duration']) > 0 else 0
            if worked > 0:
                total_worked += worked
                days_present += 1
                total_overtime += worked  # Tout travail en jour de repos = heures sup
            overtime = worked
            late = 0
            respect = "REPOS"
        elif dur == 0 or (aa == 0 and ad == 0):
            state = "Absent(e)"
            days_absent += 1
            worked = 0
            overtime = 0
            late = 0
            respect = "ABS"
        else:
            days_present += 1
            
            # === HEURES TRAVAILLÉES ===
            # Le comptage commence au début du planning, PAS avant
            # Ex: planning 7h-17h, arrivée 6h → on compte à partir de 7h
            effective_start = max(aa, ss)
            worked = ad - effective_start if ad > effective_start else 0
            total_worked += worked
            
            # Retard : arrivée après le début prévu
            if aa > ss:
                late = aa - ss
                total_late_mins += late
                days_late += 1
                state = "Retard"
            else:
                late = 0
                days_punctual += 1
                state = "Présent(e)"
            
            # === HEURES SUPPLÉMENTAIRES ===
            # Seulement le temps APRÈS la fin prévue du planning
            # Arriver tôt ne compte PAS comme heure sup
            if ad > se:
                overtime = ad - se
            else:
                overtime = 0
            total_overtime += overtime
            
            # === RESPECT HORAIRE ===
            # Si les heures travaillées >= heures obligatoires (tolérance 5 min) → OUI
            if worked >= (required - 5):
                respect = "OUI"
            else:
                deficit = required - worked
                total_deficit += deficit
                respect = f"NON (-{m2h(deficit)})"
        
        enriched.append({
            'date': rec['date'],
            'schedule': schedule_str,
            'state': state,
            'arrival': rec['arrival'],
            'departure': rec['departure'],
            'worked': m2h(worked),
            'late': m2h(late),
            'required': m2h(required),
            'respect': respect,
            'overtime': m2h(overtime),
        })
    
    # === ASSIDUITÉ (basée sur le taux de présence) ===
    presence_rate = (days_present / len(records) * 100) if len(records) > 0 else 0
    if presence_rate >= 95:
        observation = "Assidu"
    elif presence_rate >= 80:
        observation = "Moyennement assidu"
    else:
        observation = "Non assidu"
    
    stats = {
        'days_required': len(records) - days_rest,
        'days_present': days_present,
        'days_late': days_late,
        'days_punctual': days_punctual,
        'days_absent': days_absent,
        'days_badge_error': days_badge_error,
        'days_rest': days_rest,
        'total_required': total_required,
        'total_worked': total_worked,
        'total_overtime': total_overtime,
        'total_deficit': total_deficit,
        'total_late_mins': total_late_mins,
        'presence_rate': round(presence_rate, 1),
        'observation': observation,
        'hourly_cost': hourly_cost,
        'cost_late': round(total_late_mins / 60 * hourly_cost) if hourly_cost > 0 else 0,
        'cost_deficit': round(total_deficit / 60 * hourly_cost) if hourly_cost > 0 else 0,
        'cost_absent': round(days_absent * (total_required / max(len(records), 1)) / 60 * hourly_cost) if hourly_cost > 0 else 0,
        'cost_overtime': round(total_overtime / 60 * hourly_cost) if hourly_cost > 0 else 0,
    }
    
    return enriched, stats

# ======================== STYLES PDF ========================

def make_styles():
    return {
        'co': ParagraphStyle('co', fontName='Helvetica-Bold', fontSize=10, textColor=TEAL, leading=12),
        'cl': ParagraphStyle('cl', fontName='Helvetica-Bold', fontSize=12, textColor=ORANGE, alignment=TA_RIGHT),
        'ti': ParagraphStyle('ti', fontName='Helvetica-Bold', fontSize=14, textColor=TEAL, alignment=TA_CENTER, spaceAfter=4),
        'st': ParagraphStyle('st', fontName='Helvetica', fontSize=9, alignment=TA_CENTER, spaceAfter=8),
        'ei': ParagraphStyle('ei', fontName='Helvetica-Bold', fontSize=9, spaceAfter=2),
        'eb': ParagraphStyle('eb', fontName='Helvetica-Bold', fontSize=9, textColor=BLUE, spaceAfter=2),
        'c': ParagraphStyle('c', fontName='Helvetica', fontSize=5.8, alignment=TA_CENTER, leading=7),
        'cb': ParagraphStyle('cb', fontName='Helvetica-Bold', fontSize=5.8, alignment=TA_CENTER, leading=7),
        'h': ParagraphStyle('h', fontName='Helvetica-Bold', fontSize=5.8, textColor=white, alignment=TA_CENTER, leading=7),
        'g': ParagraphStyle('g', fontName='Helvetica-Bold', fontSize=5.8, textColor=GREEN, alignment=TA_CENTER, leading=7),
        'r': ParagraphStyle('r', fontName='Helvetica-Bold', fontSize=5.8, textColor=RED, alignment=TA_CENTER, leading=7),
        'b': ParagraphStyle('b', fontName='Helvetica-Bold', fontSize=5.8, textColor=BLUE, alignment=TA_CENTER, leading=7),
        'sh': ParagraphStyle('sh', fontName='Helvetica-Bold', fontSize=6, textColor=white, alignment=TA_CENTER, leading=7),
        'sv': ParagraphStyle('sv', fontName='Helvetica', fontSize=6.5, alignment=TA_CENTER, leading=8),
        'ft': ParagraphStyle('ft', fontName='Helvetica', fontSize=6, alignment=TA_RIGHT, textColor=HexColor("#888")),
        # Styles pour les pages résumé
        'big_ti': ParagraphStyle('big_ti', fontName='Helvetica-Bold', fontSize=16, textColor=TEAL, alignment=TA_CENTER, spaceAfter=12),
        'med_ti': ParagraphStyle('med_ti', fontName='Helvetica-Bold', fontSize=12, textColor=TEAL, alignment=TA_LEFT, spaceAfter=6),
        'rh': ParagraphStyle('rh', fontName='Helvetica-Bold', fontSize=7, textColor=white, alignment=TA_CENTER, leading=9),
        'rv': ParagraphStyle('rv', fontName='Helvetica', fontSize=7, alignment=TA_CENTER, leading=9),
        'rvb': ParagraphStyle('rvb', fontName='Helvetica-Bold', fontSize=7, alignment=TA_CENTER, leading=9),
        'rvo': ParagraphStyle('rvo', fontName='Helvetica-Bold', fontSize=7, textColor=ORANGE, alignment=TA_CENTER, leading=9),
    }

# ======================== HEADER COMMUN ========================

def make_header(S, provider_name, provider_info, client_name, client_info=""):
    """Crée l'en-tête commun pour chaque page."""
    right_text = safe(client_name)
    if client_info:
        right_text += f"<br/><font size=6>{safe(client_info)}</font>"
    h = Table([
        [Paragraph(f"{safe(provider_name)}<br/><font size=6>{safe(provider_info)}</font>", S['co']),
         Paragraph(right_text, S['cl'])]
    ], colWidths=[110*mm, 80*mm])
    h.setStyle(TableStyle([('VALIGN',(0,0),(-1,-1),'MIDDLE'),
                           ('LINEBELOW',(0,0),(-1,0),1,TEAL)]))
    return h

# ======================== PAGE 1-N : RAPPORTS INDIVIDUELS ========================

def gen_individual_pages(story, emps, all_stats, S, provider_name, provider_info, client_name, client_info, period, now):
    """Génère les pages de rapport individuel."""
    total_emps = len(emps)
    
    for idx, emp in enumerate(emps):
        if idx > 0: story.append(PageBreak())
        
        enriched, stats = all_stats[idx]
        emp_num = idx + 1
        
        story.append(make_header(S, provider_name, provider_info, client_name, client_info))
        story.append(Spacer(1, 3*mm))
        story.append(Paragraph("RAPPORT INDIVIDUEL ENRICHI", S['ti']))
        story.append(Paragraph(period, S['st']))
        story.append(Paragraph(f"Employé: {emp['name']}  |  Réf: {emp['ref']}  |  Fiche {emp_num}/{total_emps}", S['ei']))
        story.append(Spacer(1, 2*mm))
        
        # Résumé compact
        sum_hdrs = ["Jours<br/>prévus","Présent","Retard","Absent","Err.<br/>badge",
                    "","H. obligat.","H. travail.","H. retard","H. absent"]
        sum_vals = [
            f"{stats['days_required']}j", f"{stats['days_present']}j",
            f"{stats['days_late']}j", f"{stats['days_absent']}j", f"{stats['days_badge_error']}j",
            "",
            m2h(stats['total_required']), m2h(stats['total_worked']),
            m2h(stats['total_late_mins']),
            m2h(stats['days_absent'] * (stats['total_required'] // max(stats['days_required'],1)))
        ]
        sh = [Paragraph(x, S['sh']) for x in sum_hdrs]
        sv = [Paragraph(x, S['sv']) for x in sum_vals]
        sw = [17*mm,15*mm,14*mm,14*mm,13*mm, 4*mm, 18*mm,18*mm,16*mm,16*mm]
        stbl = Table([sh, sv], colWidths=sw)
        stbl.setStyle(TableStyle([
            ('BACKGROUND',(0,0),(4,0),TEAL),('BACKGROUND',(6,0),(9,0),TEAL),
            ('GRID',(0,0),(4,-1),0.4,colors.grey),('GRID',(6,0),(9,-1),0.4,colors.grey),
            ('ALIGN',(0,0),(-1,-1),'CENTER'),('VALIGN',(0,0),(-1,-1),'MIDDLE'),
            ('TOPPADDING',(0,0),(-1,-1),1),('BOTTOMPADDING',(0,0),(-1,-1),1),
            ('LEFTPADDING',(0,0),(-1,-1),1),('RIGHTPADDING',(0,0),(-1,-1),1),
        ]))
        story.extend([stbl, Spacer(1, 3*mm)])
        
        # Tableau détail
        hdrs = ["N°","Date","Planning","État","Arrivée",
                "Départ","H.<br/>travail.","Retard",
                "H.<br/>obligat.","H.<br/>Respectée","H. sup."]
        cw = [7*mm,16*mm,18*mm,16*mm,13*mm,13*mm,14*mm,13*mm,14*mm,18*mm,14*mm]
        
        td = [[Paragraph(x, S['h']) for x in hdrs]]
        
        for i, rec in enumerate(enriched, 1):
            resp = rec['respect']
            if resp == "OUI":
                rp = Paragraph("OUI", S['g'])
            elif resp == "ABS":
                rp = Paragraph("ABS", S['r'])
            elif resp.startswith("NON"):
                rp = Paragraph(resp.replace(" ","<br/>"), S['r'])
            else:
                rp = Paragraph(resp, S['c'])
            
            ot_mins = t2m(rec['overtime'])
            ot = Paragraph(rec['overtime'], S['b']) if ot_mins > 0 else Paragraph(rec['overtime'], S['c'])
            
            td.append([
                Paragraph(str(i), S['c']),
                Paragraph(rec['date'], S['c']),
                Paragraph(rec['schedule'], S['c']),
                Paragraph(rec['state'], S['cb']),
                Paragraph(rec['arrival'], S['c']),
                Paragraph(rec['departure'], S['c']),
                Paragraph(rec['worked'], S['cb']),
                Paragraph(rec['late'], S['c']),
                Paragraph(rec['required'], S['cb']),
                rp, ot
            ])
        
        dt = Table(td, colWidths=cw, repeatRows=1)
        sc = [('BACKGROUND',(0,0),(-1,0),TEAL),('BACKGROUND',(8,0),(10,0),DARK_TEAL),
              ('GRID',(0,0),(-1,-1),0.3,colors.grey),
              ('ALIGN',(0,0),(-1,-1),'CENTER'),('VALIGN',(0,0),(-1,-1),'MIDDLE'),
              ('TOPPADDING',(0,0),(-1,-1),1),('BOTTOMPADDING',(0,0),(-1,-1),1),
              ('LEFTPADDING',(0,0),(-1,-1),1),('RIGHTPADDING',(0,0),(-1,-1),1)]
        for i in range(2, len(td), 2):
            sc.append(('BACKGROUND',(0,i),(-1,i),LGRAY))
        # Separate days with thick border when date changes
        prev_date = None
        for i, rec in enumerate(enriched, 1):
            cur_date = rec['date'][:10] if rec['date'] else ''
            if prev_date and cur_date != prev_date:
                sc.append(('LINEABOVE',(0,i),(-1,i),1.5,TEAL))
            prev_date = cur_date
        dt.setStyle(TableStyle(sc))
        story.append(dt)
        
        # Totaux
        story.append(Spacer(1, 2*mm))
        tt = Table([[
            Paragraph(f"<b>TOTAL H. SUPPLÉMENTAIRES : {m2h(stats['total_overtime'])}</b>",
                ParagraphStyle('x',fontName='Helvetica-Bold',fontSize=8,textColor=BLUE)),
            Paragraph(f"<b>TOTAL DÉFICIT : {m2h(stats['total_deficit'])}</b>",
                ParagraphStyle('x',fontName='Helvetica-Bold',fontSize=8,textColor=RED)),
        ]], colWidths=[95*mm,95*mm])
        tt.setStyle(TableStyle([('BOX',(0,0),(-1,-1),0.8,TEAL),
            ('INNERGRID',(0,0),(-1,-1),0.4,TEAL),
            ('VALIGN',(0,0),(-1,-1),'MIDDLE'),
            ('TOPPADDING',(0,0),(-1,-1),3),('BOTTOMPADDING',(0,0),(-1,-1),3),
            ('LEFTPADDING',(0,0),(-1,-1),4)]))
        story.extend([tt, Spacer(1,2*mm)])
        
        # === ENCADRÉ COÛT (si hourly_cost > 0) ===
        if stats.get('hourly_cost', 0) > 0:
            fmt_cost = lambda x: f"{x:,.0f} FCFA"
            cost_data = [
                [Paragraph("<b>💰 IMPACT FINANCIER</b>", ParagraphStyle('ct',fontName='Helvetica-Bold',fontSize=8,textColor=colors.white)),
                 Paragraph(f"<b>Coût horaire : {fmt_cost(stats['hourly_cost'])}</b>", ParagraphStyle('ct2',fontName='Helvetica-Bold',fontSize=8,textColor=colors.white,alignment=2))],
                [Paragraph(f"Perte retards ({m2h(stats['total_late_mins'])})", ParagraphStyle('cl',fontSize=7,textColor=DARK_TEAL)),
                 Paragraph(f"<b>{fmt_cost(stats['cost_late'])}</b>", ParagraphStyle('cr',fontSize=8,fontName='Helvetica-Bold',textColor=RED,alignment=2))],
                [Paragraph(f"Perte déficit horaire ({m2h(stats['total_deficit'])})", ParagraphStyle('cl',fontSize=7,textColor=DARK_TEAL)),
                 Paragraph(f"<b>{fmt_cost(stats['cost_deficit'])}</b>", ParagraphStyle('cr',fontSize=8,fontName='Helvetica-Bold',textColor=RED,alignment=2))],
                [Paragraph(f"Perte absences ({stats['days_absent']} jour(s))", ParagraphStyle('cl',fontSize=7,textColor=DARK_TEAL)),
                 Paragraph(f"<b>{fmt_cost(stats['cost_absent'])}</b>", ParagraphStyle('cr',fontSize=8,fontName='Helvetica-Bold',textColor=RED,alignment=2))],
                [Paragraph("<b>TOTAL GAIN PERDU</b>", ParagraphStyle('ct3',fontName='Helvetica-Bold',fontSize=8,textColor=RED)),
                 Paragraph(f"<b>{fmt_cost(stats['cost_late'] + stats['cost_deficit'] + stats['cost_absent'])}</b>",
                    ParagraphStyle('ct4',fontName='Helvetica-Bold',fontSize=9,textColor=RED,alignment=2))],
            ]
            if stats['cost_overtime'] > 0:
                cost_data.append([
                    Paragraph(f"Heures sup. ({m2h(stats['total_overtime'])})", ParagraphStyle('cl',fontSize=7,textColor=DARK_TEAL)),
                    Paragraph(f"<b>+{fmt_cost(stats['cost_overtime'])}</b>", ParagraphStyle('cr',fontSize=8,fontName='Helvetica-Bold',textColor=BLUE,alignment=2))
                ])
            
            ct = Table(cost_data, colWidths=[120*mm, 70*mm])
            ct_style = [
                ('BACKGROUND',(0,0),(-1,0),DARK_TEAL),
                ('BACKGROUND',(0,-1),(-1,-1),HexColor('#fff3e0')),
                ('BOX',(0,0),(-1,-1),1,DARK_TEAL),
                ('INNERGRID',(0,0),(-1,-1),0.3,colors.grey),
                ('VALIGN',(0,0),(-1,-1),'MIDDLE'),
                ('TOPPADDING',(0,0),(-1,-1),3),('BOTTOMPADDING',(0,0),(-1,-1),3),
                ('LEFTPADDING',(0,0),(-1,-1),6),('RIGHTPADDING',(0,0),(-1,-1),6),
            ]
            ct.setStyle(TableStyle(ct_style))
            story.extend([ct, Spacer(1,2*mm)])
        
        story.append(
            Paragraph(f"Généré le {now} | {safe(client_name)} - Rapport {safe(emp['name'])} {emp_num}/{total_emps}", S['ft']))

# ======================== PAGE : RAPPORT DE PRÉSENCE ========================

def gen_rapport_presence(story, emps, all_stats, S, provider_name, provider_info, client_name, client_info, now):
    story.append(PageBreak())
    story.append(make_header(S, provider_name, provider_info, client_name, client_info))
    story.append(Spacer(1, 6*mm))
    story.append(Paragraph("RAPPORT DE PRÉSENCE", S['big_ti']))
    story.append(Spacer(1, 4*mm))
    
    hdrs = ["N°","Employé","Jours<br/>obligat.","Jours de<br/>présence",
            "Taux<br/>présence","Jours de<br/>retards","Jours<br/>ponctuel","Jours<br/>d'absences",
            "Observation"]
    hrow = [Paragraph(h, S['rh']) for h in hdrs]
    cw = [8*mm, 30*mm, 16*mm, 16*mm, 16*mm, 16*mm, 16*mm, 16*mm, 28*mm]
    
    td = [hrow]
    for i, (emp, (enriched, stats)) in enumerate(zip(emps, all_stats), 1):
        obs = stats['observation']
        if obs == "Non assidu":
            obs_style = S['rvo']
        elif obs == "Moyennement assidu":
            obs_style = ParagraphStyle('rvblue', fontName='Helvetica-Bold', fontSize=7, textColor=BLUE, alignment=TA_CENTER, leading=9)
        else:
            obs_style = ParagraphStyle('rvgreen', fontName='Helvetica-Bold', fontSize=7, textColor=GREEN, alignment=TA_CENTER, leading=9)
        td.append([
            Paragraph(str(i), S['rv']),
            Paragraph(emp['name'], S['rvb']),
            Paragraph(f"{stats['days_required']} j", S['rv']),
            Paragraph(f"{stats['days_present']} j", S['rv']),
            Paragraph(f"{stats.get('presence_rate', 0):.0f}%", S['rv']),
            Paragraph(f"{stats['days_late']} j", S['rv']),
            Paragraph(f"{stats['days_punctual']} j", S['rv']),
            Paragraph(f"{stats['days_absent']} j", S['rv']),
            Paragraph(obs, obs_style),
        ])
    
    t = Table(td, colWidths=cw, repeatRows=1)
    sc = [('BACKGROUND',(0,0),(-1,0),TEAL),
          ('GRID',(0,0),(-1,-1),0.4,colors.grey),
          ('ALIGN',(0,0),(-1,-1),'CENTER'),('VALIGN',(0,0),(-1,-1),'MIDDLE'),
          ('TOPPADDING',(0,0),(-1,-1),2),('BOTTOMPADDING',(0,0),(-1,-1),2)]
    for i in range(2, len(td), 2):
        sc.append(('BACKGROUND',(0,i),(-1,i),LGRAY))
    t.setStyle(TableStyle(sc))
    story.extend([t, Spacer(1,4*mm),
        Paragraph(f"Généré le {now} | {safe(client_name)} - Rapport de Présence", S['ft'])])

# ======================== PAGE : CLASSEMENT RETARDS & ABSENCES ========================

def gen_classement(story, emps, all_stats, S, provider_name, provider_info, client_name, client_info, now):
    story.append(PageBreak())
    story.append(make_header(S, provider_name, provider_info, client_name, client_info))
    story.append(Spacer(1, 6*mm))
    story.append(Paragraph("CLASSEMENT PAR DEGRÉ DE RETARDS ET D'ABSENCES", S['big_ti']))
    story.append(Spacer(1, 6*mm))
    
    # Classement par retards
    retards = [(emp['name'], stats['total_late_mins']) 
               for emp, (_, stats) in zip(emps, all_stats) if stats['total_late_mins'] > 0]
    retards.sort(key=lambda x: -x[1])
    
    story.append(Paragraph("Classement par Retards", S['med_ti']))
    hdrs = [Paragraph(h, S['rh']) for h in ["Rang","Nom Employé","Total heure de retard"]]
    td_r = [hdrs]
    for i, (name, mins) in enumerate(retards[:10], 1):
        td_r.append([Paragraph(str(i), S['rv']), Paragraph(name, S['rvb']),
                     Paragraph(m2h(mins), S['rv'])])
    
    if len(td_r) > 1:
        tr = Table(td_r, colWidths=[15*mm, 80*mm, 40*mm])
        tr.setStyle(TableStyle([
            ('BACKGROUND',(0,0),(-1,0),TEAL),
            ('GRID',(0,0),(-1,-1),0.4,colors.grey),
            ('ALIGN',(0,0),(-1,-1),'CENTER'),('VALIGN',(0,0),(-1,-1),'MIDDLE'),
            ('TOPPADDING',(0,0),(-1,-1),3),('BOTTOMPADDING',(0,0),(-1,-1),3)]))
        story.append(tr)
    
    story.append(Spacer(1, 8*mm))
    
    # Classement par absences
    absences = [(emp['name'], stats['days_absent'])
                for emp, (_, stats) in zip(emps, all_stats) if stats['days_absent'] > 0]
    absences.sort(key=lambda x: -x[1])
    
    story.append(Paragraph("Classement par Absences", S['med_ti']))
    hdrs2 = [Paragraph(h, S['rh']) for h in ["Rang","Nom Employé","Nombre de jours d'absence"]]
    td_a = [hdrs2]
    for i, (name, days) in enumerate(absences[:10], 1):
        td_a.append([Paragraph(str(i), S['rv']), Paragraph(name, S['rvb']),
                     Paragraph(str(days), S['rv'])])
    
    if len(td_a) > 1:
        ta = Table(td_a, colWidths=[15*mm, 80*mm, 45*mm])
        ta.setStyle(TableStyle([
            ('BACKGROUND',(0,0),(-1,0),TEAL),
            ('GRID',(0,0),(-1,-1),0.4,colors.grey),
            ('ALIGN',(0,0),(-1,-1),'CENTER'),('VALIGN',(0,0),(-1,-1),'MIDDLE'),
            ('TOPPADDING',(0,0),(-1,-1),3),('BOTTOMPADDING',(0,0),(-1,-1),3)]))
        story.append(ta)
    
    story.extend([Spacer(1,4*mm), Paragraph(f"Généré le {now}", S['ft'])])

# ======================== PRÉPARATION LOGO ========================

def _prepare_logo(logo_path, work_dir=None):
    """Supprime le fond noir du logo et retourne le chemin du logo nettoyé."""
    try:
        from PIL import Image
        import numpy as np
        
        img = Image.open(logo_path).convert('RGBA')
        data = np.array(img)
        
        dark = (data[:,:,0] < 45) & (data[:,:,1] < 45) & (data[:,:,2] < 45)
        data[dark, 3] = 0
        
        img_clean = Image.fromarray(data)
        bbox = img_clean.getbbox()
        if bbox:
            img_clean = img_clean.crop(bbox)
        
        w, h = img_clean.size
        size = max(w, h)
        square = Image.new('RGBA', (size, size), (255, 255, 255, 0))
        square.paste(img_clean, ((size - w) // 2, (size - h) // 2), img_clean)
        
        out_dir = work_dir or os.path.dirname(os.path.abspath(logo_path))
        clean_path = os.path.join(out_dir, 'logo_clean.png')
        square.save(clean_path)
        
        return clean_path
    except Exception as e:
        print(f"  ⚠️  Erreur traitement logo: {e}")
        return None


def _generate_chart_image(pct_presence, pct_absence, logo_path=None, work_dir=None):
    """Génère un graphique donut 3D avec légende et logo au centre."""
    try:
        from PIL import Image, ImageDraw, ImageFont
        
        W, H = 1000, 700
        cx, cy = 350, 320
        outer_r = 260
        inner_r = 130
        depth = 30  # Profondeur 3D
        
        img = Image.new('RGBA', (W, H), (255, 255, 255, 255))
        draw = ImageDraw.Draw(img)
        
        # Couleurs
        teal = (26, 122, 109, 255)
        teal_dark = (18, 90, 80, 255)
        red = (232, 93, 74, 255)
        red_dark = (180, 65, 50, 255)
        green_c = (46, 125, 50, 255)
        orange_c = (232, 103, 42, 255)
        blue_c = (26, 58, 92, 255)
        
        # === 3D DEPTH (ombres dessous) ===
        for d in range(depth, 0, -1):
            shade = int(200 - d * 3)
            draw.ellipse([cx-outer_r, cy-outer_r+d, cx+outer_r, cy+outer_r+d],
                        fill=(shade, shade, shade, 80))
        
        # Disque 3D inférieur (ombre du donut)
        for d in range(depth, 0, -1):
            draw.ellipse([cx-outer_r, cy-outer_r+d, cx+outer_r, cy+outer_r+d], fill=teal_dark)
            if pct_absence > 0:
                a_start = -90
                a_end = -90 + (360 * pct_absence / 100)
                draw.pieslice([cx-outer_r, cy-outer_r+d, cx+outer_r, cy+outer_r+d],
                             start=a_start, end=a_end, fill=red_dark)
        
        # Disque principal (dessus)
        draw.ellipse([cx-outer_r, cy-outer_r, cx+outer_r, cy+outer_r], fill=teal)
        if pct_absence > 0:
            a_start = -90
            a_end = -90 + (360 * pct_absence / 100)
            draw.pieslice([cx-outer_r, cy-outer_r, cx+outer_r, cy+outer_r],
                         start=a_start, end=a_end, fill=red)
        
        # Reflet lumineux (effet 3D)
        for i in range(20):
            alpha = int(40 - i * 2)
            draw.ellipse([cx-outer_r+i+30, cy-outer_r+i+20, cx-30, cy-30],
                        fill=(255, 255, 255, alpha))
        
        # Trou central
        if logo_path and os.path.exists(logo_path):
            clean_path = _prepare_logo(logo_path, work_dir)
            if clean_path:
                logo = Image.open(clean_path).convert('RGBA')
                logo_size = inner_r * 2 + 10
                logo = logo.resize((logo_size, logo_size), Image.LANCZOS)
                lx, ly = cx - logo_size // 2, cy - logo_size // 2
                img.paste(logo, (lx, ly), logo)
            else:
                draw.ellipse([cx-inner_r, cy-inner_r, cx+inner_r, cy+inner_r], fill=(255,255,255,255))
        else:
            draw.ellipse([cx-inner_r, cy-inner_r, cx+inner_r, cy+inner_r], fill=(255,255,255,255))
            try: font = ImageFont.truetype("/usr/share/fonts/truetype/dejavu/DejaVuSans-Bold.ttf", 42)
            except: font = ImageFont.load_default()
            text = f"{pct_presence:.1f}%"
            bbox = draw.textbbox((0, 0), text, font=font)
            tw = bbox[2] - bbox[0]
            th = bbox[3] - bbox[1]
            draw.text((cx - tw//2, cy - th//2), text, fill=teal, font=font)
        
        # === LÉGENDE (à droite) ===
        try:
            font_leg = ImageFont.truetype("/usr/share/fonts/truetype/dejavu/DejaVuSans.ttf", 22)
            font_leg_b = ImageFont.truetype("/usr/share/fonts/truetype/dejavu/DejaVuSans-Bold.ttf", 24)
            font_title = ImageFont.truetype("/usr/share/fonts/truetype/dejavu/DejaVuSans-Bold.ttf", 20)
            font_small = ImageFont.truetype("/usr/share/fonts/truetype/dejavu/DejaVuSans.ttf", 16)
        except:
            font_leg = font_leg_b = font_title = font_small = ImageFont.load_default()
        
        lx_start = 660
        ly = 80
        
        draw.text((lx_start, ly), "Légende", fill=blue_c, font=font_title)
        ly += 45
        
        # Présence
        draw.rounded_rectangle([lx_start, ly, lx_start+28, ly+28], radius=4, fill=teal)
        draw.text((lx_start+36, ly), f"Présence: {pct_presence:.1f}%", fill=(60,60,60), font=font_leg_b)
        ly += 45
        
        # Absence
        draw.rounded_rectangle([lx_start, ly, lx_start+28, ly+28], radius=4, fill=red)
        draw.text((lx_start+36, ly), f"Absence: {pct_absence:.1f}%", fill=(60,60,60), font=font_leg_b)
        ly += 60
        
        # Abréviations
        draw.text((lx_start, ly), "Abréviations", fill=blue_c, font=font_title)
        ly += 35
        abbrevs = [
            ("H. travail.", "Heures travaillées"),
            ("H. obligat.", "Heures obligatoires"),
            ("H. Respectée", "Heures respectées (OUI/NON)"),
            ("H. sup.", "Heures supplémentaires"),
            ("ABS", "Absent"),
            ("P", "Présent"),
            ("R", "Retard"),
        ]
        for abbr, full in abbrevs:
            draw.text((lx_start, ly), f"{abbr}", fill=orange_c, font=font_small)
            draw.text((lx_start + 130, ly), f"= {full}", fill=(100,100,100), font=font_small)
            ly += 28
        
        # Assiduité rules
        ly += 15
        draw.text((lx_start, ly), "Règles d'assiduité", fill=blue_c, font=font_title)
        ly += 32
        rules = [("≥ 95%", "Assidu", green_c), ("80-95%", "Moy. assidu", orange_c), ("< 80%", "Non assidu", red)]
        for pct_label, label, color in rules:
            draw.rounded_rectangle([lx_start, ly, lx_start+12, ly+12], radius=2, fill=color)
            draw.text((lx_start+20, ly-4), f"{pct_label} → {label}", fill=(80,80,80), font=font_small)
            ly += 28
        
        # Convert to RGB
        final = Image.new('RGB', (W, H), (255, 255, 255))
        final.paste(img, mask=img.split()[3])
        
        out_dir = work_dir or os.path.dirname(os.path.abspath(logo_path)) if logo_path else '/tmp'
        chart_path = os.path.join(out_dir, '_chart_donut.png')
        final.save(chart_path, 'PNG', quality=95)
        return chart_path
        
    except Exception as e:
        print(f"  ⚠️  Erreur génération graphique: {e}")
        return None

# ======================== PAGE : GRAPHIQUE D'ASSIDUITÉ ========================

def gen_graphique(story, emps, all_stats, S, provider_name, provider_info, client_name, client_info, now, logo_path=None, work_dir=None):
    story.append(PageBreak())
    story.append(make_header(S, provider_name, provider_info, client_name, client_info))
    story.append(Spacer(1, 6*mm))
    story.append(Paragraph("GRAPHIQUE D'ASSIDUITÉ DU MOIS", S['big_ti']))
    story.append(Spacer(1, 8*mm))
    
    # Calculs globaux
    total_presence = sum(s['total_worked'] for _, s in all_stats)
    total_required = sum(s['total_required'] for _, s in all_stats)
    total_absence = max(0, total_required - total_presence)
    
    if total_required > 0:
        pct_presence = (total_presence / total_required) * 100
        pct_absence = 100 - pct_presence
    else:
        pct_presence = 100
        pct_absence = 0
    
    # Générer le graphique en image PIL pour gérer la transparence du logo
    chart_path = _generate_chart_image(pct_presence, pct_absence, logo_path, work_dir)
    
    if chart_path:
        from reportlab.platypus import Image as PLImage
        img = PLImage(chart_path, width=140*mm, height=140*mm)
        story.append(img)
    
    story.append(Spacer(1, 4*mm))
    
    # Légendes texte
    story.append(Paragraph(
        f"<font color='#E85D4A'><b>Absence ({m2h(total_absence)} - {pct_absence:.0f}%)</b></font>",
        ParagraphStyle('la', fontSize=10, alignment=TA_RIGHT, spaceAfter=2)))
    story.append(Paragraph(
        f"<font color='#1A7A6D'><b>Présence ({m2h(total_presence)} - {pct_presence:.0f}%)</b></font>",
        ParagraphStyle('lp', fontSize=10, alignment=TA_LEFT, spaceAfter=6)))
    
    story.append(Spacer(1, 4*mm))
    
    # Légende en bas
    leg = Table([[
        Paragraph(f"<font color='#1A7A6D'><b>■</b></font>  Total heure de présence: {m2h(total_presence)}", 
                  ParagraphStyle('l1', fontSize=9)),
        Paragraph(f"<font color='#E85D4A'><b>■</b></font>  <font color='#E85D4A'>Total heure d'absence: {m2h(total_absence)}</font>",
                  ParagraphStyle('l2', fontSize=9)),
    ]], colWidths=[95*mm, 95*mm])
    leg.setStyle(TableStyle([('BOX',(0,0),(-1,-1),0.5,colors.grey),
        ('INNERGRID',(0,0),(-1,-1),0.5,colors.grey),
        ('TOPPADDING',(0,0),(-1,-1),4),('BOTTOMPADDING',(0,0),(-1,-1),4),
        ('LEFTPADDING',(0,0),(-1,-1),6)]))
    story.extend([leg, Spacer(1,4*mm), Paragraph(f"Généré le {now} | {safe(client_name)} - Graphique d'Assiduité", S['ft'])])

# ======================== FICHE DE PRÉSENCE SIMPLE ========================

def gen_simple_pages(story, emps, all_stats, S, provider_name, provider_info, client_name, client_info, period, now):
    """Génère une fiche de présence simple : uniquement N°, Date, Planning, Arrivée, Départ — sans retards, absences, totaux."""
    
    for idx, emp in enumerate(emps):
        if idx > 0: story.append(PageBreak())
        
        enriched, stats = all_stats[idx]
        
        story.append(make_header(S, provider_name, provider_info, client_name, client_info))
        story.append(Spacer(1, 3*mm))
        story.append(Paragraph("RAPPORT INDIVIDUEL", S['ti']))
        story.append(Paragraph(period, S['st']))
        story.append(Paragraph(f"Employé: {emp['name']}  |  Réf: {emp['ref']}", S['ei']))
        story.append(Spacer(1, 3*mm))
        
        # Résumé ultra-compact : jours prévus seulement
        sum_data = [[
            Paragraph("<b>Nbre de jours à Effectuer</b>", S['sh']),
            Paragraph(f"<b>{stats['days_required']} jours</b>", S['sv']),
        ]]
        st = Table(sum_data, colWidths=[95*mm, 95*mm])
        st.setStyle(TableStyle([
            ('BACKGROUND',(0,0),(0,0),TEAL), ('BACKGROUND',(1,0),(1,0),HexColor('#f8faf9')),
            ('BOX',(0,0),(-1,-1),0.5,TEAL),
            ('ALIGN',(0,0),(-1,-1),'CENTER'), ('VALIGN',(0,0),(-1,-1),'MIDDLE'),
            ('TOPPADDING',(0,0),(-1,-1),4), ('BOTTOMPADDING',(0,0),(-1,-1),4),
        ]))
        story.extend([st, Spacer(1, 3*mm)])
        
        # Tableau simple : N°, Date, Emploi du temps, Heure d'arrivée, Heure de départ
        hdrs = ["N°", "Date", "Emploi du temps", "Heure d'arrivée", "Heure de départ"]
        cw = [10*mm, 25*mm, 40*mm, 55*mm, 55*mm]
        
        td = [[Paragraph(x, S['h']) for x in hdrs]]
        
        for i, rec in enumerate(enriched, 1):
            td.append([
                Paragraph(str(i), S['c']),
                Paragraph(rec['date'], S['c']),
                Paragraph(rec['schedule'], S['c']),
                Paragraph(rec['arrival'] if rec['arrival'] and rec['state'] != 'Absent(e)' else '-', S['c']),
                Paragraph(rec['departure'] if rec['departure'] and rec['state'] != 'Absent(e)' else '', S['c']),
            ])
        
        dt = Table(td, colWidths=cw, repeatRows=1)
        sc = [('BACKGROUND',(0,0),(-1,0),TEAL),
              ('GRID',(0,0),(-1,-1),0.3,colors.grey),
              ('ALIGN',(0,0),(-1,-1),'CENTER'),('VALIGN',(0,0),(-1,-1),'MIDDLE'),
              ('TOPPADDING',(0,0),(-1,-1),2),('BOTTOMPADDING',(0,0),(-1,-1),2),
              ('LEFTPADDING',(0,0),(-1,-1),2),('RIGHTPADDING',(0,0),(-1,-1),2)]
        for i in range(2, len(td), 2):
            sc.append(('BACKGROUND',(0,i),(-1,i),LGRAY))
        dt.setStyle(TableStyle(sc))
        story.append(dt)
        
        # Coût si applicable
        if stats.get('hourly_cost', 0) > 0 and (stats['cost_late'] > 0 or stats['cost_deficit'] > 0 or stats['cost_absent'] > 0):
            story.append(Spacer(1, 3*mm))
            fmt_cost = lambda x: f"{x:,.0f} FCFA"
            total_loss = stats['cost_late'] + stats['cost_deficit'] + stats['cost_absent']
            cost_line = Table([[
                Paragraph(f"<b>💰 Coût horaire: {fmt_cost(stats['hourly_cost'])}</b>", ParagraphStyle('x',fontName='Helvetica-Bold',fontSize=8,textColor=DARK_TEAL)),
                Paragraph(f"<b>Gain perdu: {fmt_cost(total_loss)}</b>", ParagraphStyle('x',fontName='Helvetica-Bold',fontSize=8,textColor=RED,alignment=2)),
            ]], colWidths=[95*mm, 95*mm])
            cost_line.setStyle(TableStyle([('BOX',(0,0),(-1,-1),0.8,DARK_TEAL),
                ('INNERGRID',(0,0),(-1,-1),0.4,DARK_TEAL),
                ('VALIGN',(0,0),(-1,-1),'MIDDLE'),
                ('TOPPADDING',(0,0),(-1,-1),4),('BOTTOMPADDING',(0,0),(-1,-1),4),
                ('LEFTPADDING',(0,0),(-1,-1),6),('RIGHTPADDING',(0,0),(-1,-1),6)]))
            story.append(cost_line)
        
        story.append(Spacer(1, 3*mm))
        story.append(Paragraph(f"Imprimé par : RH, le {now}", S['ft']))

# ======================== FICHE DE PRÉSENCE SIMPLE ========================

def gen_simple_pages(story, emps, all_stats, S, provider_name, provider_info, client_name, client_info, period, now):
    """Génère une fiche simple : N°, Date, Planning, Arrivée, Départ — sans retard/absence/totaux."""
    
    for idx, emp in enumerate(emps):
        if idx > 0: story.append(PageBreak())
        
        enriched, stats = all_stats[idx]
        
        story.append(make_header(S, provider_name, provider_info, client_name, client_info))
        story.append(Spacer(1, 3*mm))
        story.append(Paragraph("RAPPORT INDIVIDUEL", S['ti']))
        story.append(Paragraph(period, S['st']))
        story.append(Paragraph(f"Employé : {emp['name']}  |  Réf : {emp['ref']}", S['ei']))
        story.append(Spacer(1, 2*mm))
        
        # Summary: just days count
        sum_hdrs = ["Nbre de jours à Effectuer", "Ponctuel", "Retard", "Absent", "Erreurs de Badge"]
        sum_vals = [
            f"{stats['days_required']} jours", f"{stats['days_punctual']} jours",
            f"{stats['days_late']} jours", f"{stats['days_absent']} jours", f"{stats['days_badge_error']} jours",
        ]
        sh = [Paragraph(x, S['sh']) for x in sum_hdrs]
        sv = [Paragraph(x, S['sv']) for x in sum_vals]
        sw = [36*mm, 30*mm, 28*mm, 28*mm, 28*mm]
        stbl = Table([sh, sv], colWidths=sw)
        stbl.setStyle(TableStyle([
            ('BACKGROUND',(0,0),(-1,0),TEAL),
            ('GRID',(0,0),(-1,-1),0.4,colors.grey),
            ('ALIGN',(0,0),(-1,-1),'CENTER'),('VALIGN',(0,0),(-1,-1),'MIDDLE'),
            ('TOPPADDING',(0,0),(-1,-1),3),('BOTTOMPADDING',(0,0),(-1,-1),3),
        ]))
        story.extend([stbl, Spacer(1, 2*mm)])
        
        # Hours summary row
        hrs_hdrs = ["Total heure obligatoire", "Présence", "Retard", "Absent"]
        hrs_vals = [
            m2h(stats['total_required']) + " heures",
            m2h(stats['total_worked']) + " heures",
            m2h(stats['total_late_mins']) + " heures",
            m2h(stats['days_absent'] * (stats['total_required'] // max(stats['days_required'],1))) + " heures",
        ]
        hh = [Paragraph(x, S['sh']) for x in hrs_hdrs]
        hv = [Paragraph(x, S['sv']) for x in hrs_vals]
        hw = [40*mm, 40*mm, 35*mm, 35*mm]
        htbl = Table([hh, hv], colWidths=hw)
        htbl.setStyle(TableStyle([
            ('BACKGROUND',(0,0),(-1,0),TEAL),
            ('GRID',(0,0),(-1,-1),0.4,colors.grey),
            ('ALIGN',(0,0),(-1,-1),'CENTER'),('VALIGN',(0,0),(-1,-1),'MIDDLE'),
            ('TOPPADDING',(0,0),(-1,-1),3),('BOTTOMPADDING',(0,0),(-1,-1),3),
        ]))
        story.extend([htbl, Spacer(1, 3*mm)])
        
        # Detail table: simple — N°, Date, Emploi du temps, Arrivée, Départ
        hdrs = ["N°", "Date", "Emploi du temps", "Heure d'arrivée", "Heure de départ"]
        cw = [12*mm, 28*mm, 40*mm, 40*mm, 40*mm]
        
        td = [[Paragraph(x, S['h']) for x in hdrs]]
        
        for i, rec in enumerate(enriched, 1):
            td.append([
                Paragraph(str(i), S['c']),
                Paragraph(rec['date'], S['c']),
                Paragraph(rec['schedule'], S['c']),
                Paragraph(rec['arrival'] if rec['arrival'] != '00:00' else '-', S['c']),
                Paragraph(rec['departure'] if rec['departure'] != '00:00' else '-', S['c']),
            ])
        
        dt = Table(td, colWidths=cw, repeatRows=1)
        sc = [('BACKGROUND',(0,0),(-1,0),TEAL),
              ('GRID',(0,0),(-1,-1),0.3,colors.grey),
              ('ALIGN',(0,0),(-1,-1),'CENTER'),('VALIGN',(0,0),(-1,-1),'MIDDLE'),
              ('TOPPADDING',(0,0),(-1,-1),2),('BOTTOMPADDING',(0,0),(-1,-1),2),
              ('LEFTPADDING',(0,0),(-1,-1),2),('RIGHTPADDING',(0,0),(-1,-1),2)]
        for i in range(2, len(td), 2):
            sc.append(('BACKGROUND',(0,i),(-1,i),LGRAY))
        dt.setStyle(TableStyle(sc))
        story.append(dt)
        
        # Cost box if applicable
        if stats.get('hourly_cost', 0) > 0:
            story.append(Spacer(1, 2*mm))
            fmt_cost = lambda x: f"{x:,.0f} FCFA"
            total_lost = stats['cost_late'] + stats['cost_deficit'] + stats['cost_absent']
            cost_data = [
                [Paragraph("<b>💰 IMPACT FINANCIER</b>", ParagraphStyle('ct',fontName='Helvetica-Bold',fontSize=8,textColor=colors.white)),
                 Paragraph(f"<b>Coût horaire : {fmt_cost(stats['hourly_cost'])}</b>", ParagraphStyle('ct2',fontName='Helvetica-Bold',fontSize=8,textColor=colors.white,alignment=2))],
                [Paragraph(f"Perte retards ({m2h(stats['total_late_mins'])})", ParagraphStyle('cl',fontSize=7,textColor=DARK_TEAL)),
                 Paragraph(f"<b>{fmt_cost(stats['cost_late'])}</b>", ParagraphStyle('cr',fontSize=8,fontName='Helvetica-Bold',textColor=RED,alignment=2))],
                [Paragraph(f"Perte déficit ({m2h(stats['total_deficit'])})", ParagraphStyle('cl',fontSize=7,textColor=DARK_TEAL)),
                 Paragraph(f"<b>{fmt_cost(stats['cost_deficit'])}</b>", ParagraphStyle('cr',fontSize=8,fontName='Helvetica-Bold',textColor=RED,alignment=2))],
                [Paragraph(f"Perte absences ({stats['days_absent']}j)", ParagraphStyle('cl',fontSize=7,textColor=DARK_TEAL)),
                 Paragraph(f"<b>{fmt_cost(stats['cost_absent'])}</b>", ParagraphStyle('cr',fontSize=8,fontName='Helvetica-Bold',textColor=RED,alignment=2))],
                [Paragraph("<b>TOTAL GAIN PERDU</b>", ParagraphStyle('ct3',fontName='Helvetica-Bold',fontSize=8,textColor=RED)),
                 Paragraph(f"<b>{fmt_cost(total_lost)}</b>", ParagraphStyle('ct4',fontName='Helvetica-Bold',fontSize=9,textColor=RED,alignment=2))],
            ]
            ct = Table(cost_data, colWidths=[100*mm, 60*mm])
            ct.setStyle(TableStyle([
                ('BACKGROUND',(0,0),(-1,0),DARK_TEAL),('BACKGROUND',(0,-1),(-1,-1),HexColor('#fff3e0')),
                ('BOX',(0,0),(-1,-1),1,DARK_TEAL),('INNERGRID',(0,0),(-1,-1),0.3,colors.grey),
                ('VALIGN',(0,0),(-1,-1),'MIDDLE'),('TOPPADDING',(0,0),(-1,-1),3),('BOTTOMPADDING',(0,0),(-1,-1),3),
                ('LEFTPADDING',(0,0),(-1,-1),6),('RIGHTPADDING',(0,0),(-1,-1),6),
            ]))
            story.append(ct)
        
        story.append(Spacer(1, 2*mm))
        story.append(Paragraph(f"Imprimé par : RH, le {now}", S['ft']))


# ======================== GENERATION PDF COMPLETE ========================

def generate_full_pdf(emps, output_path, provider_name, provider_info, client_name, period, logo_path=None, hp=0, client_info="", work_dir=None, hp_weekend=0, hourly_cost=0, employee_costs=None, rest_days=None, employee_rest_days=None):
    if not employee_costs: employee_costs = {}
    if rest_days is None: rest_days = []
    if employee_rest_days is None: employee_rest_days = {}
    if not work_dir:
        work_dir = os.path.dirname(os.path.abspath(output_path))
    doc = SimpleDocTemplate(output_path, pagesize=A4,
        leftMargin=6*mm, rightMargin=6*mm, topMargin=6*mm, bottomMargin=6*mm)
    S = make_styles()
    story = []
    now = datetime.now().strftime("%d/%m/%Y à %H:%M")
    
    # Pré-calculer toutes les stats avec coût par employé et jours de repos
    all_stats = []
    for emp in emps:
        emp_cost = employee_costs.get(emp['name'], hourly_cost)
        emp_rest = employee_rest_days.get(emp['name'], rest_days)
        all_stats.append(calc_employee_stats(emp, hp, hp_weekend, emp_cost, rest_days=emp_rest))
    
    # 1. Rapports individuels
    gen_individual_pages(story, emps, all_stats, S, provider_name, provider_info, client_name, client_info, period, now)
    
    # 2. Rapport de présence
    gen_rapport_presence(story, emps, all_stats, S, provider_name, provider_info, client_name, client_info, now)
    
    # 3. Classement retards & absences
    gen_classement(story, emps, all_stats, S, provider_name, provider_info, client_name, client_info, now)
    
    # 4. Graphique d'assiduité
    gen_graphique(story, emps, all_stats, S, provider_name, provider_info, client_name, client_info, now, logo_path, work_dir)
    
    doc.build(story)

# ======================== MAIN ========================

def main():
    print("\n╔══════════════════════════════════════════════════════════════╗")
    print("║   Générateur de Rapport de Pointage Enrichi               ║")
    print("║   (heures sup / respect horaire / classement / graphique) ║")
    print("╚══════════════════════════════════════════════════════════════╝\n")
    
    # ---- 1. Fichier Excel ----
    file_path = (sys.argv[1] if len(sys.argv)>1 
                 else input("📄 Fichier Excel (.xlsx) — glissez-déposez : ").strip().strip('"').strip("'"))
    
    if not os.path.exists(file_path):
        print(f"\n❌ Fichier '{file_path}' introuvable."); sys.exit(1)
    print(f"  ✅ {os.path.basename(file_path)}")
    
    # ---- 2. Extraction ----
    print(f"\n  🔄 Extraction des données...")
    emps, detected_client = extract_from_excel(file_path)
    
    if not emps:
        print("\n  ❌ Aucun employé trouvé."); sys.exit(1)
    
    print(f"  ✅ {len(emps)} employé(s) détecté(s)")
    
    # ---- 3. Noms des entreprises ----
    print(f"\n  🏢 VOTRE SOCIÉTÉ (Entrée = RAMYA TECHNOLOGIE & INNOVATION) :")
    new_prov = input("     → ").strip()
    provider_name = new_prov if new_prov else "RAMYA TECHNOLOGIE & INNOVATION"
    
    if new_prov:
        new_info = input("     Tél & Email : ").strip()
        provider_info = new_info if new_info else "Tél: 2722204498 | Email: techniqueramya@gmail.com"
    else:
        provider_info = "Tél: 2722204498 | Email: techniqueramya@gmail.com"
    
    print(f"\n  🏬 ENTREPRISE CLIENTE (Entrée = {detected_client}) :")
    new_client = input("     → ").strip()
    client_name = new_client if new_client else detected_client
    
    # ---- 4. Période ----
    # Détecter depuis les dates des enregistrements
    all_dates = []
    for emp in emps:
        for rec in emp['records']:
            all_dates.append(rec['date'])
    if all_dates:
        all_dates.sort()
        period = f"Période du {all_dates[0]} au {all_dates[-1]}"
    else:
        period = "Rapport de pointage"
    
    print(f"\n  📅 {period}")
    
    # ---- 5. Logo ----
    # Chercher le logo dans le même dossier que le fichier Excel, ou le dossier du script
    logo_path = None
    search_dirs = [os.path.dirname(os.path.abspath(file_path)), os.getcwd(), 
                   os.path.dirname(os.path.abspath(__file__))]
    logo_names = ['logo_ramya_ROIND.png', 'logo_ramya.png', 'logo.png']
    for d in search_dirs:
        for ln in logo_names:
            candidate = os.path.join(d, ln)
            if os.path.exists(candidate):
                logo_path = candidate
                break
        if logo_path:
            break
    
    if logo_path:
        print(f"  🖼️  Logo trouvé : {os.path.basename(logo_path)}")
    else:
        print(f"  ℹ️  Pas de logo trouvé (placez logo_ramya_ROIND.png dans le même dossier)")
    
    # ---- 6. Liste des employés ----
    print(f"\n  👥 Employés :")
    for i, emp in enumerate(emps, 1):
        print(f"     {i:2d}. {emp['name']:<25s} ({emp['ref']}) → {len(emp['records'])} jours")
    
    # ---- 6. Génération ----
    base = os.path.splitext(os.path.basename(file_path))[0]
    out_dir = os.path.dirname(os.path.abspath(file_path))
    out = os.path.join(out_dir, f"{base}_RAPPORT_COMPLET.pdf")
    try:
        with open(out, 'wb') as f: pass
        os.remove(out)
    except OSError:
        out = os.path.join(os.getcwd(), f"{base}_RAPPORT_COMPLET.pdf")
    
    print(f"\n  🔄 Génération du PDF complet...")
    generate_full_pdf(emps, out, provider_name, provider_info, client_name, period, logo_path)
    
    print(f"\n  ✅ SUCCÈS → {out}")
    print(f"     🏢 {provider_name} → {client_name}")
    print(f"     👥 {len(emps)} employés")
    print(f"     📄 Contenu : Rapports individuels + Présence + Classement + Graphique\n")
    return out

if __name__ == "__main__":
    main()
#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
Fusion de fichiers Enregistrement des arrivées/départs + Transactions
→ Génère le fichier Présence au format attendu par rapport_core.py
"""

import openpyxl
from collections import defaultdict
from datetime import datetime, timedelta
import os


def parse_time_str(val):
    """Convertit une valeur en string HH:MM."""
    if val is None or str(val).strip() in ('', '-', 'None'):
        return None
    s = str(val).strip()
    # Format datetime
    if hasattr(val, 'strftime'):
        return val.strftime('%H:%M')
    # Format "HH:MM"
    if ':' in s and len(s) <= 5:
        return s
    return s


def time_to_minutes(t_str):
    """Convertit HH:MM en minutes depuis minuit."""
    if not t_str or t_str == '-':
        return None
    parts = t_str.split(':')
    return int(parts[0]) * 60 + int(parts[1])


def minutes_to_hhmm(mins):
    """Convertit des minutes en HH:MM."""
    if mins is None or mins < 0:
        return '00:00'
    h = int(mins) // 60
    m = int(mins) % 60
    return f"{h:02d}:{m:02d}"


def parse_enregistrement(filepath):
    """
    Parse le fichier Enregistrement des arrivées et départs.
    Retourne: dict[employee_id] = {
        'prenom': str, 'nom': str, 'service': str,
        'dates': dict[date_str] = {
            'sched_start': 'HH:MM', 'sched_end': 'HH:MM',
            'arrival': 'HH:MM' or None, 'departure': 'HH:MM' or None,
            'duration': 'HH:MM' or None
        }
    }
    """
    wb = openpyxl.load_workbook(filepath)
    ws = wb[wb.sheetnames[0]]
    
    # Trouver la ligne d'en-tête
    header_row = None
    for i, row in enumerate(ws.iter_rows(min_row=1, max_row=15, values_only=True), start=1):
        vals = [str(v).strip().lower() if v else '' for v in row]
        if 'prénom' in vals or 'prenom' in vals:
            header_row = i
            break
    
    if not header_row:
        return {}
    
    employees = {}
    
    for row in ws.iter_rows(min_row=header_row + 1, values_only=True):
        prenom = str(row[0]).strip() if row[0] else None
        nom = str(row[1]).strip() if row[1] else '-'
        emp_id = str(row[2]).strip() if row[2] else None
        service = str(row[3]).strip() if row[3] else ''
        date_val = str(row[4]).strip() if row[4] else None
        
        if not prenom or not emp_id or not date_val:
            continue
        
        # Horaire obligatoire
        sched_start = parse_time_str(row[6])  # Heure d'arrivée obligatoire
        sched_end = parse_time_str(row[8])     # Heure de départ obligatoire
        arrival = parse_time_str(row[9])       # Heure de contrôle d'arrivée
        departure = parse_time_str(row[10])    # Sortie à
        
        # Durée
        dur_raw = str(row[11]).strip() if row[11] else '00:00'
        dur_raw = dur_raw.replace(' : ', ':').replace(' :', ':').replace(': ', ':')
        
        if emp_id not in employees:
            employees[emp_id] = {
                'prenom': prenom,
                'nom': nom,
                'service': service,
                'dates': {},
                'schedules': []
            }
        
        # Normaliser date
        date_str = str(date_val)[:10]
        
        employees[emp_id]['dates'][date_str] = {
            'sched_start': sched_start,
            'sched_end': sched_end,
            'arrival': arrival if arrival != '-' else None,
            'departure': departure if departure != '-' else None,
            'duration': dur_raw
        }
        
        # Stocker le planning pour déduire plus tard
        if sched_start and sched_end:
            employees[emp_id]['schedules'].append((sched_start, sched_end))
    
    return employees


def parse_transactions(filepath):
    """
    Parse le fichier Transactions.
    Retourne: dict[employee_id] = {
        'prenom': str, 'nom': str, 'service': str,
        'dates': dict[date_str] = [list de 'HH:MM' triées]
    }
    """
    wb = openpyxl.load_workbook(filepath)
    ws = wb[wb.sheetnames[0]]
    
    header_row = None
    for i, row in enumerate(ws.iter_rows(min_row=1, max_row=15, values_only=True), start=1):
        vals = [str(v).strip().lower() if v else '' for v in row]
        if 'prénom' in vals or 'prenom' in vals:
            header_row = i
            break
    
    if not header_row:
        return {}
    
    employees = {}
    
    for row in ws.iter_rows(min_row=header_row + 1, values_only=True):
        prenom = str(row[0]).strip() if row[0] else None
        nom = str(row[1]).strip() if row[1] else '-'
        emp_id = str(row[2]).strip() if row[2] else None
        service = str(row[3]).strip() if row[3] else ''
        date_val = str(row[4]).strip() if row[4] else None
        heure = parse_time_str(row[5])
        
        if not prenom or not emp_id or not date_val or not heure:
            continue
        
        if emp_id not in employees:
            employees[emp_id] = {
                'prenom': prenom,
                'nom': nom,
                'service': service,
                'dates': defaultdict(list)
            }
        
        date_str = str(date_val)[:10]
        employees[emp_id]['dates'][date_str].append(heure)
    
    # Trier les heures par date
    for emp_id in employees:
        for date_str in employees[emp_id]['dates']:
            employees[emp_id]['dates'][date_str].sort()
    
    return employees


def get_typical_schedule(enr_emp):
    """Détermine le planning typique d'un employé depuis l'Enregistrement."""
    if not enr_emp or not enr_emp.get('schedules'):
        return ('07:00', '17:00')  # Défaut
    
    # Prendre le planning le plus fréquent
    from collections import Counter
    counter = Counter(enr_emp['schedules'])
    return counter.most_common(1)[0][0]


def merge_files(enr_path, trans_path):
    """
    Fusionne Enregistrement + Transactions → données Présence.
    Retourne une liste de lignes au format Présence.
    """
    enr_data = parse_enregistrement(enr_path)
    trans_data = parse_transactions(trans_path)
    
    # Collecter tous les IDs d'employés
    all_ids = set(list(enr_data.keys()) + list(trans_data.keys()))
    
    rows = []
    
    for emp_id in sorted(all_ids):
        enr_emp = enr_data.get(emp_id, {})
        trans_emp = trans_data.get(emp_id, {})
        
        prenom = enr_emp.get('prenom') or trans_emp.get('prenom', '')
        nom = enr_emp.get('nom') or trans_emp.get('nom', '-')
        service = enr_emp.get('service') or trans_emp.get('service', '')
        
        # Plannings typiques de cet employé
        typical_start, typical_end = get_typical_schedule(enr_emp)
        is_night_typical = time_to_minutes(typical_start) > time_to_minutes(typical_end)
        
        # Dédupliquer les transactions
        trans_dates = {}
        if trans_emp and 'dates' in trans_emp:
            for d, times in trans_emp['dates'].items():
                trans_dates[d] = sorted(set(times))
        
        # Collecter toutes les dates
        all_dates = set()
        if enr_emp and 'dates' in enr_emp:
            all_dates.update(enr_emp['dates'].keys())
        all_dates.update(trans_dates.keys())
        
        for date_str in sorted(all_dates, reverse=True):
            enr_day = enr_emp.get('dates', {}).get(date_str, {}) if enr_emp else {}
            times = trans_dates.get(date_str, [])
            
            # --- Planning ---
            sched_start = enr_day.get('sched_start') or typical_start
            sched_end = enr_day.get('sched_end') or typical_end
            
            ss_m = time_to_minutes(sched_start)
            se_m = time_to_minutes(sched_end)
            is_night = ss_m > se_m  # ex: 19:00 > 07:00
            
            # --- Arrivée & Départ depuis Transactions ---
            arrival = None
            departure = None
            
            if times:
                if is_night:
                    # Poste de nuit : arrivée = badge >= 14h, départ = badge < 14h
                    evening_badges = [t for t in times if time_to_minutes(t) >= 840]  # >= 14h
                    morning_badges = [t for t in times if time_to_minutes(t) < 840]   # < 14h
                    
                    arrival = evening_badges[0] if evening_badges else None
                    
                    # Départ = premier badge du lendemain matin
                    next_date = None
                    try:
                        from datetime import datetime as dt, timedelta
                        d = dt.strptime(date_str, '%Y-%m-%d')
                        next_date = (d + timedelta(days=1)).strftime('%Y-%m-%d')
                    except:
                        pass
                    
                    if next_date and next_date in trans_dates:
                        next_morning = [t for t in trans_dates[next_date] if time_to_minutes(t) < 840]
                        if next_morning:
                            departure = next_morning[0]
                    
                    # Fallback sur les badges matin du même jour
                    if not departure and morning_badges:
                        departure = morning_badges[0]
                    
                    # Si pas d'arrivée soir mais des badges matin → c'est le départ d'un poste précédent, ignorer
                    if not arrival and morning_badges:
                        continue  # Ce jour est juste le départ d'une nuit précédente
                        
                else:
                    # Poste de jour : premier badge = arrivée, dernier = départ
                    arrival = times[0]
                    departure = times[-1] if len(times) > 1 else times[0]
            
            # --- Fallback sur Enregistrement si Transactions incomplet ---
            if not arrival and enr_day:
                arr_val = enr_day.get('arrival')
                if arr_val and arr_val != '-':
                    arrival = arr_val
            
            if not departure and enr_day:
                dep_val = enr_day.get('departure')
                if dep_val and dep_val != '-':
                    departure = dep_val
            
            if not arrival:
                arrival = '-'
            if not departure:
                departure = '-'
            
            # --- Calculer la durée ---
            if arrival != '-' and departure != '-':
                arr_m = time_to_minutes(arrival)
                dep_m = time_to_minutes(departure)
                if arr_m is not None and dep_m is not None:
                    if is_night or dep_m < arr_m:
                        dur_m = (24 * 60 - arr_m) + dep_m
                    else:
                        dur_m = dep_m - arr_m
                    # Éviter les durées absurdes (> 18h)
                    if dur_m > 18 * 60:
                        dur_m = 0
                    duration = minutes_to_hhmm(dur_m)
                else:
                    duration = '00:00'
            else:
                duration = '00:00'
            
            rows.append([
                prenom, nom, emp_id, service, date_str,
                sched_start, sched_end, arrival, departure, duration
            ])
    
    return rows


def generate_presence_xlsx(enr_path, trans_path, output_path):
    """Génère le fichier Présence .xlsx à partir des 2 fichiers source."""
    rows = merge_files(enr_path, trans_path)
    
    if not rows:
        return None
    
    # Détecter le nom du service/client
    services = set()
    for r in rows:
        if r[3]:
            parts = r[3].split('>')
            if len(parts) >= 2:
                services.add(parts[1].strip())
    client_name = list(services)[0] if services else "CLIENT"
    
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Présence"
    
    # Titre
    ws.append([f"Présence - {client_name}"])
    
    # En-têtes
    ws.append([
        'Prénom', 'Nom de famille', 'ID', 'Service', 'Date',
        "Heure d'arrivée obligatoire", 'Heure de départ obligatoire',
        "Heure de contrôle d'arrivée", 'Sortie à', 'Durée'
    ])
    
    # Données
    for row in rows:
        ws.append(row)
    
    # Ajuster largeurs
    col_widths = [15, 18, 10, 45, 12, 12, 12, 12, 10, 10]
    for i, w in enumerate(col_widths):
        ws.column_dimensions[chr(65 + i)].width = w
    
    wb.save(output_path)
    
    emp_count = len(set(r[2] for r in rows))
    emp_names = sorted(set(f"{r[0]} {r[1]}".strip() for r in rows if r[0]))
    # Group by service for frontend
    emp_by_service = {}
    for r in rows:
        name = f"{r[0]} {r[1]}".strip() if r[0] else None
        service = r[3] if len(r) > 3 else ''
        if name and name not in emp_by_service:
            emp_by_service[name] = service or 'Non défini'
    return {
        'path': output_path,
        'client': client_name,
        'employees': emp_names,
        'emp_services': emp_by_service,
        'emp_count': emp_count,
        'rows': len(rows)
    }


# ======================== TEST ========================
if __name__ == '__main__':
    result = generate_presence_xlsx(
        '/mnt/user-data/uploads/Enregistrement_des_arrivées_et_départs_2026-02-01_2026-02-28.xlsx',
        '/mnt/user-data/uploads/Transactions_2026-02-01_2026-02-28.xlsx',
        '/home/claude/test_presence.xlsx'
    )
    print(f"Résultat: {result}")
    
    # Vérifier
    wb = openpyxl.load_workbook('/home/claude/test_presence.xlsx')
    ws = wb.active
    for i, row in enumerate(ws.iter_rows(min_row=1, max_row=10, values_only=True)):
        print(f"  Row {i+1}: {list(row)}")
    print(f"  ... Total: {ws.max_row} lignes")
#!/usr/bin/env python3
"""Module DPCI — Calcul d'heures avec 4 temps (arrivée, pause début, pause fin, départ)."""

import os
from datetime import datetime
from collections import OrderedDict
import openpyxl

from reportlab.lib.pagesizes import A4
from reportlab.lib.units import mm
from reportlab.lib.colors import HexColor, white, black
from reportlab.lib import colors
from reportlab.platypus import (
    SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer, PageBreak, HRFlowable
)
from reportlab.lib.styles import ParagraphStyle
from reportlab.lib.enums import TA_CENTER, TA_RIGHT, TA_LEFT

TEAL = HexColor('#1a7a6d')
DARK_TEAL = HexColor('#0d5b50')
ORANGE = HexColor('#e8672a')
RED = HexColor('#c53030')
BLUE = HexColor('#1565c0')
LGRAY = HexColor('#f5f6fa')


def t2m(t):
    """Convertit HH:MM en minutes."""
    if not t or t in ('', '-', '00:00'):
        return 0
    try:
        parts = t.replace('h', ':').split(':')
        return int(parts[0]) * 60 + int(parts[1])
    except:
        return 0


def m2h(mins):
    """Convertit minutes en HH:MM."""
    if not mins or mins <= 0:
        return "00:00"
    mins = int(mins)
    return f"{mins // 60:02d}:{mins % 60:02d}"


def parse_dpci_excel(xlsx_path):
    """Parse le fichier Excel DPCI et retourne les employés groupés par département."""
    wb = openpyxl.load_workbook(xlsx_path, read_only=True)
    ws = wb.active

    # Find header row
    header_row = None
    time_period = ""
    for row in ws.iter_rows(min_row=1, max_row=15, values_only=True):
        vals = [str(v or '') for v in row]
        if 'Time Period' in vals[0]:
            time_period = vals[0].replace('Time Period: ', '')
        if vals[0] == 'First Name' or 'First' in vals[0]:
            header_row = True
            break

    employees = OrderedDict()
    for row in ws.iter_rows(min_row=1, max_row=ws.max_row, values_only=True):
        if not row[0] or not row[2]:
            continue
        first = str(row[0]).strip()
        last = str(row[1] or '').strip()
        emp_id = str(row[2]).strip()
        dept = str(row[3] or '').strip()
        date_val = str(row[4] or '').strip()
        record = str(row[5] or '').strip()

        if first == 'First Name' or emp_id == 'ID':
            continue

        # Parse date
        if len(date_val) < 8:
            continue

        # Parse 4 times from record
        times = record.split(';')
        arrival = times[0].strip() if len(times) > 0 else ''
        pause_start = times[1].strip() if len(times) > 1 else ''
        pause_end = times[2].strip() if len(times) > 2 else ''
        departure = times[3].strip() if len(times) > 3 else ''

        full_name = f"{first} {last}".strip()
        key = emp_id

        if key not in employees:
            employees[key] = {
                'name': full_name,
                'id': emp_id,
                'department': dept,
                'records': []
            }

        employees[key]['records'].append({
            'date': date_val[:10],
            'arrival': arrival[:5] if len(arrival) >= 5 else arrival,
            'pause_start': pause_start[:5] if len(pause_start) >= 5 else pause_start,
            'pause_end': pause_end[:5] if len(pause_end) >= 5 else pause_end,
            'departure': departure[:5] if len(departure) >= 5 else departure,
        })

    wb.close()

    # Sort records by date
    for emp in employees.values():
        emp['records'].sort(key=lambda x: x['date'])

    return list(employees.values()), time_period


def calc_dpci_stats(emp, schedule=None, hourly_cost=0, hp=0, hp_weekend=0):
    """Calcule les stats pour un employé DPCI. hp/hp_weekend en heures."""
    records = emp['records']
    total_worked = 0
    total_pause = 0
    total_late = 0
    total_overtime = 0
    total_required = 0
    days_present = 0
    days_late = 0
    days_absent = 0

    # Default schedule from DB or fallback
    sched_start = t2m(schedule.get('start_time', '07:00')) if schedule else t2m('07:00')
    sched_end = t2m(schedule.get('end_time', '17:00')) if schedule else t2m('17:00')
    sched_break_start = t2m(schedule.get('break_start', '12:00')) if schedule else t2m('12:00')
    sched_break_end = t2m(schedule.get('break_end', '13:00')) if schedule else t2m('13:00')
    
    hm = hp * 60  # heures obligatoires semaine en minutes
    hm_we = hp_weekend * 60

    enriched = []

    for rec in records:
        arr = t2m(rec['arrival'])
        ps = t2m(rec['pause_start'])
        pe = t2m(rec['pause_end'])
        dep = t2m(rec['departure'])

        # Detect weekend
        is_weekend = False
        try:
            from datetime import datetime as _dt
            d = _dt.strptime(rec['date'][:10], '%Y-%m-%d')
            is_weekend = d.weekday() >= 5
        except:
            pass

        # Determine required hours for this day
        if is_weekend and hp_weekend > 0:
            required = hm_we
        elif not is_weekend and hp > 0:
            required = hm
        else:
            required = (sched_end - sched_start) - (sched_break_end - sched_break_start)
        
        total_required += required

        if arr == 0 and dep == 0:
            days_absent += 1
            enriched.append({
                'date': rec['date'],
                'arrival': '-', 'pause_start': '-', 'pause_end': '-', 'departure': '-',
                'worked': '00:00', 'pause': '00:00', 'presence': '00:00',
                'required': m2h(required), 'state': 'Absent', 'respect': 'ABS',
            })
            continue

        days_present += 1

        # Pause duration
        pause = pe - ps if pe > ps else 0
        total_pause += pause

        # Worked = morning + afternoon
        morning = ps - arr if ps > arr else 0
        afternoon = dep - pe if dep > pe else 0
        worked = morning + afternoon
        total_worked += worked

        # Presence (total on site)
        presence = dep - arr if dep > arr else 0

        # Late (tracked internally for cost but not displayed)
        late = arr - sched_start if arr > sched_start else 0
        if late > 0:
            total_late += late
            days_late += 1

        # Overtime
        overtime = dep - sched_end if dep > sched_end else 0
        total_overtime += overtime

        # Respect hours
        if worked >= required - 5:
            respect = "OUI"
        else:
            respect = "NON"

        enriched.append({
            'date': rec['date'],
            'arrival': rec['arrival'],
            'pause_start': rec['pause_start'],
            'pause_end': rec['pause_end'],
            'departure': rec['departure'],
            'worked': m2h(worked),
            'pause': m2h(pause),
            'presence': m2h(presence),
            'required': m2h(required),
            'state': 'Présent',
            'respect': respect,
        })

    presence_rate = (days_present / len(records) * 100) if len(records) > 0 else 0

    stats = {
        'days_required': len(records),
        'days_present': days_present,
        'days_late': days_late,
        'days_punctual': days_present - days_late,
        'days_absent': days_absent,
        'total_required': total_required,
        'total_worked': total_worked,
        'total_pause': total_pause,
        'total_late': total_late,
        'total_overtime': total_overtime,
        'presence_rate': round(presence_rate, 1),
        'sched_str': f"{m2h(sched_start)}-{m2h(sched_end)}",
        'hourly_cost': hourly_cost,
        'cost_late': round(total_late / 60 * hourly_cost) if hourly_cost > 0 else 0,
        'cost_absent': round(days_absent * required / 60 * hourly_cost) if hourly_cost > 0 and len(records) > 0 else 0,
    }
    return enriched, stats


def generate_dpci_pdf(emps, output_path, client_name, period, schedules_map=None, employee_costs=None, default_cost=0, hp=0, hp_weekend=0, provider_name='', treated_by='', period_mode='all', rest_days=None):
    """Génère le rapport PDF DPCI — design identique à la fiche de présence."""
    if not schedules_map:
        schedules_map = {}
    if not employee_costs:
        employee_costs = {}
    if rest_days is None:
        rest_days = []

    doc = SimpleDocTemplate(output_path, pagesize=A4,
                            leftMargin=12 * mm, rightMargin=12 * mm, topMargin=10 * mm, bottomMargin=10 * mm)

    # Couleurs exactes de l'image
    HEADER_BG = HexColor('#44546A')   # Barre en-tête gris-bleu foncé
    BLUE_HDR  = HexColor('#4472C4')   # En-têtes tableaux résumé + détail
    BLUE_DARK = HexColor('#305496')   # Sous-en-tête résumé 2
    BORDER_BL = HexColor('#8EAADB')   # Bordures bleu clair
    WHITE     = white
    BLK       = HexColor('#333333')
    LGREY     = HexColor('#F2F2F2')

    hw  = ParagraphStyle('hw', fontName='Helvetica-Bold', fontSize=9, textColor=WHITE, alignment=TA_CENTER)
    hv  = ParagraphStyle('hv', fontSize=9, alignment=TA_CENTER, textColor=BLK)
    th  = ParagraphStyle('th', fontName='Helvetica-Bold', fontSize=7.5, textColor=WHITE, alignment=TA_CENTER, leading=9)
    tc  = ParagraphStyle('tc', fontSize=8, alignment=TA_CENTER, textColor=BLK, leading=10)
    tcb = ParagraphStyle('tcb', fontName='Helvetica-Bold', fontSize=8, alignment=TA_CENTER, textColor=BLK, leading=10)
    ft_s = ParagraphStyle('ft', fontSize=7, textColor=HexColor('#888'), alignment=TA_LEFT)

    story = []
    period_labels = {'all': 'mois', 'week': 'semaine', 'day': 'jour', 'custom': 'période'}
    period_label = period_labels.get(period_mode, 'période')
    now = datetime.now().strftime("%d/%m/%Y \u00e0 %H:%M")

    depts = OrderedDict()
    for emp in emps:
        dept = emp.get('department', 'Non assign\u00e9')
        if dept not in depts:
            depts[dept] = []
        depts[dept].append(emp)

    first_page = True
    pw = 186 * mm
    total_emps = sum(len(v) for v in depts.values())
    emp_counter = 0

    for dept_name, dept_emps in depts.items():
        for emp in dept_emps:
            emp_counter += 1
            if not first_page:
                story.append(PageBreak())
            first_page = False

            sched = schedules_map.get(emp['name'], None)
            cost = employee_costs.get(emp['name'], default_cost)
            enriched, stats = calc_dpci_stats(emp, schedule=sched, hourly_cost=cost, hp=hp, hp_weekend=hp_weekend)

            # BARRE EN-TETE
            prov = provider_name or 'RAMYA TECHNOLOGIE & INNOVATION'
            hbar = Table([[
                Paragraph(f"<b>{prov}</b>", ParagraphStyle('hl', fontName='Helvetica-Bold', fontSize=10, textColor=WHITE)),
                Paragraph(f"<b>{client_name}</b>", ParagraphStyle('hr', fontName='Helvetica-Bold', fontSize=10, textColor=WHITE, alignment=TA_RIGHT)),
            ]], colWidths=[pw * 0.55, pw * 0.45])
            hbar.setStyle(TableStyle([
                ('BACKGROUND', (0, 0), (-1, -1), HEADER_BG),
                ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
                ('TOPPADDING', (0, 0), (-1, -1), 10), ('BOTTOMPADDING', (0, 0), (-1, -1), 10),
                ('LEFTPADDING', (0, 0), (-1, -1), 14), ('RIGHTPADDING', (0, 0), (-1, -1), 14),
            ]))
            story.extend([hbar, Spacer(1, 8 * mm)])

            # TITRE + PERIODE (s\u00e9par\u00e9s)
            story.append(Paragraph("<b>RAPPORT INDIVIDUEL</b>",
                ParagraphStyle('tit', fontName='Helvetica-Bold', fontSize=18, textColor=BLK, alignment=TA_CENTER, spaceAfter=3*mm)))
            story.append(Paragraph(period,
                ParagraphStyle('sub', fontSize=9, textColor=HexColor('#666'), alignment=TA_CENTER, spaceBefore=1*mm)))
            story.append(Spacer(1, 5 * mm))

            # EMPLOYE
            story.append(Paragraph(f"<b>Employ\u00e9 : {emp['name']}</b>",
                ParagraphStyle('emp', fontName='Helvetica-Bold', fontSize=11, textColor=BLK)))
            story.append(Paragraph(f"R\u00e9f\u00e9rence : {emp['id']}",
                ParagraphStyle('ref', fontSize=9, textColor=HexColor('#555'))))
            story.append(Spacer(1, 4 * mm))

            # RESUME 1 : JOURS
            s1_h = ["Nbre de jours \u00e0 Effectuer", "Ponctualité", "Absence"]
            s1_v = [f"{stats['days_required']} jours", f"{stats['days_punctual']} jours", f"{stats['days_absent']} jours"]
            cw1 = [pw * 0.40, pw * 0.30, pw * 0.30]
            t1 = Table([
                [Paragraph(x, hw) for x in s1_h],
                [Paragraph(x, hv) for x in s1_v],
            ], colWidths=cw1)
            t1.setStyle(TableStyle([
                ('BACKGROUND', (0, 0), (-1, 0), BLUE_HDR),
                ('BOX', (0, 0), (-1, -1), 0.6, BORDER_BL),
                ('INNERGRID', (0, 0), (-1, -1), 0.4, BORDER_BL),
                ('ALIGN', (0, 0), (-1, -1), 'CENTER'), ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
                ('TOPPADDING', (0, 0), (-1, -1), 6), ('BOTTOMPADDING', (0, 0), (-1, -1), 6),
            ]))
            story.extend([t1, Spacer(1, 2 * mm)])

            # RESUME 2 : HEURES
            s2_h = ["Total heure obligatoire", "Pr\u00e9sence", "Absence"]
            abs_hrs = m2h(stats['days_absent'] * (stats['total_required'] // max(stats['days_required'], 1)))
            s2_v = [f"{m2h(stats['total_required'])} heures", f"{m2h(stats['total_worked'])} heures", f"{abs_hrs} heures"]
            t2 = Table([
                [Paragraph(x, hw) for x in s2_h],
                [Paragraph(x, hv) for x in s2_v],
            ], colWidths=cw1)
            t2.setStyle(TableStyle([
                ('BACKGROUND', (0, 0), (-1, 0), BLUE_DARK),
                ('BOX', (0, 0), (-1, -1), 0.6, BORDER_BL),
                ('INNERGRID', (0, 0), (-1, -1), 0.4, BORDER_BL),
                ('ALIGN', (0, 0), (-1, -1), 'CENTER'), ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
                ('TOPPADDING', (0, 0), (-1, -1), 6), ('BOTTOMPADDING', (0, 0), (-1, -1), 6),
            ]))
            story.extend([t2, Spacer(1, 4 * mm)])

            # TABLEAU DETAIL
            hdrs = ["Jour", "Date", "Emploi du\ntemps", "Heure\nd'arriv\u00e9e",
                    "D\u00e9but de\npause", "Retour de\npause", "Heure de\nd\u00e9part", "Pause\neffectu\u00e9e",
                    "H.\nobligatoire", "H.\ntravaill\u00e9es", "Emploi du temps\nrespect\u00e9"]
            cw_d = [9*mm, 18*mm, 20*mm, 16*mm, 15*mm, 15*mm, 16*mm, 15*mm, 16*mm, 16*mm, 18*mm]

            td = [[Paragraph(x.replace("\n", "<br/>"), th) for x in hdrs]]

            total_pause_mins = 0

            for i, rec in enumerate(enriched, 1):
                sched_str = stats['sched_str']
                resp = rec['respect']
                if resp == 'OUI':
                    rp = Paragraph("OUI", ParagraphStyle('g', fontName='Helvetica-Bold', fontSize=7, textColor=HexColor('#2e7d32'), alignment=TA_CENTER))
                elif resp == 'ABS':
                    rp = Paragraph("ABS", ParagraphStyle('r', fontName='Helvetica-Bold', fontSize=7, textColor=HexColor('#c53030'), alignment=TA_CENTER))
                else:
                    rp = Paragraph("NON", ParagraphStyle('r', fontName='Helvetica-Bold', fontSize=7, textColor=HexColor('#c53030'), alignment=TA_CENTER))

                req_display = rec.get('required', '') or m2h(stats['total_required'] // max(stats['days_required'], 1))

                # Track pause total
                total_pause_mins += t2m(rec.get('pause', '00:00'))

                td.append([
                    Paragraph(str(i), tc),
                    Paragraph(rec['date'], tc),
                    Paragraph(f"({sched_str.replace('-', '_')})", tc),
                    Paragraph(rec['arrival'] if rec['arrival'] != '-' else '-', tcb),
                    Paragraph(rec['pause_start'] if rec['pause_start'] != '-' else '-', tc),
                    Paragraph(rec['pause_end'] if rec['pause_end'] != '-' else '-', tc),
                    Paragraph(rec['departure'] if rec['departure'] != '-' else '-', tcb),
                    Paragraph(rec.get('pause', '00:00'), tc),
                    Paragraph(req_display, tc),
                    Paragraph(rec['worked'], tcb),
                    rp,
                ])

            dt = Table(td, colWidths=cw_d, repeatRows=1)
            sc = [
                ('BACKGROUND', (0, 0), (-1, 0), BLUE_HDR),
                ('TEXTCOLOR', (0, 0), (-1, 0), WHITE),
                ('BOX', (0, 0), (-1, -1), 0.6, BORDER_BL),
                ('INNERGRID', (0, 0), (-1, -1), 0.3, BORDER_BL),
                ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
                ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
                ('TOPPADDING', (0, 0), (-1, -1), 3),
                ('BOTTOMPADDING', (0, 0), (-1, -1), 3),
                ('LEFTPADDING', (0, 0), (-1, -1), 1),
                ('RIGHTPADDING', (0, 0), (-1, -1), 1),
            ]
            for i in range(2, len(td), 2):
                sc.append(('BACKGROUND', (0, i), (-1, i), LGREY))
            dt.setStyle(TableStyle(sc))
            story.append(dt)

            # RÉSUMÉ CUMULS EN BAS
            story.append(Spacer(1, 3 * mm))
            cum_h = [f"Cumul pause ({period_label})", f"Cumul H. travaill\u00e9es ({period_label})", f"Cumul H. obligatoire ({period_label})", "Taux pr\u00e9sence"]
            cum_v = [f"{m2h(total_pause_mins)}", f"{m2h(stats['total_worked'])}", f"{m2h(stats['total_required'])}", f"{stats['presence_rate']}%"]
            ct_cum = Table([
                [Paragraph(x, hw) for x in cum_h],
                [Paragraph(x, ParagraphStyle('cv', fontName='Helvetica-Bold', fontSize=9, alignment=TA_CENTER, textColor=BLK)) for x in cum_v],
            ], colWidths=[pw * 0.25] * 4)
            ct_cum.setStyle(TableStyle([
                ('BACKGROUND', (0, 0), (-1, 0), BLUE_DARK),
                ('BOX', (0, 0), (-1, -1), 0.6, BORDER_BL),
                ('INNERGRID', (0, 0), (-1, -1), 0.4, BORDER_BL),
                ('ALIGN', (0, 0), (-1, -1), 'CENTER'), ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
                ('TOPPADDING', (0, 0), (-1, -1), 5), ('BOTTOMPADDING', (0, 0), (-1, -1), 5),
            ]))
            story.append(ct_cum)

            # IMPACT FINANCIER (absences uniquement)
            if cost > 0 and stats['cost_absent'] > 0:
                story.append(Spacer(1, 3 * mm))
                fmt = lambda x: f"{x:,.0f} FCFA"
                cd = [
                    [Paragraph("<b>IMPACT FINANCIER</b>", ParagraphStyle('x', fontName='Helvetica-Bold', fontSize=9, textColor=WHITE)),
                     Paragraph(f"<b>Co\u00fbt : {fmt(cost)}/h</b>", ParagraphStyle('x2', fontName='Helvetica-Bold', fontSize=9, textColor=WHITE, alignment=TA_RIGHT))],
                    [Paragraph(f"Perte absences ({stats['days_absent']} jour(s))", ParagraphStyle('x3', fontSize=8, textColor=BLK)),
                     Paragraph(f"<b>{fmt(stats['cost_absent'])}</b>", ParagraphStyle('x4', fontSize=9, fontName='Helvetica-Bold', textColor=HexColor('#c53030'), alignment=TA_RIGHT))],
                    [Paragraph("<b>TOTAL GAIN PERDU</b>", ParagraphStyle('x5', fontName='Helvetica-Bold', fontSize=9, textColor=HexColor('#c53030'))),
                     Paragraph(f"<b>{fmt(stats['cost_absent'])}</b>", ParagraphStyle('x6', fontName='Helvetica-Bold', fontSize=10, textColor=HexColor('#c53030'), alignment=TA_RIGHT))],
                ]
                ct = Table(cd, colWidths=[pw * 0.65, pw * 0.35])
                ct.setStyle(TableStyle([
                    ('BACKGROUND', (0, 0), (-1, 0), HEADER_BG),
                    ('BACKGROUND', (0, -1), (-1, -1), HexColor('#FFF2CC')),
                    ('BOX', (0, 0), (-1, -1), 0.6, BORDER_BL),
                    ('INNERGRID', (0, 0), (-1, -1), 0.3, BORDER_BL),
                    ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
                    ('TOPPADDING', (0, 0), (-1, -1), 5), ('BOTTOMPADDING', (0, 0), (-1, -1), 5),
                    ('LEFTPADDING', (0, 0), (-1, -1), 8), ('RIGHTPADDING', (0, 0), (-1, -1), 8),
                ]))
                story.append(ct)

            # FOOTER
            story.append(Spacer(1, 6 * mm))
            story.append(Paragraph(f"Généré le {now} | {client_name} - Rapport {emp['name']} {emp_counter}/{total_emps}  —  Traité par : {treated_by or 'Admin'}", ft_s))

    doc.build(story)
#!/usr/bin/env python3
"""Générateur de Devis/Proforma PDF — Format RAMYA TECHNOLOGIE"""

import os, json
from datetime import datetime
from reportlab.lib.pagesizes import A4
from reportlab.lib.units import mm
from reportlab.lib.colors import HexColor, white
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer, Image as RLImage
from reportlab.lib.styles import ParagraphStyle
from reportlab.lib.enums import TA_CENTER, TA_LEFT, TA_RIGHT

TEAL = HexColor('#1a7a6d')
ORANGE = HexColor('#e8672a')


def number_to_words_fr(n):
    """Convertit un nombre en mots français (simplifié)."""
    units = ['', 'Un', 'Deux', 'Trois', 'Quatre', 'Cinq', 'Six', 'Sept', 'Huit', 'Neuf',
             'Dix', 'Onze', 'Douze', 'Treize', 'Quatorze', 'Quinze', 'Seize', 'Dix-sept',
             'Dix-huit', 'Dix-neuf']
    tens = ['', '', 'Vingt', 'Trente', 'Quarante', 'Cinquante', 'Soixante',
            'Soixante', 'Quatre-vingt', 'Quatre-vingt']
    
    if n == 0: return 'Zéro'
    if n < 0: return 'Moins ' + number_to_words_fr(-n)
    
    result = ''
    if n >= 1000000:
        m = n // 1000000
        result += ('Un Million' if m == 1 else number_to_words_fr(m) + ' Millions') + ' '
        n %= 1000000
    if n >= 1000:
        t = n // 1000
        result += ('Mille' if t == 1 else number_to_words_fr(t) + ' Mille') + ' '
        n %= 1000
    if n >= 100:
        c = n // 100
        result += ('Cent' if c == 1 else units[c] + ' Cent') + ' '
        n %= 100
    if n >= 20:
        d = n // 10
        if d == 7 or d == 9:
            result += tens[d] + '-' + units[10 + n % 10] + ' '
            n = 0
        else:
            result += tens[d]
            if n % 10 == 1 and d != 8:
                result += ' et Un '
            elif n % 10 > 0:
                result += '-' + units[n % 10] + ' '
            else:
                result += ' '
            n = 0
    if 0 < n < 20:
        result += units[n] + ' '
    
    return result.strip()


def fmt(amount):
    """Format number with thousands separator."""
    return f"{amount:,.0f}".replace(',', ' ')


def generate_devis_pdf(devis_data, output_path, logo_path=None):
    """Génère un PDF de devis/proforma au format RAMYA."""
    
    doc = SimpleDocTemplate(output_path, pagesize=A4,
        leftMargin=15*mm, rightMargin=15*mm, topMargin=10*mm, bottomMargin=12*mm)
    
    story = []
    
    # Styles
    s_title = ParagraphStyle('title', fontSize=20, fontName='Helvetica-Bold', 
                              alignment=TA_RIGHT, textColor=TEAL)
    s_ref = ParagraphStyle('ref', fontSize=10, alignment=TA_RIGHT, textColor=HexColor('#555'))
    s_normal = ParagraphStyle('normal', fontSize=10, leading=13)
    s_bold = ParagraphStyle('bold', fontSize=10, fontName='Helvetica-Bold')
    s_small = ParagraphStyle('small', fontSize=8, textColor=HexColor('#888'))
    s_center = ParagraphStyle('center', fontSize=9, alignment=TA_CENTER)
    s_right = ParagraphStyle('right', fontSize=10, alignment=TA_RIGHT)
    s_footer = ParagraphStyle('footer', fontSize=6, alignment=TA_CENTER, textColor=TEAL)
    
    doc_type = devis_data.get('doc_type', 'devis').upper()
    ref = devis_data.get('reference', '')
    date_str = devis_data.get('date', datetime.now().strftime('%d-%m-%Y'))
    contact = devis_data.get('contact_commercial', '')
    client_name = devis_data.get('client_name', '')
    client_code = devis_data.get('client_code', '')
    objet = devis_data.get('objet', '')
    items = json.loads(devis_data.get('items_json', '[]')) if isinstance(devis_data.get('items_json'), str) else devis_data.get('items_json', [])
    
    total_ht = devis_data.get('total_ht', 0)
    petites_fourn = devis_data.get('petites_fournitures', 0)
    total_ttc = devis_data.get('total_ttc', 0)
    main_oeuvre = devis_data.get('main_oeuvre', 0)
    remise = devis_data.get('remise', 0)
    
    # === HEADER ===
    s_svc = ParagraphStyle('services', fontSize=7, textColor=ORANGE, leading=13, alignment=TA_RIGHT)
    
    logo_el = Paragraph("<b>RAMYA<br/>TECHNOLOGIE &amp; INNOVATION</b>", 
                ParagraphStyle('co', fontSize=12, fontName='Helvetica-Bold', textColor=TEAL))
    if logo_path and os.path.exists(logo_path):
        try:
            logo_el = RLImage(logo_path, width=25*mm, height=25*mm)
        except: pass
    
    header_data = [
        [logo_el,
         Paragraph("<b>RAMYA TECHNOLOGIE &amp; INNOVATION</b><br/><font size='7' color='#888'>Abidjan, Côte d'Ivoire · RCCM: CI-ABJ-03-2017_A10-25092</font>", 
                    ParagraphStyle('co', fontSize=9, fontName='Helvetica-Bold', textColor=TEAL, leading=14)),
         Paragraph("""<font color='#1a7a6d'>■</font> <i>Caméras de surveillance,</i><br/>
<font color='#1a7a6d'>■</font> <i>Clôture électrique,</i><br/>
<font color='#1a7a6d'>■</font> <i>Kit visiophone alarme anti-intrusion,</i><br/>
<font color='#1a7a6d'>■</font> <i>Domotique, Poignées intelligentes</i>""", s_svc)]
    ]
    ht = Table(header_data, colWidths=[30*mm, 75*mm, 75*mm])
    ht.setStyle(TableStyle([('VALIGN', (0,0), (-1,-1), 'MIDDLE')]))
    story.append(ht)
    story.append(Spacer(1, 8*mm))
    
    # === DEVIS / PROFORMA title + info right-aligned ===
    devis_info = [
        [Paragraph(f"<b>{doc_type}</b>", s_title),
         Paragraph(f"<b>{doc_type}#</b> {ref}<br/>Date: {date_str}" + 
                   (f"<br/>Contact commercial: {contact}" if contact else ""),
                   ParagraphStyle('dinfo', fontSize=10, alignment=TA_RIGHT, leading=14))]
    ]
    dt = Table(devis_info, colWidths=[90*mm, 90*mm])
    dt.setStyle(TableStyle([('VALIGN', (0,0), (-1,-1), 'MIDDLE')]))
    story.append(dt)
    story.append(Spacer(1, 8*mm))
    
    # === CLIENT ===
    story.append(Paragraph("<b>À</b>", s_normal))
    story.append(Paragraph(f"<b>{client_name}</b>", ParagraphStyle('cl', fontSize=12, fontName='Helvetica-Bold')))
    if client_code:
        story.append(Spacer(1, 3*mm))
        story.append(Paragraph(f"Code client: {client_code}", s_normal))
    story.append(Spacer(1, 4*mm))
    
    # === OBJET ===
    if objet:
        story.append(Paragraph(f"<b>Objet : {objet}</b>", s_bold))
    story.append(Spacer(1, 6*mm))
    
    # === TABLE DES ARTICLES ===
    hdrs = ['#', 'Désignation', 'Qté.', 'Prix unitaire', 'Remise', 'Montant HT']
    table_data = [[Paragraph(h, ParagraphStyle('th', fontSize=9, fontName='Helvetica-Bold', textColor=white)) for h in hdrs]]
    
    for item in items:
        desc = str(item.get('designation', ''))
        detail = str(item.get('detail', ''))
        full_desc = f"<b>{desc}</b>"
        if detail:
            full_desc += f"<br/>{detail}"
        
        qty = item.get('qty', 1)
        prix = item.get('prix', 0)
        rem = item.get('remise', 0)
        montant = qty * prix - rem
        
        table_data.append([
            Paragraph(str(item.get('num', '')), s_center),
            Paragraph(full_desc, ParagraphStyle('desc', fontSize=9, leading=12)),
            Paragraph(str(qty), s_center),
            Paragraph(fmt(prix), s_right),
            Paragraph(fmt(rem) if rem else '', s_right),
            Paragraph(fmt(montant), s_right),
        ])
    
    col_widths = [12*mm, 68*mm, 14*mm, 28*mm, 20*mm, 28*mm]
    t = Table(table_data, colWidths=col_widths)
    t.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), TEAL),
        ('TEXTCOLOR', (0, 0), (-1, 0), white),
        ('FONTSIZE', (0, 0), (-1, -1), 9),
        ('GRID', (0, 0), (-1, -1), 0.5, HexColor('#cccccc')),
        ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
        ('TOPPADDING', (0, 0), (-1, -1), 5),
        ('BOTTOMPADDING', (0, 0), (-1, -1), 5),
        ('ROWBACKGROUNDS', (0, 1), (-1, -1), [white, HexColor('#f8f8f8')]),
    ]))
    story.append(t)
    story.append(Spacer(1, 6*mm))
    
    # === TOTAUX ===
    total_pieces = total_ht - main_oeuvre
    total_brut = total_ht
    total_net = total_brut - remise
    
    totals = [
        ['', '', '', '', Paragraph("<b>Total HT</b>", s_right), Paragraph(f"<b>{fmt(total_ht)}XOF</b>", s_right)],
        ['', '', '', '', Paragraph("petites fournitures", s_right), Paragraph(f"{fmt(petites_fourn)}XOF", s_right)],
    ]
    tt = Table(totals, colWidths=col_widths)
    tt.setStyle(TableStyle([('LINEABOVE', (4, 0), (5, 0), 1, HexColor('#cccccc'))]))
    story.append(tt)
    
    # Total TTC bar
    ttc_data = [
        [Paragraph("<b>Total TTC</b>", ParagraphStyle('ttc', fontSize=12, fontName='Helvetica-Bold', textColor=white, alignment=TA_RIGHT)),
         Paragraph(f"<b>{fmt(total_ttc)}XOF</b>", ParagraphStyle('ttcv', fontSize=12, fontName='Helvetica-Bold', textColor=white, alignment=TA_RIGHT))]
    ]
    ttc_t = Table(ttc_data, colWidths=[140*mm, 30*mm])
    ttc_t.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, -1), ORANGE),
        ('TOPPADDING', (0, 0), (-1, -1), 6),
        ('BOTTOMPADDING', (0, 0), (-1, -1), 6),
        ('LEFTPADDING', (0, 0), (-1, -1), 10),
        ('RIGHTPADDING', (0, 0), (-1, -1), 10),
    ]))
    story.append(ttc_t)
    story.append(Spacer(1, 4*mm))
    
    # === BARRE RÉSUMÉ ===
    summary_hdrs = ['TOTAL PIÈCES', "MAIN D'ŒUVRE", 'TOTAL BRUT', 'REMISE', 'TOTAL NET', 'PETITES FOURN.', 'TOTAL TTC']
    summary_vals = [fmt(total_pieces), fmt(main_oeuvre), fmt(total_brut), fmt(remise), fmt(total_net), fmt(petites_fourn), fmt(total_ttc)]
    
    s_hdr = ParagraphStyle('sh', fontSize=5, fontName='Helvetica-Bold', textColor=white, alignment=TA_CENTER)
    s_val = ParagraphStyle('sv', fontSize=7, fontName='Helvetica-Bold', textColor=white, alignment=TA_CENTER)
    
    bar_data = [
        [Paragraph(h, s_hdr) for h in summary_hdrs],
        [Paragraph(f"{v}XOF", s_val) for v in summary_vals],
    ]
    bar = Table(bar_data, colWidths=[24*mm]*7)
    bar.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, -1), TEAL),
        ('GRID', (0, 0), (-1, -1), 0.5, HexColor('#0d6b5e')),
        ('TOPPADDING', (0, 0), (-1, -1), 3),
        ('BOTTOMPADDING', (0, 0), (-1, -1), 3),
    ]))
    story.append(bar)
    story.append(Spacer(1, 5*mm))
    
    # === MONTANT EN LETTRES ===
    words = number_to_words_fr(int(total_ttc))
    story.append(Paragraph(
        f"<i>Sauf erreur, arrêté à la somme de: <b>{words} Francs CFA</b></i>",
        ParagraphStyle('words', fontSize=9, alignment=TA_CENTER, textColor=TEAL)
    ))
    story.append(Spacer(1, 15*mm))
    
    # === SIGNATURES ===
    sig_data = [
        [Paragraph("Note:", s_bold), '', Paragraph("Visa Client", s_bold)],
        [Paragraph("MODE DE REGLEMENT (Espèce, Chèque, Virement, Mobile money)", s_small), '', ''],
        ['', '', ''],
        [Paragraph("Signature autorisée", s_bold), '', ''],
    ]
    sig = Table(sig_data, colWidths=[85*mm, 15*mm, 70*mm], rowHeights=[12*mm, 8*mm, 20*mm, 8*mm])
    sig.setStyle(TableStyle([('VALIGN', (0, 0), (-1, -1), 'TOP')]))
    story.append(sig)
    story.append(Spacer(1, 10*mm))
    
    # === FOOTER ===
    story.append(Paragraph(
        "<b>Siège social ABIDJAN Cocody ABATTA derrière la station OLA ENERGY / N°RCCM : CI-ABJ-03-2017_A10-25092 / NCC : 1746141.B</b><br/>"
        "<b>Compte bancaire : Orabank N° : 033201001901 / Bdu N° : 20401160186 / Cel : + 225 2722204498 / 07 09 50 02 43 / 07 47 68 20 27</b><br/>"
        "<b>Email: dg@ramyaci.tech - admin@ramyaci.tech - www.ramyatechnologie.com</b>",
        ParagraphStyle('ft', fontSize=7, alignment=TA_CENTER, textColor=TEAL, leading=10)
    ))
    
    doc.build(story)
    return output_path
