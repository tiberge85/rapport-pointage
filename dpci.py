#!/usr/bin/env python3
"""Module DPCI — Calcul d'heures avec 4 temps (arrivée, pause début, pause fin, départ)."""

import os
from datetime import datetime
from collections import OrderedDict
import openpyxl

from reportlab.lib.pagesizes import A4
from reportlab.lib.units import mm
from reportlab.lib.colors import HexColor, white, black
from reportlab.lib import colors
from reportlab.platypus import (
    SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer, PageBreak, HRFlowable
)
from reportlab.lib.styles import ParagraphStyle
from reportlab.lib.enums import TA_CENTER, TA_RIGHT, TA_LEFT

TEAL = HexColor('#1a7a6d')
DARK_TEAL = HexColor('#0d5b50')
ORANGE = HexColor('#e8672a')
RED = HexColor('#c53030')
BLUE = HexColor('#1565c0')
LGRAY = HexColor('#f5f6fa')


def t2m(t):
    """Convertit HH:MM en minutes."""
    if not t or t in ('', '-', '00:00'):
        return 0
    try:
        parts = t.replace('h', ':').split(':')
        return int(parts[0]) * 60 + int(parts[1])
    except:
        return 0


def m2h(mins):
    """Convertit minutes en HH:MM."""
    if not mins or mins <= 0:
        return "00:00"
    mins = int(mins)
    return f"{mins // 60:02d}:{mins % 60:02d}"


def parse_dpci_excel(xlsx_path):
    """Parse le fichier Excel DPCI et retourne les employés groupés par département."""
    wb = openpyxl.load_workbook(xlsx_path, read_only=True)
    ws = wb.active

    # Find header row
    header_row = None
    time_period = ""
    for row in ws.iter_rows(min_row=1, max_row=15, values_only=True):
        vals = [str(v or '') for v in row]
        if 'Time Period' in vals[0]:
            time_period = vals[0].replace('Time Period: ', '')
        if vals[0] == 'First Name' or 'First' in vals[0]:
            header_row = True
            break

    employees = OrderedDict()
    for row in ws.iter_rows(min_row=1, max_row=ws.max_row, values_only=True):
        if not row[0] or not row[2]:
            continue
        first = str(row[0]).strip()
        last = str(row[1] or '').strip()
        emp_id = str(row[2]).strip()
        dept = str(row[3] or '').strip()
        date_val = str(row[4] or '').strip()
        record = str(row[5] or '').strip()

        if first == 'First Name' or emp_id == 'ID':
            continue

        # Parse date
        if len(date_val) < 8:
            continue

        # Parse 4 times from record
        times = record.split(';')
        arrival = times[0].strip() if len(times) > 0 else ''
        pause_start = times[1].strip() if len(times) > 1 else ''
        pause_end = times[2].strip() if len(times) > 2 else ''
        departure = times[3].strip() if len(times) > 3 else ''

        full_name = f"{first} {last}".strip()
        key = emp_id

        if key not in employees:
            employees[key] = {
                'name': full_name,
                'id': emp_id,
                'department': dept,
                'records': []
            }

        employees[key]['records'].append({
            'date': date_val[:10],
            'arrival': arrival[:5] if len(arrival) >= 5 else arrival,
            'pause_start': pause_start[:5] if len(pause_start) >= 5 else pause_start,
            'pause_end': pause_end[:5] if len(pause_end) >= 5 else pause_end,
            'departure': departure[:5] if len(departure) >= 5 else departure,
        })

    wb.close()

    # Sort records by date
    for emp in employees.values():
        emp['records'].sort(key=lambda x: x['date'])

    return list(employees.values()), time_period


def calc_dpci_stats(emp, schedule=None, hourly_cost=0, hp=0, hp_weekend=0):
    """Calcule les stats pour un employé DPCI. hp/hp_weekend en heures."""
    records = emp['records']
    total_worked = 0
    total_pause = 0
    total_late = 0
    total_overtime = 0
    total_required = 0
    days_present = 0
    days_late = 0
    days_absent = 0

    # Default schedule from DB or fallback
    sched_start = t2m(schedule.get('start_time', '07:00')) if schedule else t2m('07:00')
    sched_end = t2m(schedule.get('end_time', '17:00')) if schedule else t2m('17:00')
    sched_break_start = t2m(schedule.get('break_start', '12:00')) if schedule else t2m('12:00')
    sched_break_end = t2m(schedule.get('break_end', '13:00')) if schedule else t2m('13:00')
    
    hm = hp * 60  # heures obligatoires semaine en minutes
    hm_we = hp_weekend * 60

    enriched = []

    for rec in records:
        arr = t2m(rec['arrival'])
        ps = t2m(rec['pause_start'])
        pe = t2m(rec['pause_end'])
        dep = t2m(rec['departure'])

        # Detect weekend
        is_weekend = False
        try:
            from datetime import datetime as _dt
            d = _dt.strptime(rec['date'][:10], '%Y-%m-%d')
            is_weekend = d.weekday() >= 5
        except:
            pass

        # Determine required hours for this day
        if is_weekend and hp_weekend > 0:
            required = hm_we
        elif not is_weekend and hp > 0:
            required = hm
        else:
            required = (sched_end - sched_start) - (sched_break_end - sched_break_start)
        
        total_required += required

        if arr == 0 and dep == 0:
            days_absent += 1
            enriched.append({
                'date': rec['date'],
                'arrival': '-', 'pause_start': '-', 'pause_end': '-', 'departure': '-',
                'worked': '00:00', 'pause': '00:00', 'presence': '00:00',
                'required': m2h(required), 'state': 'Absent', 'respect': 'ABS',
            })
            continue

        days_present += 1

        # Pause duration
        pause = pe - ps if pe > ps else 0
        total_pause += pause

        # Worked = morning + afternoon
        morning = ps - arr if ps > arr else 0
        afternoon = dep - pe if dep > pe else 0
        worked = morning + afternoon
        total_worked += worked

        # Presence (total on site)
        presence = dep - arr if dep > arr else 0

        # Late (tracked internally for cost but not displayed)
        late = arr - sched_start if arr > sched_start else 0
        if late > 0:
            total_late += late
            days_late += 1

        # Overtime
        overtime = dep - sched_end if dep > sched_end else 0
        total_overtime += overtime

        # Respect hours
        if worked >= required - 5:
            respect = "OUI"
        else:
            respect = "NON"

        enriched.append({
            'date': rec['date'],
            'arrival': rec['arrival'],
            'pause_start': rec['pause_start'],
            'pause_end': rec['pause_end'],
            'departure': rec['departure'],
            'worked': m2h(worked),
            'pause': m2h(pause),
            'presence': m2h(presence),
            'required': m2h(required),
            'state': 'Présent',
            'respect': respect,
        })

    presence_rate = (days_present / len(records) * 100) if len(records) > 0 else 0

    stats = {
        'days_required': len(records),
        'days_present': days_present,
        'days_late': days_late,
        'days_punctual': days_present - days_late,
        'days_absent': days_absent,
        'total_required': total_required,
        'total_worked': total_worked,
        'total_pause': total_pause,
        'total_late': total_late,
        'total_overtime': total_overtime,
        'presence_rate': round(presence_rate, 1),
        'sched_str': f"{m2h(sched_start)}-{m2h(sched_end)}",
        'hourly_cost': hourly_cost,
        'cost_late': round(total_late / 60 * hourly_cost) if hourly_cost > 0 else 0,
        'cost_absent': round(days_absent * required / 60 * hourly_cost) if hourly_cost > 0 and len(records) > 0 else 0,
    }
    return enriched, stats


def generate_dpci_pdf(emps, output_path, client_name, period, schedules_map=None, employee_costs=None, default_cost=0, hp=0, hp_weekend=0, provider_name='', treated_by=''):
    """Génère le rapport PDF DPCI — design identique à la fiche de présence."""
    if not schedules_map:
        schedules_map = {}
    if not employee_costs:
        employee_costs = {}

    doc = SimpleDocTemplate(output_path, pagesize=A4,
                            leftMargin=12 * mm, rightMargin=12 * mm, topMargin=10 * mm, bottomMargin=10 * mm)

    # Couleurs exactes de l'image
    HEADER_BG = HexColor('#44546A')   # Barre en-tête gris-bleu foncé
    BLUE_HDR  = HexColor('#4472C4')   # En-têtes tableaux résumé + détail
    BLUE_DARK = HexColor('#305496')   # Sous-en-tête résumé 2
    BORDER_BL = HexColor('#8EAADB')   # Bordures bleu clair
    WHITE     = white
    BLK       = HexColor('#333333')
    LGREY     = HexColor('#F2F2F2')

    hw  = ParagraphStyle('hw', fontName='Helvetica-Bold', fontSize=9, textColor=WHITE, alignment=TA_CENTER)
    hv  = ParagraphStyle('hv', fontSize=9, alignment=TA_CENTER, textColor=BLK)
    th  = ParagraphStyle('th', fontName='Helvetica-Bold', fontSize=7.5, textColor=WHITE, alignment=TA_CENTER, leading=9)
    tc  = ParagraphStyle('tc', fontSize=8, alignment=TA_CENTER, textColor=BLK, leading=10)
    tcb = ParagraphStyle('tcb', fontName='Helvetica-Bold', fontSize=8, alignment=TA_CENTER, textColor=BLK, leading=10)
    ft_s = ParagraphStyle('ft', fontSize=7, textColor=HexColor('#888'), alignment=TA_LEFT)

    story = []
    now = datetime.now().strftime("%d/%m/%Y \u00e0 %H:%M")

    depts = OrderedDict()
    for emp in emps:
        dept = emp.get('department', 'Non assign\u00e9')
        if dept not in depts:
            depts[dept] = []
        depts[dept].append(emp)

    first_page = True
    pw = 186 * mm

    for dept_name, dept_emps in depts.items():
        for emp in dept_emps:
            if not first_page:
                story.append(PageBreak())
            first_page = False

            sched = schedules_map.get(emp['name'], None)
            cost = employee_costs.get(emp['name'], default_cost)
            enriched, stats = calc_dpci_stats(emp, schedule=sched, hourly_cost=cost, hp=hp, hp_weekend=hp_weekend)

            # BARRE EN-TETE
            prov = provider_name or 'RAMYA TECHNOLOGIE & INNOVATION'
            hbar = Table([[
                Paragraph(f"<b>{prov}</b>", ParagraphStyle('hl', fontName='Helvetica-Bold', fontSize=10, textColor=WHITE)),
                Paragraph(f"<b>{client_name}</b>", ParagraphStyle('hr', fontName='Helvetica-Bold', fontSize=10, textColor=WHITE, alignment=TA_RIGHT)),
            ]], colWidths=[pw * 0.55, pw * 0.45])
            hbar.setStyle(TableStyle([
                ('BACKGROUND', (0, 0), (-1, -1), HEADER_BG),
                ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
                ('TOPPADDING', (0, 0), (-1, -1), 10), ('BOTTOMPADDING', (0, 0), (-1, -1), 10),
                ('LEFTPADDING', (0, 0), (-1, -1), 14), ('RIGHTPADDING', (0, 0), (-1, -1), 14),
            ]))
            story.extend([hbar, Spacer(1, 8 * mm)])

            # TITRE + PERIODE (s\u00e9par\u00e9s)
            story.append(Paragraph("<b>RAPPORT INDIVIDUEL</b>",
                ParagraphStyle('tit', fontName='Helvetica-Bold', fontSize=18, textColor=BLK, alignment=TA_CENTER, spaceAfter=3*mm)))
            story.append(Paragraph(period,
                ParagraphStyle('sub', fontSize=9, textColor=HexColor('#666'), alignment=TA_CENTER, spaceBefore=1*mm)))
            story.append(Spacer(1, 5 * mm))

            # EMPLOYE
            story.append(Paragraph(f"<b>Employ\u00e9 : {emp['name']}</b>",
                ParagraphStyle('emp', fontName='Helvetica-Bold', fontSize=11, textColor=BLK)))
            story.append(Paragraph(f"R\u00e9f\u00e9rence : {emp['id']}",
                ParagraphStyle('ref', fontSize=9, textColor=HexColor('#555'))))
            story.append(Spacer(1, 4 * mm))

            # RESUME 1 : JOURS
            s1_h = ["Nbre de jours \u00e0 Effectuer", "Ponctualité", "Absence"]
            s1_v = [f"{stats['days_required']} jours", f"{stats['days_punctual']} jours", f"{stats['days_absent']} jours"]
            cw1 = [pw * 0.40, pw * 0.30, pw * 0.30]
            t1 = Table([
                [Paragraph(x, hw) for x in s1_h],
                [Paragraph(x, hv) for x in s1_v],
            ], colWidths=cw1)
            t1.setStyle(TableStyle([
                ('BACKGROUND', (0, 0), (-1, 0), BLUE_HDR),
                ('BOX', (0, 0), (-1, -1), 0.6, BORDER_BL),
                ('INNERGRID', (0, 0), (-1, -1), 0.4, BORDER_BL),
                ('ALIGN', (0, 0), (-1, -1), 'CENTER'), ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
                ('TOPPADDING', (0, 0), (-1, -1), 6), ('BOTTOMPADDING', (0, 0), (-1, -1), 6),
            ]))
            story.extend([t1, Spacer(1, 2 * mm)])

            # RESUME 2 : HEURES
            s2_h = ["Total heure obligatoire", "Pr\u00e9sence", "Absence"]
            abs_hrs = m2h(stats['days_absent'] * (stats['total_required'] // max(stats['days_required'], 1)))
            s2_v = [f"{m2h(stats['total_required'])} heures", f"{m2h(stats['total_worked'])} heures", f"{abs_hrs} heures"]
            t2 = Table([
                [Paragraph(x, hw) for x in s2_h],
                [Paragraph(x, hv) for x in s2_v],
            ], colWidths=cw1)
            t2.setStyle(TableStyle([
                ('BACKGROUND', (0, 0), (-1, 0), BLUE_DARK),
                ('BOX', (0, 0), (-1, -1), 0.6, BORDER_BL),
                ('INNERGRID', (0, 0), (-1, -1), 0.4, BORDER_BL),
                ('ALIGN', (0, 0), (-1, -1), 'CENTER'), ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
                ('TOPPADDING', (0, 0), (-1, -1), 6), ('BOTTOMPADDING', (0, 0), (-1, -1), 6),
            ]))
            story.extend([t2, Spacer(1, 4 * mm)])

            # TABLEAU DETAIL
            hdrs = ["Jour", "Date", "Emploi du\ntemps", "Heure\nd'arriv\u00e9e",
                    "D\u00e9but de\npause", "Retour de\npause", "Heure de\nd\u00e9part", "Pause\neffectu\u00e9e",
                    "H.\nobligatoire", "H.\ntravaill\u00e9es", "Emploi du temps\nrespect\u00e9"]
            cw_d = [9*mm, 18*mm, 20*mm, 16*mm, 15*mm, 15*mm, 16*mm, 15*mm, 16*mm, 16*mm, 18*mm]

            td = [[Paragraph(x.replace("\n", "<br/>"), th) for x in hdrs]]

            total_pause_mins = 0

            for i, rec in enumerate(enriched, 1):
                sched_str = stats['sched_str']
                resp = rec['respect']
                if resp == 'OUI':
                    rp = Paragraph("OUI", ParagraphStyle('g', fontName='Helvetica-Bold', fontSize=7, textColor=HexColor('#2e7d32'), alignment=TA_CENTER))
                elif resp == 'ABS':
                    rp = Paragraph("ABS", ParagraphStyle('r', fontName='Helvetica-Bold', fontSize=7, textColor=HexColor('#c53030'), alignment=TA_CENTER))
                else:
                    rp = Paragraph("NON", ParagraphStyle('r', fontName='Helvetica-Bold', fontSize=7, textColor=HexColor('#c53030'), alignment=TA_CENTER))

                req_display = rec.get('required', '') or m2h(stats['total_required'] // max(stats['days_required'], 1))

                # Track pause total
                total_pause_mins += t2m(rec.get('pause', '00:00'))

                td.append([
                    Paragraph(str(i), tc),
                    Paragraph(rec['date'], tc),
                    Paragraph(f"({sched_str.replace('-', '_')})", tc),
                    Paragraph(rec['arrival'] if rec['arrival'] != '-' else '-', tcb),
                    Paragraph(rec['pause_start'] if rec['pause_start'] != '-' else '-', tc),
                    Paragraph(rec['pause_end'] if rec['pause_end'] != '-' else '-', tc),
                    Paragraph(rec['departure'] if rec['departure'] != '-' else '-', tcb),
                    Paragraph(rec.get('pause', '00:00'), tc),
                    Paragraph(req_display, tc),
                    Paragraph(rec['worked'], tcb),
                    rp,
                ])

            dt = Table(td, colWidths=cw_d, repeatRows=1)
            sc = [
                ('BACKGROUND', (0, 0), (-1, 0), BLUE_HDR),
                ('TEXTCOLOR', (0, 0), (-1, 0), WHITE),
                ('BOX', (0, 0), (-1, -1), 0.6, BORDER_BL),
                ('INNERGRID', (0, 0), (-1, -1), 0.3, BORDER_BL),
                ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
                ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
                ('TOPPADDING', (0, 0), (-1, -1), 3),
                ('BOTTOMPADDING', (0, 0), (-1, -1), 3),
                ('LEFTPADDING', (0, 0), (-1, -1), 1),
                ('RIGHTPADDING', (0, 0), (-1, -1), 1),
            ]
            for i in range(2, len(td), 2):
                sc.append(('BACKGROUND', (0, i), (-1, i), LGREY))
            dt.setStyle(TableStyle(sc))
            story.append(dt)

            # RÉSUMÉ CUMULS EN BAS
            story.append(Spacer(1, 3 * mm))
            cum_h = ["Cumul pause effectu\u00e9e", "Cumul H. travaill\u00e9es", "Cumul H. obligatoire", "Taux pr\u00e9sence"]
            cum_v = [f"{m2h(total_pause_mins)}", f"{m2h(stats['total_worked'])}", f"{m2h(stats['total_required'])}", f"{stats['presence_rate']}%"]
            ct_cum = Table([
                [Paragraph(x, hw) for x in cum_h],
                [Paragraph(x, ParagraphStyle('cv', fontName='Helvetica-Bold', fontSize=9, alignment=TA_CENTER, textColor=BLK)) for x in cum_v],
            ], colWidths=[pw * 0.25] * 4)
            ct_cum.setStyle(TableStyle([
                ('BACKGROUND', (0, 0), (-1, 0), BLUE_DARK),
                ('BOX', (0, 0), (-1, -1), 0.6, BORDER_BL),
                ('INNERGRID', (0, 0), (-1, -1), 0.4, BORDER_BL),
                ('ALIGN', (0, 0), (-1, -1), 'CENTER'), ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
                ('TOPPADDING', (0, 0), (-1, -1), 5), ('BOTTOMPADDING', (0, 0), (-1, -1), 5),
            ]))
            story.append(ct_cum)

            # IMPACT FINANCIER (absences uniquement)
            if cost > 0 and stats['cost_absent'] > 0:
                story.append(Spacer(1, 3 * mm))
                fmt = lambda x: f"{x:,.0f} FCFA"
                cd = [
                    [Paragraph("<b>IMPACT FINANCIER</b>", ParagraphStyle('x', fontName='Helvetica-Bold', fontSize=9, textColor=WHITE)),
                     Paragraph(f"<b>Co\u00fbt : {fmt(cost)}/h</b>", ParagraphStyle('x2', fontName='Helvetica-Bold', fontSize=9, textColor=WHITE, alignment=TA_RIGHT))],
                    [Paragraph(f"Perte absences ({stats['days_absent']} jour(s))", ParagraphStyle('x3', fontSize=8, textColor=BLK)),
                     Paragraph(f"<b>{fmt(stats['cost_absent'])}</b>", ParagraphStyle('x4', fontSize=9, fontName='Helvetica-Bold', textColor=HexColor('#c53030'), alignment=TA_RIGHT))],
                    [Paragraph("<b>TOTAL GAIN PERDU</b>", ParagraphStyle('x5', fontName='Helvetica-Bold', fontSize=9, textColor=HexColor('#c53030'))),
                     Paragraph(f"<b>{fmt(stats['cost_absent'])}</b>", ParagraphStyle('x6', fontName='Helvetica-Bold', fontSize=10, textColor=HexColor('#c53030'), alignment=TA_RIGHT))],
                ]
                ct = Table(cd, colWidths=[pw * 0.65, pw * 0.35])
                ct.setStyle(TableStyle([
                    ('BACKGROUND', (0, 0), (-1, 0), HEADER_BG),
                    ('BACKGROUND', (0, -1), (-1, -1), HexColor('#FFF2CC')),
                    ('BOX', (0, 0), (-1, -1), 0.6, BORDER_BL),
                    ('INNERGRID', (0, 0), (-1, -1), 0.3, BORDER_BL),
                    ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
                    ('TOPPADDING', (0, 0), (-1, -1), 5), ('BOTTOMPADDING', (0, 0), (-1, -1), 5),
                    ('LEFTPADDING', (0, 0), (-1, -1), 8), ('RIGHTPADDING', (0, 0), (-1, -1), 8),
                ]))
                story.append(ct)

            # FOOTER
            story.append(Spacer(1, 6 * mm))
            story.append(Paragraph(f"Rapport trait\u00e9 par : {treated_by or 'Admin'}, le {now}", ft_s))

    doc.build(story)
