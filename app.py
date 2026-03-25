#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
RAMYA TECHNOLOGIE & INNOVATION
Application Web v3 — Gestion des Rapports de Pointage
Auth + Rôles + Dashboard + Clients + Fichiers RH
"""

import os, uuid, shutil, functools, smtplib, json
from datetime import datetime
from email.mime.multipart import MIMEMultipart
from email.mime.text import MIMEText
from email.mime.base import MIMEBase
from email import encoders

BASE_DIR = os.path.dirname(os.path.abspath(__file__))

from flask import (Flask, render_template, request, send_file, flash,
                   redirect, url_for, jsonify, session, send_from_directory)
from werkzeug.utils import secure_filename

from rapport_core import extract_from_excel, generate_full_pdf
from merge_presence import generate_presence_xlsx
from models import (init_db, create_user, authenticate_user, get_user_by_id,
                    get_all_users, update_user, delete_user,
                    create_client, get_all_clients, get_client_by_id,
                    find_client_by_name, update_client, delete_client,
                    create_job, get_jobs_by_status, get_all_jobs, mark_job_sent,
                    get_dashboard_stats, has_permission, get_role_permissions,
                    update_role_permissions,
                    reset_jobs, reset_clients, reset_users, reset_all,
                    log_activity, get_activity_logs,
                    add_job_comment, get_job_comments, update_job_notes,
                    get_job_by_id, get_db_path,
                    create_contract, get_client_contracts, get_all_contracts,
                    get_contract_by_id, update_contract, delete_contract,
                    get_client_monthly_stats,
                    save_smtp_settings, get_smtp_settings,
                    create_invoice, get_invoices_by_status, get_all_invoices,
                    update_invoice_status, get_invoice_stats,
                    create_visit_report, get_visit_reports, get_visit_by_id,
                    update_visit_proforma, get_visit_stats,
                    create_devis, get_all_devis, get_devis_by_id,
                    update_devis_status, get_devis_stats)
from devis_generator import generate_devis_pdf

app = Flask(__name__, template_folder=BASE_DIR, static_folder=BASE_DIR, static_url_path='/static')
app.secret_key = os.environ.get('SECRET_KEY', 'ramya-tech-2026-secret-v3')
app.config['MAX_CONTENT_LENGTH'] = 16 * 1024 * 1024
PERSISTENT_DIR = os.environ.get('PERSISTENT_DIR', BASE_DIR)
app.config['UPLOAD_FOLDER'] = os.path.join(PERSISTENT_DIR, 'uploads')
app.config['FILES_FOLDER'] = os.path.join(PERSISTENT_DIR, 'files')

os.makedirs(app.config['UPLOAD_FOLDER'], exist_ok=True)
os.makedirs(app.config['FILES_FOLDER'], exist_ok=True)

init_db()

# Init extra tables
from models import (init_rh_tables, get_all_employees, get_employee_by_id,
                    create_employee, update_employee, get_employee_stats,
                    get_leaves, create_leave, update_leave_status,
                    get_payslips, create_payslip, update_payslip,
                    init_extra_tables, record_login_attempt, get_failed_attempts,
                    save_otp, verify_otp, db_count, db_sum, db_get_all)
init_rh_tables()
init_extra_tables()

from models import init_mg_tables
init_mg_tables()

from models import (init_chat_tables, get_messages, get_direct_messages, 
                    send_message, get_unread_count, mark_chat_read, db_get_all,
                    migrate_v4, get_payslip_detail, get_maintenance_due,
                    migrate_v5, log_audit, get_audit_trail, get_executive_stats,
                    get_devis_templates, get_devis_template,
                    migrate_payslip_v2, get_payslip_detail_v2,
                    migrate_caisse, gen_caisse_ref, get_caisse_sorties, get_caisse_stats)
from models import migrate_caisse_v2, delete_caisse
from models import migrate_v6
init_chat_tables()
migrate_v4()
migrate_payslip_v2()
migrate_v5()
migrate_caisse()
migrate_caisse_v2()
migrate_v6()
from models import migrate_v7
migrate_v7()
from models import migrate_v8
migrate_v8()
from models import migrate_v9
migrate_v9()
from models import migrate_v10
migrate_v10()
from models import migrate_v11
migrate_v11()
from models import migrate_v12
migrate_v12()
from models import migrate_v13
migrate_v13()
from models import migrate_v14
migrate_v14()
from models import migrate_v15
migrate_v15()
from models import migrate_v16
migrate_v16()
from models import migrate_v17
migrate_v17()
from models import migrate_v18
migrate_v18()
from models import migrate_v19
migrate_v19()
from models import migrate_v20
migrate_v20()
from models import migrate_v15
migrate_v15()
from models import migrate_v16
migrate_v16()
from models import migrate_v17
migrate_v17()
from models import migrate_v18
migrate_v18()
from models import migrate_v19
migrate_v19()
from models import migrate_v20
migrate_v20()

# Register module routes
from modules_routes import modules_bp
app.register_blueprint(modules_bp)

from models import (init_devis_tables, create_devis, get_all_devis, get_devis_by_id,
                    update_devis_status, get_devis_stats, get_next_devis_ref)
init_devis_tables()

ALLOWED_EXTENSIONS = {'xlsx', 'xls'}
ALL_PERMISSIONS = ['traitement', 'fichiers', 'clients', 'clients_edit', 'admin', 'dashboard', 'dashboard_general',
                   'envoyer', 'logs', 
                   'contrats', 'comptabilite', 'comptabilite_edit', 'visites', 'visites_edit', 'proforma', 'proforma_edit',
                   'moyens_generaux', 'moyens_generaux_edit', 'informatique', 'projets', 'caisse_sortie', 'rapports_j', 'convertir_devis',
                   'resp_projet', 'resp_projet_edit', 'centre_technique', 'centre_technique_edit', 'chat', 'tracking']

# Permission categories for admin display
PERM_CATEGORIES = {
    'Comptabilité': [('comptabilite', 'Lecture'), ('comptabilite_edit', 'Modification'), ('convertir_devis', 'Convertir devis'), ('caisse_sortie', 'Caisse')],
    'Commercial / CRM': [('clients', 'Clients lecture'), ('clients_edit', 'Clients modification'), ('proforma', 'Devis lecture'), ('proforma_edit', 'Devis modification'), ('visites', 'Visites lecture'), ('visites_edit', 'Visites modification')],
    'Technique': [('centre_technique', 'Centre technique'), ('centre_technique_edit', 'Centre tech. modif'), ('traitement', 'Traitement/DPCI')],
    'Projets': [('resp_projet', 'Resp. projet lecture'), ('resp_projet_edit', 'Resp. projet modif'), ('projets', 'Projets info')],
    'RH & Admin': [('fichiers', 'Employés/RH'), ('contrats', 'Contrats'), ('envoyer', 'Envoi paie'), ('logs', 'Logs')],
    'Général': [('dashboard', 'Dashboard'), ('dashboard_general', 'Dashboard général'), ('rapports_j', 'Rapports journaliers'), ('chat', 'Chat'), ('informatique', 'Informatique'), ('moyens_generaux', 'Stock lecture'), ('moyens_generaux_edit', 'Stock modification')],
    'Administration': [('admin', 'Admin système')],
}

def allowed_file(fn):
    return '.' in fn and fn.rsplit('.', 1)[1].lower() in ALLOWED_EXTENSIONS


# ======================== AUTH HELPERS ========================

def login_required(f):
    @functools.wraps(f)
    def decorated(*args, **kwargs):
        if 'user_id' not in session:
            flash("Veuillez vous connecter", "error")
            return redirect(url_for('login'))
        return f(*args, **kwargs)
    return decorated

def permission_required(perm):
    def decorator(f):
        @functools.wraps(f)
        def decorated(*args, **kwargs):
            if 'user_id' not in session:
                return redirect(url_for('login'))
            user = get_user_by_id(session['user_id'])
            if not user or not has_permission(user['role'], perm):
                flash("Accès non autorisé", "error")
                return redirect(url_for('dashboard'))
            return f(*args, **kwargs)
        return decorated
    return decorator

@app.context_processor
def inject_globals():
    """Injecte les variables globales dans tous les templates."""
    ctx = {'current_user': None, 'permissions': [], 'pending_count': 0, 'unread_messages': 0, 'caisse_pending': 0, 'weekly_champion': None}
    if 'user_id' in session:
        user = get_user_by_id(session['user_id'])
        if user:
            perms = get_role_permissions(user['role'])
            ctx['current_user'] = user
            ctx['permissions'] = perms
            ctx['can_edit'] = lambda module: (module + '_edit') in perms or 'admin' in perms
            ctx['pending_count'] = len(get_jobs_by_status('traite'))
            try: ctx['unread_messages'] = get_unread_count(user['id'])
            except: pass
            if user['role'] in ('admin', 'dg', 'directeur'):
                try:
                    from models import get_db as _gdb
                    _c = _gdb()
                    ctx['caisse_pending'] = _c.execute("SELECT COUNT(*) FROM caisse_sorties WHERE status='en_attente'").fetchone()[0]
                    _c.close()
                except: pass
            # Weekly champion
            try:
                from models import get_current_champion, update_weekly_champion, get_live_champion
                update_weekly_champion()
                champ = get_current_champion()
                live = get_live_champion()
                # Prefer stored champion, fall back to live leader
                if champ:
                    ctx['weekly_champion'] = champ
                elif live:
                    ctx['weekly_champion'] = live
            except: pass
    else:
        ctx['can_edit'] = lambda module: False
    return ctx


# ======================== AUTH ROUTES ========================

@app.route('/robots.txt')
def robots_txt():
    return "User-agent: *\nAllow: /\n", 200, {'Content-Type': 'text/plain'}

@app.route('/favicon.ico')
def favicon():
    return send_from_directory(app.config.get('STATIC_FOLDER', 'static'), 'logo_wannygest.png', mimetype='image/png')

@app.errorhandler(500)
def internal_error(e):
    import traceback
    traceback.print_exc()
    return f"<h1>Erreur serveur</h1><p>{str(e)}</p><a href='/'>Retour</a>", 500

@app.errorhandler(404)
def not_found(e):
    return f"<h1>Page non trouvée</h1><a href='/'>Retour à l'accueil</a>", 404

@app.route('/')
def welcome():
    if 'user_id' in session:
        return redirect(url_for('dashboard'))
    return render_template('welcome.html')

@app.before_request
def check_session_timeout():
    """Déconnexion automatique après 30 min d'inactivité."""
    if 'user_id' in session:
        last = session.get('last_active')
        if last:
            from datetime import datetime, timedelta
            try:
                last_dt = datetime.fromisoformat(last)
                if datetime.now() - last_dt > timedelta(minutes=30):
                    session.clear()
                    flash("Session expirée — veuillez vous reconnecter", "info")
                    return redirect(url_for('login'))
            except:
                pass
        session['last_active'] = datetime.now().isoformat()

@app.route('/login', methods=['GET', 'POST'])
def login():
    if request.method == 'POST':
        username = request.form['username']
        ip = request.remote_addr
        
        # Vérifier verrouillage (5 tentatives en 15 min)
        if get_failed_attempts(username) >= 5:
            flash("Compte temporairement verrouillé (trop de tentatives). Réessayez dans 15 minutes.", "error")
            return render_template('login.html')
        
        user = authenticate_user(username, request.form['password'])
        if user:
            # Vérifier politique mot de passe
            record_login_attempt(username, True, ip)
            session['user_id'] = user['id']
            session['last_active'] = datetime.now().isoformat()
            log_activity(user['id'], user['full_name'], 'Connexion', 
                        f"Connexion réussie", ip)
            flash(f"Bienvenue {user['full_name']} !", "success")
            return redirect(url_for('dashboard'))
        
        record_login_attempt(username, False, ip)
        remaining = 5 - get_failed_attempts(username)
        flash(f"Identifiants incorrects ({remaining} tentative(s) restante(s))", "error")
    return render_template('login.html')

@app.route('/register', methods=['GET', 'POST'])
@login_required
def register():
    # Only admin can create users
    u = get_user_by_id(session['user_id'])
    if u['role'] not in ('admin', 'dg'):
        flash("Seul l'administrateur peut créer des comptes", "error")
        return redirect('/dashboard')
    if request.method == 'POST':
        pwd = request.form['password']
        pwd2 = request.form['password2']
        if pwd != pwd2:
            flash("Les mots de passe ne correspondent pas", "error")
            return render_template('register.html')
        if len(pwd) < 6:
            flash("Le mot de passe doit faire au moins 6 caractères", "error")
            return render_template('register.html')
        role = request.form.get('role', 'technicien')
        ok, msg = create_user(
            request.form['username'], request.form['email'],
            pwd, request.form['full_name'], role
        )
        if ok:
            flash(f"Compte '{request.form['username']}' créé avec le rôle {role} !", "success")
            return redirect('/admin')
        flash(msg, "error")
    return render_template('register.html')

@app.route('/logout')
def logout():
    session.clear()
    flash("Déconnexion réussie", "success")
    return redirect(url_for('welcome'))


# ======================== DASHBOARD ========================

@app.route('/dashboard')
@login_required
def dashboard():
    # Show guide on first visit (only redirect once)
    if request.args.get('skip_guide'):
        session['guide_seen'] = True
    if not session.get('guide_seen'):
        session['guide_seen'] = True
        return redirect(url_for('guide'))
    user = get_user_by_id(session['user_id'])
    role = user['role'] if user else 'technicien'
    stats = get_dashboard_stats()
    inv_stats = get_invoice_stats()
    v_stats = get_visit_stats()
    d_stats = get_devis_stats()
    emp_stats = get_employee_stats()
    # Notifications for dashboard
    announcements = db_get_all('rh_announcements', order='created_at DESC', limit=5) if 'rh_announcements' in _get_tables() else []
    trainings_upcoming = db_get_all('rh_trainings', order='date ASC', limit=5) if 'rh_trainings' in _get_tables() else []
    return render_template('dashboard.html', page='dashboard', stats=stats, 
                          inv_stats=inv_stats, v_stats=v_stats, d_stats=d_stats,
                          emp_stats=emp_stats, user_role=role,
                          announcements=announcements, trainings_upcoming=trainings_upcoming)

@app.route('/dashboard-general')
@permission_required('dashboard_general')
def dashboard_general():
    """Tableau de bord général — vue consolidée de toutes les activités."""
    conn = _gdb()
    data = {}
    
    # CRM
    data['clients'] = conn.execute("SELECT COUNT(*) FROM clients").fetchone()[0]
    data['prospects'] = conn.execute("SELECT COUNT(*) FROM prospects").fetchone()[0] if 'prospects' in _get_tables() else 0
    data['devis'] = conn.execute("SELECT COUNT(*) FROM devis").fetchone()[0]
    data['devis_acceptes'] = conn.execute("SELECT COUNT(*) FROM devis WHERE status='accepte'").fetchone()[0]
    
    # Comptabilité
    data['invoices'] = conn.execute("SELECT COUNT(*) FROM invoices").fetchone()[0] if 'invoices' in _get_tables() else 0
    data['revenue'] = conn.execute("SELECT COALESCE(SUM(amount),0) FROM invoices WHERE status='payee'").fetchone()[0] if 'invoices' in _get_tables() else 0
    data['expenses'] = conn.execute("SELECT COALESCE(SUM(amount),0) FROM pieces_caisse").fetchone()[0] if 'pieces_caisse' in _get_tables() else 0
    data['profit'] = data['revenue'] - data['expenses']
    
    # RH
    data['employees'] = conn.execute("SELECT COUNT(*) FROM employees").fetchone()[0] if 'employees' in _get_tables() else 0
    data['formations'] = conn.execute("SELECT COUNT(*) FROM rh_trainings").fetchone()[0] if 'rh_trainings' in _get_tables() else 0
    data['annonces'] = conn.execute("SELECT COUNT(*) FROM rh_announcements").fetchone()[0] if 'rh_announcements' in _get_tables() else 0
    
    # Stock & Achats
    data['stock_items'] = conn.execute("SELECT COUNT(*) FROM stock_items").fetchone()[0] if 'stock_items' in _get_tables() else 0
    data['stock_value'] = conn.execute("SELECT COALESCE(SUM(quantity*unit_price),0) FROM stock_items").fetchone()[0] if 'stock_items' in _get_tables() else 0
    data['stock_low'] = conn.execute("SELECT COUNT(*) FROM stock_items WHERE quantity<=min_stock").fetchone()[0] if 'stock_items' in _get_tables() else 0
    data['fournisseurs'] = conn.execute("SELECT COUNT(*) FROM achats_fournisseurs").fetchone()[0] if 'achats_fournisseurs' in _get_tables() else 0
    data['commandes'] = conn.execute("SELECT COUNT(*) FROM achats_commandes").fetchone()[0] if 'achats_commandes' in _get_tables() else 0
    
    # Gestion du temps
    data['rapports'] = conn.execute("SELECT COUNT(*) FROM jobs").fetchone()[0] if 'jobs' in _get_tables() else 0
    data['rapports_j'] = conn.execute("SELECT COUNT(*) FROM rapports_journaliers").fetchone()[0] if 'rapports_journaliers' in _get_tables() else 0
    
    # Informatique
    data['projets'] = conn.execute("SELECT COUNT(*) FROM projects").fetchone()[0] if 'projects' in _get_tables() else 0
    data['tickets'] = conn.execute("SELECT COUNT(*) FROM tickets").fetchone()[0] if 'tickets' in _get_tables() else 0
    
    # Recent activity
    data['recent_logs'] = [dict(r) for r in conn.execute("""SELECT * FROM activity_log ORDER BY timestamp DESC LIMIT 15""").fetchall()] if 'activity_log' in _get_tables() else []
    
    # Monthly revenue chart
    data['monthly_rev'] = [dict(r) for r in conn.execute("""SELECT strftime('%Y-%m', created_at) as month, SUM(amount) as total 
        FROM invoices WHERE status='payee' GROUP BY month ORDER BY month DESC LIMIT 6""").fetchall()] if 'invoices' in _get_tables() else []
    data['monthly_rev'].reverse()
    
    # Invoice stats
    data['inv_stats'] = {}
    if 'invoices' in _get_tables():
        for s in ('a_envoyer','envoyee','en_attente_paiement','payee'):
            data['inv_stats'][s] = conn.execute(f"SELECT COUNT(*) FROM invoices WHERE status='{s}'").fetchone()[0]
        data['pending_amount'] = conn.execute("SELECT COALESCE(SUM(amount),0) FROM invoices WHERE status IN ('envoyee','en_attente_paiement')").fetchone()[0]
    
    # Devis stats
    data['devis_stats'] = {}
    for s in ('brouillon','envoye','accepte','refuse'):
        data['devis_stats'][s] = conn.execute(f"SELECT COUNT(*) FROM devis WHERE status='{s}'").fetchone()[0]
    
    # Prospect stats
    data['prospect_stats'] = {}
    if 'prospects' in _get_tables():
        for s in ('nouveau','contacte','qualifie','proposition','gagne','perdu'):
            data['prospect_stats'][s] = conn.execute(f"SELECT COUNT(*) FROM prospects WHERE status='{s}'").fetchone()[0]
    
    conn.close()
    return render_template('dashboard_general.html', page='dashboard_general', data=data)

def _get_tables():
    from models import get_db
    conn = get_db()
    tables = [r['name'] for r in conn.execute("SELECT name FROM sqlite_master WHERE type='table'").fetchall()]
    conn.close()
    return tables

@app.route('/guide')
@login_required
def guide():
    return render_template('onboarding.html', page='guide')


# ======================== TRAITEMENT ========================

@app.route('/traitement')
@permission_required('traitement')
def traitement():
    clients = get_all_clients()
    return render_template('traitement.html', page='traitement', clients=clients)

@app.route('/traitement/preview', methods=['POST'])
@login_required
def traitement_preview():
    if 'excel_file' not in request.files:
        return jsonify({"error": "Aucun fichier"}), 400
    file = request.files['excel_file']
    if not allowed_file(file.filename):
        return jsonify({"error": "Format non supporté"}), 400
    job_dir = os.path.join(app.config['UPLOAD_FOLDER'], 'preview')
    os.makedirs(job_dir, exist_ok=True)
    try:
        path = os.path.join(job_dir, secure_filename(file.filename))
        file.save(path)
        emps, client = extract_from_excel(path)
        from models import save_known_employees
        save_known_employees([e['name'] for e in emps])
        return jsonify({"client": client, "count": len(emps),
            "employees": [{"name": e['name'], "ref": e['ref'], "days": len(e['records'])} for e in emps]})
    except Exception as e:
        return jsonify({"error": str(e)}), 500
    finally:
        shutil.rmtree(job_dir, ignore_errors=True)

@app.route('/traitement/merge', methods=['POST'])
@login_required
def traitement_merge():
    if 'enr_file' not in request.files or 'trans_file' not in request.files:
        return jsonify({"error": "Les 2 fichiers sont requis"}), 400
    enr = request.files['enr_file']
    trans = request.files['trans_file']
    if not enr.filename or not trans.filename:
        return jsonify({"error": "Les 2 fichiers sont requis"}), 400
    merge_dir = os.path.join(app.config['UPLOAD_FOLDER'], 'merge_' + str(uuid.uuid4())[:8])
    os.makedirs(merge_dir, exist_ok=True)
    try:
        enr_path = os.path.join(merge_dir, secure_filename(enr.filename))
        trans_path = os.path.join(merge_dir, secure_filename(trans.filename))
        enr.save(enr_path); trans.save(trans_path)
        out = os.path.join(merge_dir, 'Presence_fusionnee.xlsx')
        result = generate_presence_xlsx(enr_path, trans_path, out)
        if not result:
            return jsonify({"error": "Échec de la fusion."}), 400
        merge_id = os.path.basename(merge_dir)
        # Save employee names for schedule
        if result.get('employees'):
            from models import save_known_employees
            save_known_employees(result['employees'], result.get('emp_services', {}))
        return jsonify({"success": True, "merge_id": merge_id, "client": result['client'],
            "employees": result['employees'], "emp_services": result.get('emp_services', {}),
            "rows": result['rows'], "filename": 'Presence_fusionnee.xlsx'})
    except Exception as e:
        return jsonify({"error": str(e)}), 500

@app.route('/traitement/merge/download/<merge_id>')
@login_required
def traitement_merge_download(merge_id):
    safe_id = secure_filename(merge_id)
    path = os.path.join(app.config['UPLOAD_FOLDER'], safe_id, 'Presence_fusionnee.xlsx')
    if not os.path.exists(path):
        flash("Fichier expiré", "error")
        return redirect(url_for('traitement'))
    return send_file(path, as_attachment=True, download_name='Presence_fusionnee.xlsx')

@app.route('/traitement/generate', methods=['POST'])
@permission_required('traitement')
def traitement_generate():
    merge_id = request.form.get('merge_id', '').strip()
    
    if merge_id:
        merge_dir = os.path.join(app.config['UPLOAD_FOLDER'], merge_id)
        merged_file = os.path.join(merge_dir, 'Presence_fusionnee.xlsx')
        if not os.path.exists(merged_file):
            flash("Fichier fusionné expiré. Refaites la fusion.", "error")
            return redirect(url_for('traitement'))
        xlsx_source = merged_file
        filename = 'Presence_fusionnee.xlsx'
    else:
        if 'excel_file' not in request.files or not request.files['excel_file'].filename:
            flash("Aucun fichier sélectionné", "error")
            return redirect(url_for('traitement'))
        file = request.files['excel_file']
        if not allowed_file(file.filename):
            flash("Format non supporté", "error")
            return redirect(url_for('traitement'))
        xlsx_source = None
        filename = secure_filename(file.filename)
    
    provider_name = request.form.get('provider_name', '').strip() or "RAMYA TECHNOLOGIE & INNOVATION"
    provider_info = request.form.get('provider_info', '').strip() or "Tél: 2722204498 | Email: techniqueramya@gmail.com"
    client_name = request.form.get('client_name', '').strip()
    client_tel = request.form.get('client_tel', '').strip()
    client_email = request.form.get('client_email', '').strip()
    client_id_str = request.form.get('client_id', '').strip()
    hp_str = request.form.get('required_hours', '0').strip()
    hp_we_str = request.form.get('required_hours_weekend', '0').strip()
    hourly_cost_str = request.form.get('hourly_cost', '0').strip()
    employee_costs_json = request.form.get('employee_costs_json', '{}')

    
    # Auto-fill from client database
    client_id = int(client_id_str) if client_id_str else None
    if client_id:
        db_client = get_client_by_id(client_id)
        if db_client:
            if not client_name: client_name = db_client['name']
            if not client_tel: client_tel = db_client.get('tel', '')
            if not client_email: client_email = db_client.get('email', '')
    
    client_info_parts = []
    if client_tel: client_info_parts.append(f"Tél: {client_tel}")
    if client_email: client_info_parts.append(f"Email: {client_email}")
    client_info = " | ".join(client_info_parts)
    
    try:
        hp = float(hp_str) if hp_str else 0
    except:
        hp = 0
    try:
        hp_weekend = float(hp_we_str) if hp_we_str else 0
    except:
        hp_weekend = 0
    try:
        hourly_cost = float(hourly_cost_str) if hourly_cost_str else 0
    except:
        hourly_cost = 0
    try:
        employee_costs = json.loads(employee_costs_json) if employee_costs_json else {}
    except:
        employee_costs = {}
    
    job_id = str(uuid.uuid4())[:8]
    job_dir = os.path.join(app.config['UPLOAD_FOLDER'], job_id)
    os.makedirs(job_dir, exist_ok=True)
    
    try:
        if xlsx_source:
            xlsx_path = os.path.join(job_dir, filename)
            shutil.copy2(xlsx_source, xlsx_path)
        else:
            file = request.files['excel_file']
            xlsx_path = os.path.join(job_dir, filename)
            file.save(xlsx_path)
        
        # Logo
        logo_path = None
        if 'logo_file' in request.files:
            lf = request.files['logo_file']
            if lf.filename:
                logo_path = os.path.join(job_dir, 'custom_logo.png')
                lf.save(logo_path)
        if not logo_path:
            for n in ['logo_wannygest.png', 'logo.png']:
                c = os.path.join(BASE_DIR, n)
                if os.path.exists(c):
                    logo_path = os.path.join(job_dir, n)
                    shutil.copy2(c, logo_path)
                    break
        
        emps, detected_client = extract_from_excel(xlsx_path)
        from models import save_known_employees
        save_known_employees([e['name'] for e in emps])
        if not emps:
            flash("Aucun employé trouvé", "error")
            return redirect(url_for('traitement'))
        
        if not client_name:
            client_name = detected_client
            # Try to find client in database
            db_client = find_client_by_name(detected_client)
            if db_client:
                client_id = db_client['id']
                if not client_tel: client_tel = db_client.get('tel', '')
                if not client_email: client_email = db_client.get('email', '')
                client_info_parts = []
                if client_tel: client_info_parts.append(f"Tél: {client_tel}")
                if client_email: client_info_parts.append(f"Email: {client_email}")
                client_info = " | ".join(client_info_parts)
        
        all_dates = [rec['date'] for emp in emps for rec in emp['records']]
        all_dates.sort()
        period = f"Période du {all_dates[0]} au {all_dates[-1]}" if all_dates else "Rapport"
        
        base = os.path.splitext(filename)[0]
        pdf_name = f"{base}_RAPPORT_DE_PRESENCE.pdf"
        output_path = os.path.join(job_dir, pdf_name)
        
        generate_full_pdf(emps, output_path, provider_name, provider_info,
                         client_name, period, logo_path, hp=hp, client_info=client_info,
                         work_dir=job_dir, hp_weekend=hp_weekend, hourly_cost=hourly_cost,
                         employee_costs=employee_costs)
        
        if not os.path.exists(output_path):
            flash("Erreur génération PDF", "error")
            return redirect(url_for('traitement'))
        
        # Sauvegarder dans le dossier files
        files_dir = os.path.join(app.config['FILES_FOLDER'], job_id)
        os.makedirs(files_dir, exist_ok=True)
        shutil.copy2(output_path, os.path.join(files_dir, pdf_name))
        
        # Copier le xlsx fusionné si dispo
        xlsx_out = None
        if merge_id:
            merged_xlsx = os.path.join(app.config['UPLOAD_FOLDER'], merge_id, 'Presence_fusionnee.xlsx')
            if os.path.exists(merged_xlsx):
                xlsx_out = 'Presence_fusionnee.xlsx'
                shutil.copy2(merged_xlsx, os.path.join(files_dir, xlsx_out))
        
        # Enregistrer dans la BDD
        hp_text = f"{hp}h/sem" if hp > 0 else "Auto"
        if hp_weekend > 0: hp_text += f" | {hp_weekend}h/we"
        create_job(job_id, session['user_id'], client_name, provider_name,
                   filename, pdf_name, xlsx_out, len(emps), period, hp_text, client_id)
        
        user = get_user_by_id(session['user_id'])
        log_activity(session['user_id'], user['full_name'] if user else '?',
                    'Traitement', f"Rapport PDF généré — {client_name}, {len(emps)} employés",
                    request.remote_addr)
        flash(f"Rapport généré avec succès — {len(emps)} employés, {client_name}", "success")
        return send_file(output_path, as_attachment=True, download_name=pdf_name, mimetype='application/pdf')
    
    except Exception as e:
        flash(f"Erreur : {str(e)}", "error")
        return redirect(url_for('traitement'))
    finally:
        _cleanup_old(app.config['UPLOAD_FOLDER'])


# ======================== FICHIERS RH ========================

@app.route('/fichiers')
@permission_required('fichiers')
def fichiers():
    tab = request.args.get('tab', 'pending')
    pending = get_jobs_by_status('traite')
    sent = get_jobs_by_status('envoye')
    return render_template('fichiers.html', page='fichiers', tab=tab, pending=pending, sent=sent)

@app.route('/fichiers/download/<job_id>/<ftype>')
@login_required
def fichiers_download(job_id, ftype):
    safe_id = secure_filename(job_id)
    files_dir = os.path.join(app.config['FILES_FOLDER'], safe_id)
    if not os.path.isdir(files_dir):
        flash("Fichier non trouvé", "error")
        return redirect(url_for('fichiers'))
    for f in os.listdir(files_dir):
        if ftype == 'pdf' and f.endswith('.pdf'):
            return send_from_directory(files_dir, f, as_attachment=True)
        if ftype == 'xlsx' and f.endswith('.xlsx'):
            return send_from_directory(files_dir, f, as_attachment=True)
    flash("Fichier non trouvé", "error")
    return redirect(url_for('fichiers'))

@app.route('/fichiers/preview/<job_id>')
@login_required
def fichiers_preview(job_id):
    """Prévisualisation PDF dans le navigateur avant envoi."""
    safe_id = secure_filename(job_id)
    files_dir = os.path.join(app.config['FILES_FOLDER'], safe_id)
    if not os.path.isdir(files_dir):
        flash("Fichier non trouvé", "error")
        return redirect(url_for('fichiers'))
    for f in os.listdir(files_dir):
        if f.endswith('.pdf'):
            return send_from_directory(files_dir, f, as_attachment=False, mimetype='application/pdf')
    flash("PDF non trouvé", "error")
    return redirect(url_for('fichiers'))

@app.route('/fichiers/marquer/<job_id>')
@permission_required('envoyer')
def fichiers_marquer(job_id):
    mark_job_sent(job_id, session['user_id'])
    job = get_job_by_id(job_id)
    if job:
        create_invoice(job_id, job.get('client_id'), job.get('client_name', ''))
    user = get_user_by_id(session['user_id'])
    log_activity(session['user_id'], user['full_name'] if user else '?',
                'Envoi', f"Rapport {job_id} marqué comme envoyé", request.remote_addr)
    flash("Fichier envoyé — facture créée pour la comptabilité", "success")
    return redirect(url_for('fichiers'))


# ======================== CLIENTS ========================

@app.route('/clients')
@permission_required('clients')
def clients_page():
    clients = get_all_clients()
    return render_template('clients.html', page='clients', clients=clients)

@app.route('/clients/add', methods=['POST'])
@permission_required('clients_edit')
def clients_add():
    create_client(
        request.form['name'], request.form.get('tel', ''),
        request.form.get('email', ''), request.form.get('contact_name', ''),
        request.form.get('address', ''), request.form.get('notes', ''),
        session['user_id']
    )
    # Update enriched fields
    from models import get_db as _gdb2
    conn = _gdb2()
    cid = conn.execute("SELECT id FROM clients ORDER BY id DESC LIMIT 1").fetchone()['id']
    for field in ['sector','city','country','website','rc_number','cnps_number',
                  'contact_title','contact_tel2','contact_email2','payment_terms',
                  'source','status']:
        val = request.form.get(field, '')
        if val:
            try: conn.execute(f"UPDATE clients SET {field}=? WHERE id=?", (val, cid))
            except: pass
    if request.form.get('credit_limit'):
        conn.execute("UPDATE clients SET credit_limit=? WHERE id=?", (float(request.form['credit_limit']), cid))
    conn.commit(); conn.close()
    user = get_user_by_id(session['user_id'])
    log_audit(session['user_id'], user['full_name'] if user else '?', 'clients', 0, 'create', 'name', '', request.form['name'])
    flash("Client ajouté", "success")
    return redirect(url_for('clients_page'))

@app.route('/clients/edit/<int:cid>', methods=['GET', 'POST'])
@permission_required('clients_edit')
def clients_edit(cid):
    client = get_client_by_id(cid)
    if not client:
        flash("Client non trouvé", "error")
        return redirect(url_for('clients_page'))
    if request.method == 'POST':
        update_client(cid, name=request.form['name'], tel=request.form.get('tel', ''),
                      email=request.form.get('email', ''), contact_name=request.form.get('contact_name', ''),
                      address=request.form.get('address', ''), notes=request.form.get('notes', ''))
        # Update enriched fields
        from models import get_db as _gdb3
        conn = _gdb3()
        for field in ['sector','city','country','website','rc_number','cnps_number',
                      'contact_title','contact_tel2','contact_email2','payment_terms',
                      'source','status']:
            val = request.form.get(field, '')
            try: conn.execute(f"UPDATE clients SET {field}=? WHERE id=?", (val, cid))
            except: pass
        if request.form.get('credit_limit'):
            try: conn.execute("UPDATE clients SET credit_limit=? WHERE id=?", (float(request.form.get('credit_limit',0) or 0), cid))
            except: pass
        conn.commit(); conn.close()
        user = get_user_by_id(session['user_id'])
        log_audit(session['user_id'], user['full_name'] if user else '?', 'clients', cid, 'update', 'name', client['name'], request.form['name'])
        flash("Client modifié", "success")
        return redirect(url_for('clients_page'))
    return render_template('edit_client.html', page='clients', client=client)

@app.route('/clients/delete/<int:cid>')
@permission_required('clients_edit')
def clients_delete(cid):
    delete_client(cid)
    flash("Client supprimé", "success")
    return redirect(url_for('clients_page'))


# ======================== ADMIN ========================

@app.route('/admin')
@permission_required('admin')
def admin_page():
    users = get_all_users()
    stats = get_dashboard_stats()
    role_perms = {r: get_role_permissions(r) for r in ['admin', 'dg', 'rh', 'technicien', 'commercial', 'comptable', 'moyens_generaux', 'informatique']}
    return render_template('admin.html', page='admin', users=users, stats=stats,
                          all_permissions=ALL_PERMISSIONS, role_perms=role_perms, perm_categories=PERM_CATEGORIES)

@app.route('/admin/add', methods=['POST'])
@permission_required('admin')
def admin_add_user():
    ok, msg = create_user(
        request.form['username'], request.form['email'],
        request.form['password'], request.form['full_name'],
        request.form.get('role', 'technicien')
    )
    flash(msg, "success" if ok else "error")
    return redirect(url_for('admin_page'))

@app.route('/admin/reset', methods=['POST'])
@permission_required('admin')
def admin_reset():
    target = request.form.get('target', '')
    user = get_user_by_id(session['user_id'])
    uname = user['full_name'] if user else '?'
    
    if target == 'jobs':
        reset_jobs()
        files_dir = app.config['FILES_FOLDER']
        if os.path.exists(files_dir):
            shutil.rmtree(files_dir)
            os.makedirs(files_dir, exist_ok=True)
        log_activity(session['user_id'], uname, 'Réinitialisation', 'Tous les rapports supprimés', request.remote_addr)
        flash("Tous les rapports ont été supprimés", "success")
    
    elif target == 'clients':
        reset_clients()
        log_activity(session['user_id'], uname, 'Réinitialisation', 'Tous les clients supprimés', request.remote_addr)
        flash("Tous les clients ont été supprimés", "success")
    
    elif target == 'users':
        reset_users()
        log_activity(session['user_id'], uname, 'Réinitialisation', 'Utilisateurs non-admin supprimés', request.remote_addr)
        flash("Tous les utilisateurs (sauf admin) ont été supprimés", "success")
    
    elif target == 'all':
        reset_all()
        for folder in [app.config['FILES_FOLDER'], app.config['UPLOAD_FOLDER']]:
            if os.path.exists(folder):
                shutil.rmtree(folder)
                os.makedirs(folder, exist_ok=True)
        log_activity(session['user_id'], uname, 'Réinitialisation', 'RÉINITIALISATION COMPLÈTE', request.remote_addr)
        flash("Réinitialisation complète effectuée", "success")
    
    return redirect(url_for('admin_page'))

@app.route('/admin/edit/<int:uid>', methods=['GET', 'POST'])
@permission_required('admin')
def admin_edit_user(uid):
    u = get_user_by_id(uid)
    if not u:
        flash("Utilisateur non trouvé", "error")
        return redirect(url_for('admin_page'))
    if request.method == 'POST':
        updates = {'full_name': request.form['full_name'],
                   'email': request.form['email'],
                   'role': request.form['role']}
        pwd = request.form.get('password', '').strip()
        if pwd:
            updates['password'] = pwd
        update_user(uid, **updates)
        flash("Utilisateur modifié", "success")
        return redirect(url_for('admin_page'))
    return render_template('edit_user.html', page='admin', edit_user=u)

@app.route('/admin/toggle/<int:uid>')
@permission_required('admin')
def admin_toggle_user(uid):
    u = get_user_by_id(uid)
    if u and u['role'] != 'admin':
        update_user(uid, is_active=0 if u['is_active'] else 1)
        flash(f"Utilisateur {'désactivé' if u['is_active'] else 'activé'}", "success")
    return redirect(url_for('admin_page'))

@app.route('/admin/delete/<int:uid>')
@permission_required('admin')
def admin_delete_user(uid):
    u = get_user_by_id(uid)
    if not u:
        flash("Utilisateur non trouvé", "error")
    elif u['role'] == 'admin' and u['username'] == 'admin':
        flash("Impossible de supprimer le compte admin principal", "error")
    elif uid == session.get('user_id'):
        flash("Impossible de supprimer votre propre compte", "error")
    else:
        delete_user(uid)
        user = get_user_by_id(session['user_id'])
        log_activity(session['user_id'], user['full_name'] if user else '?',
                    'Admin', f"Utilisateur {u['full_name']} ({u['username']}) supprimé", request.remote_addr)
        flash(f"Compte '{u['username']}' supprimé définitivement", "success")
    return redirect(url_for('admin_page'))

@app.route('/admin/permissions', methods=['POST'])
@permission_required('admin')
def admin_permissions():
    for role in ['dg', 'rh', 'technicien', 'commercial', 'comptable', 'moyens_generaux', 'informatique', 'resp_projet', 'gestionnaire_projet']:
        perms = [p for p in ALL_PERMISSIONS if request.form.get(f'{role}_{p}')]
        update_role_permissions(role, perms)
    # Admin always has all
    update_role_permissions('admin', ALL_PERMISSIONS)
    flash("Permissions mises à jour", "success")
    return redirect(url_for('admin_page'))


# ======================== LOGS D'ACTIVITÉ ========================

@app.route('/logs')
@permission_required('logs')
def logs_page():
    logs = get_activity_logs(200)
    return render_template('logs.html', page='logs', logs=logs)


# ======================== DÉTAIL RAPPORT + COMMENTAIRES ========================

@app.route('/job/<job_id>')
@login_required
def job_detail(job_id):
    job = get_job_by_id(job_id)
    if not job:
        flash("Rapport non trouvé", "error")
        return redirect(url_for('dashboard'))
    comments = get_job_comments(job_id)
    return render_template('job_detail.html', page='fichiers', job=job, comments=comments)

@app.route('/job/<job_id>/comment', methods=['POST'])
@login_required
def job_add_comment(job_id):
    comment = request.form.get('comment', '').strip()
    if comment:
        user = get_user_by_id(session['user_id'])
        add_job_comment(job_id, session['user_id'], user['full_name'] if user else '?', comment)
        log_activity(session['user_id'], user['full_name'] if user else '?',
                    'Commentaire', f"Commentaire ajouté sur rapport {job_id}", request.remote_addr)
    return redirect(url_for('job_detail', job_id=job_id))

@app.route('/job/<job_id>/notes', methods=['POST'])
@login_required
def job_update_notes(job_id):
    notes = request.form.get('notes', '').strip()
    update_job_notes(job_id, notes)
    flash("Notes mises à jour", "success")
    return redirect(url_for('job_detail', job_id=job_id))


# ======================== SAUVEGARDE BDD ========================

@app.route('/admin/backup')
@permission_required('admin')
def admin_backup():
    db_path = get_db_path()
    if os.path.exists(db_path):
        user = get_user_by_id(session['user_id'])
        log_activity(session['user_id'], user['full_name'] if user else '?',
                    'Sauvegarde', 'Téléchargement de la base de données', request.remote_addr)
        return send_file(db_path, as_attachment=True,
                        download_name=f"ramya_backup_{datetime.now().strftime('%Y%m%d_%H%M')}.db")
    flash("Base de données introuvable", "error")
    return redirect(url_for('admin_page'))


# ======================== PWA MANIFEST ========================

@app.route('/manifest.json')
def pwa_manifest():
    manifest = {
        "name": "WannyGest",
        "short_name": "WannyGest",
        "start_url": "/dashboard",
        "display": "standalone",
        "background_color": "#0d2137",
        "theme_color": "#1a3a5c",
        "icons": [
            {"src": "/static/logo_wannygest.png", "sizes": "192x192", "type": "image/png"},
            {"src": "/static/logo_wannygest.png", "sizes": "512x512", "type": "image/png"}
        ]
    }
    return jsonify(manifest)

@app.route('/sw.js')
def service_worker():
    return app.send_static_file('sw.js') if os.path.exists(os.path.join(BASE_DIR, 'sw.js')) else ('', 204)


# ======================== CONTRATS ========================

@app.route('/contrats')
@permission_required('contrats')
def contrats_page():
    contracts = get_all_contracts()
    clients = get_all_clients()
    return render_template('contrats.html', page='contrats', contracts=contracts, clients=clients)

@app.route('/contrats/add', methods=['POST'])
@permission_required('contrats')
def contrats_add():
    create_contract(
        int(request.form['client_id']),
        request.form.get('reference', ''),
        request.form.get('start_date', ''),
        request.form.get('end_date', ''),
        float(request.form.get('monthly_rate', 0) or 0),
        request.form.get('description', ''),
        session['user_id']
    )
    user = get_user_by_id(session['user_id'])
    log_activity(session['user_id'], user['full_name'] if user else '?',
                'Contrat', 'Nouveau contrat ajouté', request.remote_addr)
    flash("Contrat ajouté", "success")
    return redirect(url_for('contrats_page'))

@app.route('/contrats/edit/<int:cid>', methods=['GET', 'POST'])
@permission_required('contrats')
def contrats_edit(cid):
    contract = get_contract_by_id(cid)
    if not contract:
        flash("Contrat non trouvé", "error")
        return redirect(url_for('contrats_page'))
    if request.method == 'POST':
        update_contract(cid,
            client_id=int(request.form['client_id']),
            reference=request.form.get('reference', ''),
            start_date=request.form.get('start_date', ''),
            end_date=request.form.get('end_date', ''),
            monthly_rate=float(request.form.get('monthly_rate', 0) or 0),
            description=request.form.get('description', ''),
            status=request.form.get('status', 'actif'))
        flash("Contrat modifié", "success")
        return redirect(url_for('contrats_page'))
    clients = get_all_clients()
    return render_template('edit_contract.html', page='contrats', contract=contract, clients=clients)

@app.route('/contrats/delete/<int:cid>')
@permission_required('contrats')
def contrats_delete(cid):
    delete_contract(cid)
    flash("Contrat supprimé", "success")
    return redirect(url_for('contrats_page'))


# ======================== COMPARAISON MENSUELLE ========================

@app.route('/comparaison')
@login_required
def comparaison_page():
    stats = get_client_monthly_stats()
    # Collect all months
    all_months = set()
    for client_data in stats.values():
        all_months.update(client_data.keys())
    months = sorted(all_months)
    return render_template('comparaison.html', page='comparaison', stats=stats, months=months)


# ======================== ALERTES ========================

@app.route('/alertes')
@login_required
def alertes_page():
    # Analyser les derniers rapports pour trouver les alertes
    jobs = get_all_jobs()
    alerts = []
    
    for job in jobs:
        if not job.get('job_id'):
            continue
        # Charger les données du rapport
        files_dir = os.path.join(app.config['FILES_FOLDER'], job['job_id'])
        xlsx_path = None
        if os.path.isdir(files_dir):
            for f in os.listdir(files_dir):
                if f.endswith('.xlsx'):
                    xlsx_path = os.path.join(files_dir, f)
                    break
        
        if not xlsx_path or not os.path.exists(xlsx_path):
            continue
        
        try:
            emps, _ = extract_from_excel(xlsx_path)
            from rapport_core import calc_employee_stats
            for emp in emps:
                enriched, stats = calc_employee_stats(emp)
                # Alertes retards excessifs (>5 jours)
                if stats['days_late'] >= 5:
                    alerts.append({
                        'type': 'retard',
                        'severity': 'high' if stats['days_late'] >= 10 else 'medium',
                        'employee': emp['name'],
                        'client': job['client_name'],
                        'detail': f"{stats['days_late']} jours de retard",
                        'period': job.get('period', ''),
                        'job_id': job['job_id']
                    })
                # Alertes absences (>3 jours)
                if stats['days_absent'] >= 3:
                    alerts.append({
                        'type': 'absence',
                        'severity': 'high' if stats['days_absent'] >= 5 else 'medium',
                        'employee': emp['name'],
                        'client': job['client_name'],
                        'detail': f"{stats['days_absent']} jours d'absence",
                        'period': job.get('period', ''),
                        'job_id': job['job_id']
                    })
        except:
            continue
    
    # Trier: high en premier
    alerts.sort(key=lambda a: (0 if a['severity'] == 'high' else 1, a['client']))
    return render_template('alertes.html', page='alertes', alerts=alerts)


# ======================== EMPLOI DU TEMPS & ANOMALIES ========================

@app.route('/emploi-du-temps')
@permission_required('traitement')
def schedule_page():
    conn = _gdb()
    schedules = [dict(r) for r in conn.execute("SELECT * FROM schedules ORDER BY employee_name, day_of_week").fetchall()]
    emp_schedules = {}
    for s in schedules:
        name = s['employee_name']
        if name not in emp_schedules: emp_schedules[name] = {}
        emp_schedules[name][s['day_of_week']] = s
    
    anomalies = [dict(r) for r in conn.execute("""SELECT * FROM presence_anomalies 
        ORDER BY date DESC LIMIT 100""").fetchall()]
    conn.close()
    
    # Collect employee names from DB (saved when files are processed)
    from models import get_known_employees
    known = get_known_employees()
    all_employees = {e['name']: e.get('service','') for e in known}
    for s in schedules:
        if s['employee_name'] not in all_employees:
            all_employees[s['employee_name']] = ''
    
    # Group employees by service
    by_service = {}
    for name, svc in sorted(all_employees.items()):
        svc = svc or 'Non défini'
        if svc not in by_service: by_service[svc] = []
        by_service[svc].append(name)
    
    days = {0:'Lundi', 1:'Mardi', 2:'Mercredi', 3:'Jeudi', 4:'Vendredi', 5:'Samedi', 6:'Dimanche'}
    return render_template('emploi_du_temps.html', page='schedule', emp_schedules=emp_schedules,
        days=days, anomalies=anomalies, employees=sorted(all_employees.keys()),
        by_service=by_service)

@app.route('/emploi-du-temps/add', methods=['POST'])
@permission_required('traitement')
def schedule_add():
    names = request.form.getlist('employee_names')  # Multi-select
    if not names:
        single = request.form.get('employee_name', '').strip()
        if single: names = [single]
    
    start = request.form.get('start_time', '08:00')
    end = request.form.get('end_time', '17:00')
    bstart = request.form.get('break_start', '12:00')
    bend = request.form.get('break_end', '13:00')
    stype = request.form.get('schedule_type', 'standard')
    
    days = request.form.getlist('days')
    if not days:
        days = ['0','1','2','3','4']
    
    conn = _gdb()
    for name in names:
        name = name.strip()
        if not name: continue
        conn.execute("DELETE FROM schedules WHERE employee_name=?", (name,))
        for d in days:
            conn.execute("""INSERT INTO schedules (employee_name, day_of_week, start_time, end_time, 
                break_start, break_end, schedule_type) VALUES (?,?,?,?,?,?,?)""",
                (name, int(d), start, end, bstart, bend, stype))
    conn.commit(); conn.close()
    flash(f"Emploi du temps enregistré pour {len(names)} personne(s)", "success")
    return redirect('/emploi-du-temps')

@app.route('/emploi-du-temps/delete/<path:name>')
@permission_required('traitement')
def schedule_delete(name):
    from urllib.parse import unquote
    name = unquote(name)
    conn = _gdb()
    conn.execute("DELETE FROM schedules WHERE employee_name=?", (name,))
    conn.commit(); conn.close()
    flash(f"Emploi du temps de {name} supprimé", "success")
    return redirect('/emploi-du-temps')

@app.route('/emploi-du-temps/detect-anomalies', methods=['POST'])
@permission_required('traitement')
def schedule_detect():
    """Compare les fichiers de présence avec les emplois du temps définis."""
    
    files_dir = app.config.get('FILES_FOLDER', 'files')
    xlsx_path = None
    
    # Find latest presence file
    jobs = get_all_jobs()
    for job in reversed(jobs):
        jdir = os.path.join(files_dir, job['job_id'])
        if os.path.isdir(jdir):
            for f in os.listdir(jdir):
                if f.endswith('.xlsx'):
                    xlsx_path = os.path.join(jdir, f)
                    break
            if xlsx_path: break
    
    if not xlsx_path:
        flash("Aucun fichier de présence trouvé. Traitez d'abord un fichier.", "error")
        return redirect('/emploi-du-temps')
    
    # Load our defined schedules
    conn = _gdb()
    scheds = [dict(r) for r in conn.execute("SELECT * FROM schedules").fetchall()]
    sched_map = {}
    for s in scheds:
        name = s['employee_name']
        if name not in sched_map: sched_map[name] = {}
        sched_map[name][s['day_of_week']] = s
    
    # Parse presence file
    try:
        emps, meta = extract_from_excel(xlsx_path)
    except Exception as e:
        flash(f"Erreur lecture fichier: {e}", "error")
        conn.close()
        return redirect('/emploi-du-temps')
    
    # Clear old anomalies
    conn.execute("DELETE FROM presence_anomalies")
    
    def time_to_min(t):
        """Convertit HH:MM en minutes depuis minuit."""
        if not t or len(t) < 4: return None
        try:
            parts = t.replace('h',':').split(':')
            return int(parts[0]) * 60 + int(parts[1])
        except: return None
    
    anomaly_count = 0
    tolerance = 15  # minutes de tolérance
    
    for emp in emps:
        name = emp.get('name', '').strip()
        if name not in sched_map: continue
        
        for rec in emp.get('records', []):
            date = rec.get('date', '')
            if not date: continue
            
            try:
                from datetime import datetime as dt2
                d = dt2.strptime(date[:10], '%Y-%m-%d')
                dow = d.weekday()
            except: continue
            
            if dow not in sched_map[name]: continue
            sched = sched_map[name][dow]
            
            actual_arr = rec.get('arrival', '').strip()
            actual_dep = rec.get('departure', '').strip()
            exp_start = sched['start_time']
            exp_end = sched['end_time']
            
            exp_s = time_to_min(exp_start)
            exp_e = time_to_min(exp_end)
            act_s = time_to_min(actual_arr)
            act_e = time_to_min(actual_dep)
            
            anomalies = []
            
            if not act_s and not act_e:
                # Pas de pointage du tout
                anomalies.append('absence')
            elif act_s is not None and exp_s is not None:
                # Retard : arrivée > prévu + tolérance
                if act_s > exp_s + tolerance:
                    anomalies.append('retard')
                
                # Horaire complètement décalé : arrivée APRÈS l'heure de fin prévue
                # Ex: prévu 07:00-17:00 mais pointe à 17:00-21:00
                if exp_e is not None and act_s >= exp_e:
                    anomalies.append('horaire_decale')
                
                # Départ anticipé : départ > 30min avant la fin prévue
                if act_e is not None and exp_e is not None and act_e < exp_e - 30:
                    anomalies.append('depart_anticipe')
                
                # Arrivée très en avance (plus de 2h avant l'heure prévue)
                if act_s < exp_s - 120:
                    anomalies.append('horaire_decale')
            
            if anomalies:
                anomaly_type = '+'.join(anomalies)
                conn.execute("""INSERT INTO presence_anomalies 
                    (employee_name, date, expected_start, expected_end, actual_start, actual_end, anomaly_type)
                    VALUES (?,?,?,?,?,?,?)""",
                    (name, date, exp_start, exp_end, actual_arr or '', actual_dep or '', anomaly_type))
                anomaly_count += 1
    
    conn.commit(); conn.close()
    flash(f"{anomaly_count} anomalie(s) détectée(s) sur {len(emps)} employés", "info")
    return redirect('/emploi-du-temps')

@app.route('/emploi-du-temps/correct/<int:aid>', methods=['POST'])
@permission_required('traitement')
def schedule_correct(aid):
    conn = _gdb()
    conn.execute("""UPDATE presence_anomalies SET status='corrigee', 
        corrected_start=?, corrected_end=?, notes=? WHERE id=?""",
        (request.form.get('corrected_start',''), request.form.get('corrected_end',''),
         request.form.get('notes',''), aid))
    conn.commit(); conn.close()
    flash("Anomalie corrigée", "success")
    return redirect('/emploi-du-temps')


# ======================== CALCUL D'HEURES DPCI ========================

@app.route('/dpci')
@permission_required('traitement')
def dpci_page():
    return render_template('dpci.html', page='dpci')

@app.route('/dpci/preview', methods=['POST'])
@login_required
def dpci_preview():
    if 'excel_file' not in request.files:
        return jsonify({"error": "Fichier requis"}), 400
    f = request.files['excel_file']
    if not f.filename:
        return jsonify({"error": "Fichier vide"}), 400
    
    import tempfile
    tmp = os.path.join(tempfile.gettempdir(), f'dpci_{uuid.uuid4().hex[:8]}.xlsx')
    f.save(tmp)
    try:
        from dpci import parse_dpci_excel
        emps, period = parse_dpci_excel(tmp)
        
        # Save known employees
        from models import save_known_employees
        save_known_employees([e['name'] for e in emps])
        
        # Extract client from department (first part before >)
        client = ''
        if emps:
            dept = emps[0].get('department', '')
            if '>' in dept:
                client = dept.split('>')[1].strip()
            else:
                client = dept
        
        return jsonify({
            "count": len(emps),
            "period": period,
            "client": client,
            "employees": [{"name": e['name'], "id": e['id'], "dept": e['department'], "days": len(e['records'])} for e in emps]
        })
    except Exception as e:
        return jsonify({"error": str(e)}), 500
    finally:
        if os.path.exists(tmp): os.remove(tmp)

@app.route('/dpci/generate', methods=['POST'])
@permission_required('traitement')
def dpci_generate():
    if 'excel_file' not in request.files:
        flash("Fichier requis", "error")
        return redirect('/dpci')
    
    f = request.files['excel_file']
    client_name = request.form.get('client_name', '').strip() or 'DPCI'
    provider_name = request.form.get('provider_name', '').strip() or 'RAMYA TECHNOLOGIE & INNOVATION'
    default_cost = float(request.form.get('default_cost', 0) or 0)
    try:
        hp = float(request.form.get('required_hours', 0) or 0)
    except:
        hp = 0
    try:
        hp_weekend = float(request.form.get('required_hours_weekend', 0) or 0)
    except:
        hp_weekend = 0
    
    try:
        employee_costs = json.loads(request.form.get('employee_costs_json', '{}'))
    except:
        employee_costs = {}
    
    # Save file
    job_id = str(uuid.uuid4())[:8]
    job_dir = os.path.join(app.config['UPLOAD_FOLDER'], f'dpci_{job_id}')
    os.makedirs(job_dir, exist_ok=True)
    
    xlsx_path = os.path.join(job_dir, secure_filename(f.filename))
    f.save(xlsx_path)
    
    try:
        from dpci import parse_dpci_excel, generate_dpci_pdf
        emps, period = parse_dpci_excel(xlsx_path)
        
        if not emps:
            flash("Aucun employé trouvé dans le fichier", "error")
            return redirect('/dpci')
        
        # Filter by period
        period_mode = request.form.get('period_mode', 'all')
        period_start = request.form.get('period_start', '').strip()
        period_end = request.form.get('period_end', '').strip()
        
        if period_mode == 'day' and period_start:
            period_end = period_start
        elif period_mode == 'week' and period_start:
            from datetime import datetime as dt2, timedelta
            try:
                d = dt2.strptime(period_start, '%Y-%m-%d')
                period_end = (d + timedelta(days=6)).strftime('%Y-%m-%d')
            except: pass
        
        if period_start or period_end:
            for emp in emps:
                emp['records'] = [r for r in emp['records'] 
                    if (not period_start or r['date'] >= period_start) and 
                       (not period_end or r['date'] <= period_end)]
            emps = [e for e in emps if e['records']]
        
        if not emps:
            flash("Aucun enregistrement pour la période sélectionnée", "error")
            return redirect('/dpci')
        
        # Build period string
        all_dates = sorted([r['date'] for e in emps for r in e['records']])
        if all_dates:
            period_str = f"Débuté le {all_dates[0]} au {all_dates[-1]}"
        elif period:
            parts = period.split(' - ')
            period_str = f"Débuté le {parts[0].strip()} au {parts[-1].strip()}" if len(parts) > 1 else period
        else:
            period_str = "Rapport DPCI"
        conn = _gdb()
        scheds = [dict(r) for r in conn.execute("SELECT * FROM schedules").fetchall()]
        conn.close()
        
        schedules_map = {}
        for s in scheds:
            name = s['employee_name']
            if name not in schedules_map:
                schedules_map[name] = s  # Use first found (e.g., Monday schedule)
        
        # Generate PDF
        pdf_name = f"DPCI_{client_name.replace(' ', '_')}_{job_id}.pdf"
        output_path = os.path.join(job_dir, pdf_name)
        
        user = get_user_by_id(session['user_id'])
        treated_by = user['full_name'] if user else 'Admin'
        
        generate_dpci_pdf(emps, output_path, client_name, period_str,
                         schedules_map=schedules_map, employee_costs=employee_costs,
                         default_cost=default_cost, hp=hp, hp_weekend=hp_weekend,
                         provider_name=provider_name, treated_by=treated_by, period_mode=period_mode)
        
        if not os.path.exists(output_path):
            flash("Erreur de génération PDF", "error")
            return redirect('/dpci')
        
        # Save to files folder
        files_dir = os.path.join(app.config['FILES_FOLDER'], f'dpci_{job_id}')
        os.makedirs(files_dir, exist_ok=True)
        shutil.copy2(output_path, os.path.join(files_dir, pdf_name))
        shutil.copy2(xlsx_path, os.path.join(files_dir, os.path.basename(xlsx_path)))
        
        # Log
        user = get_user_by_id(session['user_id'])
        log_activity(session['user_id'], user['full_name'] if user else '?',
                    'DPCI', f"Rapport {client_name} — {len(emps)} employés", request.remote_addr)
        
        flash(f"Rapport DPCI généré — {len(emps)} employés", "success")
        return send_file(output_path, as_attachment=True, download_name=pdf_name)
    
    except Exception as e:
        flash(f"Erreur: {e}", "error")
        return redirect('/dpci')


# ======================== ENVOI EMAIL ========================

@app.route('/fichiers/email/<job_id>', methods=['GET', 'POST'])
@permission_required('envoyer')
def fichiers_email(job_id):
    job = get_job_by_id(job_id)
    if not job:
        flash("Rapport non trouvé", "error")
        return redirect(url_for('fichiers'))
    
    # Pré-remplir avec les infos client
    client = get_client_by_id(job['client_id']) if job.get('client_id') else None
    default_email = client['email'] if client and client.get('email') else ''
    
    if request.method == 'POST':
        to_email = request.form.get('to_email', '').strip()
        subject = request.form.get('subject', '').strip()
        body = request.form.get('body', '').strip()
        smtp_host = request.form.get('smtp_host', '').strip()
        smtp_port = int(request.form.get('smtp_port', 587))
        smtp_user = request.form.get('smtp_user', '').strip()
        smtp_pass = request.form.get('smtp_pass', '').strip()
        
        # Sauvegarder les paramètres SMTP
        if smtp_host and smtp_user:
            save_smtp_settings(session['user_id'], smtp_host, smtp_port, smtp_user, smtp_pass)
        
        if not all([to_email, subject, smtp_host, smtp_user, smtp_pass]):
            flash("Tous les champs SMTP sont obligatoires", "error")
            smtp = get_smtp_settings(session['user_id'])
            return render_template('email_send.html', page='fichiers', job=job,
                                 default_email=default_email, smtp=smtp)
        
        # Préparer le fichier PDF
        files_dir = os.path.join(app.config['FILES_FOLDER'], secure_filename(job_id))
        pdf_path = None
        if os.path.isdir(files_dir):
            for f in os.listdir(files_dir):
                if f.endswith('.pdf'):
                    pdf_path = os.path.join(files_dir, f)
                    break
        
        if not pdf_path:
            flash("Fichier PDF non trouvé", "error")
            return redirect(url_for('fichiers'))
        
        try:
            msg = MIMEMultipart()
            msg['From'] = smtp_user
            msg['To'] = to_email
            msg['Subject'] = subject
            msg.attach(MIMEText(body, 'plain', 'utf-8'))
            
            # Joindre le PDF
            with open(pdf_path, 'rb') as f:
                part = MIMEBase('application', 'pdf')
                part.set_payload(f.read())
                encoders.encode_base64(part)
                part.add_header('Content-Disposition', f'attachment; filename="{os.path.basename(pdf_path)}"')
                msg.attach(part)
            
            server = smtplib.SMTP(smtp_host, smtp_port)
            server.starttls()
            server.login(smtp_user, smtp_pass)
            server.send_message(msg)
            server.quit()
            
            # Marquer comme envoyé
            mark_job_sent(job_id, session['user_id'])
            user = get_user_by_id(session['user_id'])
            log_activity(session['user_id'], user['full_name'] if user else '?',
                        'Email', f"Rapport envoyé par email à {to_email}", request.remote_addr)
            
            flash(f"Email envoyé avec succès à {to_email}", "success")
            return redirect(url_for('fichiers'))
        
        except Exception as e:
            flash(f"Erreur d'envoi : {str(e)}", "error")
    
    smtp = get_smtp_settings(session['user_id'])
    return render_template('email_send.html', page='fichiers', job=job, default_email=default_email, smtp=smtp)


# ======================== COMPTABILITÉ ========================

@app.route('/comptabilite')
@permission_required('comptabilite')
def comptabilite_page():
    tab = request.args.get('tab', 'a_envoyer')
    inv_stats = get_invoice_stats()
    invoices = get_invoices_by_status(tab) if tab != 'all' else get_all_invoices()
    
    # Dashboard data
    conn = _gdb()
    # Monthly revenue (paid invoices)
    monthly = [dict(r) for r in conn.execute("""
        SELECT strftime('%Y-%m', created_at) as month, SUM(amount) as total 
        FROM invoices WHERE status='payee' GROUP BY month ORDER BY month DESC LIMIT 12
    """).fetchall()]
    monthly.reverse()
    
    # Monthly expenses (from pieces_caisse)
    expenses = [dict(r) for r in conn.execute("""
        SELECT strftime('%Y-%m', date) as month, SUM(amount) as total 
        FROM pieces_caisse GROUP BY month ORDER BY month DESC LIMIT 12
    """).fetchall()]
    expenses.reverse()
    
    # Totals
    total_revenue = conn.execute("SELECT COALESCE(SUM(amount),0) FROM invoices WHERE status='payee'").fetchone()[0]
    total_pending = conn.execute("SELECT COALESCE(SUM(amount),0) FROM invoices WHERE status IN ('envoyee','en_attente_paiement')").fetchone()[0]
    total_expenses = conn.execute("SELECT COALESCE(SUM(amount),0) FROM pieces_caisse").fetchone()[0]
    
    # Expense by category
    exp_cats = [dict(r) for r in conn.execute("""
        SELECT category, SUM(amount) as total FROM pieces_caisse GROUP BY category ORDER BY total DESC LIMIT 8
    """).fetchall()]
    
    # Recent devis (for comptabilite to see new devis from CRM)
    recent_devis = [dict(r) for r in conn.execute("""
        SELECT * FROM devis ORDER BY created_at DESC LIMIT 5
    """).fetchall()]
    conn.close()
    
    chart_data = {
        'months': [m['month'] for m in monthly],
        'revenue': [m['total'] for m in monthly],
        'exp_months': [e['month'] for e in expenses],
        'exp_totals': [e['total'] for e in expenses],
        'exp_cats': [c['category'] for c in exp_cats],
        'exp_cat_vals': [c['total'] for c in exp_cats],
        'total_revenue': total_revenue,
        'total_pending': total_pending,
        'total_expenses': total_expenses,
        'profit': total_revenue - total_expenses,
    }
    
    return render_template('comptabilite.html', page='comptabilite', tab=tab,
                          invoices=invoices, inv_stats=inv_stats, chart=chart_data, recent_devis=recent_devis)

@app.route('/comptabilite/status/<int:inv_id>/<status>')
@permission_required('comptabilite')
def comptabilite_status(inv_id, status):
    if status in ('envoyee', 'en_attente_paiement', 'payee', 'a_envoyer'):
        update_invoice_status(inv_id, status, session.get('user_id'))
        user = get_user_by_id(session['user_id'])
        log_activity(session['user_id'], user['full_name'] if user else '?',
                    'Facture', f"Facture #{inv_id} → {status}", request.remote_addr)
        flash(f"Statut mis à jour : {status}", "success")
    return redirect(url_for('comptabilite_page'))

@app.route('/comptabilite/facture/new', methods=['GET', 'POST'])
@permission_required('comptabilite_edit')
def invoice_new():
    if request.method == 'POST':
        ref = f"FAC-{datetime.now().strftime('%Y%m%d%H%M%S')}"
        conn = _gdb()
        conn.execute("""INSERT INTO invoices (reference, client_name, client_id, amount, objet, 
            description, due_date, payment_method, status, total_ht, tva, total_ttc, notes)
            VALUES (?,?,?,?,?,?,?,?,?,?,?,?,?)""",
            (ref, request.form.get('client_name',''), int(request.form.get('client_id',0) or 0) or None,
             float(request.form.get('amount',0) or 0), request.form.get('objet',''),
             request.form.get('description',''), request.form.get('due_date',''),
             request.form.get('payment_method',''), 'a_envoyer',
             float(request.form.get('total_ht',0) or 0), float(request.form.get('tva',0) or 0),
             float(request.form.get('total_ttc',0) or 0), request.form.get('notes','')))
        conn.commit(); conn.close()
        flash(f"Facture {ref} créée", "success")
        return redirect(url_for('comptabilite_page'))
    clients = get_all_clients()
    return render_template('invoice_new.html', page='comptabilite', clients=clients)

@app.route('/devis/convert/<int:did>')
@permission_required('convertir_devis')
def devis_to_invoice(did):
    from models import db_get_by_id
    d = db_get_by_id('devis', did)
    if not d:
        flash("Devis non trouvé", "error"); return redirect('/devis')
    ref = f"FAC-{datetime.now().strftime('%Y%m%d%H%M%S')}"
    conn = _gdb()
    conn.execute("""INSERT INTO invoices (reference, client_name, client_id, amount, objet,
        total_ht, tva, total_ttc, items_json, devis_id, status, notes)
        VALUES (?,?,?,?,?,?,?,?,?,?,?,?)""",
        (ref, d.get('client_name',''), d.get('client_id'), d.get('total_ttc',0),
         d.get('objet',''), d.get('total_ht',0), d.get('total_ht',0)*0.18 if d.get('total_ht') else 0,
         d.get('total_ttc',0), d.get('items_json',''), did, 'a_envoyer',
         f"Convertie depuis devis {d.get('reference','')}"))
    conn.commit(); conn.close()
    user = get_user_by_id(session['user_id'])
    log_activity(session['user_id'], user['full_name'] if user else '?',
                'Facture', f"Devis {d.get('reference','')} converti en facture {ref}", request.remote_addr)
    flash(f"Devis {d.get('reference','')} converti en facture {ref}", "success")
    return redirect(url_for('comptabilite_page'))

@app.route('/comptabilite/facture/view/<int:fid>')
@permission_required('comptabilite')
def invoice_view(fid):
    conn = _gdb()
    inv = conn.execute("SELECT * FROM invoices WHERE id=?", (fid,)).fetchone()
    conn.close()
    if not inv: flash("Facture non trouvée","error"); return redirect(url_for('comptabilite_page'))
    inv = dict(inv)
    inv['items'] = json.loads(inv.get('items_json','[]') or '[]')
    return render_template('invoice_view.html', page='comptabilite', inv=inv, inv_items=inv['items'])

@app.route('/comptabilite/facture/edit/<int:fid>', methods=['GET', 'POST'])
@permission_required('comptabilite_edit')
def invoice_edit(fid):
    conn = _gdb()
    inv = conn.execute("SELECT * FROM invoices WHERE id=?", (fid,)).fetchone()
    if not inv: flash("Non trouvée","error"); return redirect(url_for('comptabilite_page'))
    if request.method == 'POST':
        conn.execute("""UPDATE invoices SET client_name=?, objet=?, amount=?, total_ht=?, tva=?, total_ttc=?,
            description=?, due_date=?, payment_method=?, notes=? WHERE id=?""",
            (request.form.get('client_name',''), request.form.get('objet',''),
             float(request.form.get('amount',0) or 0), float(request.form.get('total_ht',0) or 0),
             float(request.form.get('tva',0) or 0), float(request.form.get('total_ttc',0) or 0),
             request.form.get('description',''), request.form.get('due_date',''),
             request.form.get('payment_method',''), request.form.get('notes',''), fid))
        conn.commit(); conn.close()
        flash("Facture modifiée","success"); return redirect(f'/comptabilite/facture/view/{fid}')
    conn.close()
    clients = get_all_clients()
    return render_template('invoice_edit.html', page='comptabilite', inv=dict(inv), clients=clients)

# ======================== CAISSE (Entrées + Sorties) ========================

@app.route('/caisse-entree/add', methods=['POST'])
@permission_required('caisse_sortie')
def caisse_entree_add():
    conn = _gdb()
    ref = f"CE-{datetime.now().strftime('%Y%m%d%H%M%S')}"
    conn.execute("""INSERT INTO caisse_entrees (reference, date, source, montant, description, payment_method, created_by)
        VALUES (?,?,?,?,?,?,?)""",
        (ref, request.form.get('date',''), request.form.get('source',''),
         float(request.form.get('montant',0) or 0), request.form.get('description',''),
         request.form.get('payment_method',''), session['user_id']))
    conn.commit(); conn.close()
    flash("Entrée de caisse enregistrée","success")
    return redirect('/caisse-sortie?tab=entrees')

# ======================== BILAN COMPTABLE ========================

@app.route('/comptabilite/bilan')
@permission_required('comptabilite')
def bilan_comptable():
    tab = request.args.get('tab', 'mensuel')
    month = request.args.get('month', datetime.now().strftime('%Y-%m'))
    exercice = request.args.get('exercice', datetime.now().strftime('%Y'))
    conn = _gdb()
    
    if tab == 'plan':
        comptes = [dict(r) for r in conn.execute("SELECT * FROM plan_comptable ORDER BY numero").fetchall()]
        conn.close()
        return render_template('bilan_comptable.html', page='bilan', tab=tab, comptes=comptes, exercice=exercice)
    
    if tab == 'ecritures':
        ecritures = [dict(r) for r in conn.execute("""SELECT e.*, u.full_name FROM ecritures_comptables e 
            LEFT JOIN users u ON e.created_by=u.id WHERE strftime('%Y',e.date)=? ORDER BY e.date DESC, e.id DESC""",
            (exercice,)).fetchall()]
        comptes = [dict(r) for r in conn.execute("SELECT numero, libelle FROM plan_comptable ORDER BY numero").fetchall()]
        conn.close()
        return render_template('bilan_comptable.html', page='bilan', tab=tab, ecritures=ecritures, comptes=comptes, exercice=exercice)
    
    if tab == 'bilan_officiel':
        # Compute official bilan from écritures
        rows = conn.execute("""SELECT pc.numero, pc.libelle, pc.type, pc.categorie, pc.classe,
            COALESCE(SUM(CASE WHEN ec.compte_debit=pc.numero THEN ec.montant ELSE 0 END),0) as total_debit,
            COALESCE(SUM(CASE WHEN ec.compte_credit=pc.numero THEN ec.montant ELSE 0 END),0) as total_credit
            FROM plan_comptable pc
            LEFT JOIN ecritures_comptables ec ON (ec.compte_debit=pc.numero OR ec.compte_credit=pc.numero) AND strftime('%Y',ec.date)=?
            GROUP BY pc.numero ORDER BY pc.numero""", (exercice,)).fetchall()
        
        actif_immob, actif_circ, actif_treso = [], [], []
        passif_cap, passif_dettes = [], []
        charges, produits = [], []
        
        for r in rows:
            d = dict(r)
            solde = d['total_debit'] - d['total_credit']
            d['solde'] = abs(solde)
            if d['categorie'] in ('immobilise',) and d['classe'] == '2':
                actif_immob.append(d)
            elif d['categorie'] in ('circulant',) and d['classe'] in ('3','4'):
                if d['type'] == 'actif': actif_circ.append(d)
                else: passif_dettes.append(d)
            elif d['categorie'] in ('tresorerie',):
                actif_treso.append(d)
            elif d['categorie'] in ('capitaux','dettes_financieres'):
                passif_cap.append(d)
            elif d['categorie'] in ('dettes_circulant',):
                passif_dettes.append(d)
            elif d['categorie'] == 'charges':
                charges.append(d)
            elif d['categorie'] == 'produits':
                produits.append(d)
        
        t_actif = sum(c['solde'] for c in actif_immob + actif_circ + actif_treso if c['solde'] > 0)
        t_passif = sum(c['solde'] for c in passif_cap + passif_dettes if c['solde'] > 0)
        t_charges = sum(c['solde'] for c in charges)
        t_produits = sum(c['solde'] for c in produits)
        resultat = t_produits - t_charges
        
        bilans_hist = [dict(r) for r in conn.execute("SELECT * FROM bilans WHERE exercice=? ORDER BY created_at DESC", (exercice,)).fetchall()]
        conn.close()
        
        bilan_data = {
            'actif_immob': actif_immob, 'actif_circ': actif_circ, 'actif_treso': actif_treso,
            'passif_cap': passif_cap, 'passif_dettes': passif_dettes,
            'charges': charges, 'produits': produits,
            't_actif': t_actif, 't_passif': t_passif + resultat,
            't_charges': t_charges, 't_produits': t_produits, 'resultat': resultat,
            'equilibre': abs(t_actif - (t_passif + resultat)) < 1,
            'bilans_hist': bilans_hist
        }
        return render_template('bilan_comptable.html', page='bilan', tab=tab, bilan=bilan_data, exercice=exercice)
    
    # Default: mensuel overview
    fp = conn.execute("SELECT COALESCE(SUM(amount),0) FROM invoices WHERE status='payee' AND strftime('%Y-%m',paid_at)=?", (month,)).fetchone()[0]
    fpen = conn.execute("SELECT COALESCE(SUM(amount),0) FROM invoices WHERE status IN ('envoyee','en_attente_paiement') AND strftime('%Y-%m',created_at)=?", (month,)).fetchone()[0]
    ce = conn.execute("SELECT COALESCE(SUM(montant),0) FROM caisse_entrees WHERE strftime('%Y-%m',date)=?", (month,)).fetchone()[0]
    cs = conn.execute("SELECT COALESCE(SUM(montant),0) FROM caisse_sorties WHERE status='approuve' AND strftime('%Y-%m',date)=?", (month,)).fetchone()[0]
    tr = conn.execute("SELECT COALESCE(SUM(amount),0) FROM treasury WHERE movement_type='recette' AND strftime('%Y-%m',created_at)=?", (month,)).fetchone()[0]
    td = conn.execute("SELECT COALESCE(SUM(amount),0) FROM treasury WHERE movement_type='depense' AND strftime('%Y-%m',created_at)=?", (month,)).fetchone()[0]
    dp = conn.execute("SELECT COALESCE(SUM(amount),0) FROM pieces_caisse WHERE strftime('%Y-%m',date)=?", (month,)).fetchone()[0]
    banks = [dict(r) for r in conn.execute("SELECT * FROM bank_accounts ORDER BY name").fetchall()]
    inv_list = [dict(r) for r in conn.execute("SELECT * FROM invoices WHERE strftime('%Y-%m',created_at)=? ORDER BY created_at DESC", (month,)).fetchall()]
    conn.close()
    
    te = fp + ce + tr; ts = cs + td + dp
    data = {'month': month, 'factures_payees': fp, 'factures_pending': fpen,
            'caisse_entrees': ce, 'caisse_sorties': cs, 'treso_recettes': tr, 'treso_depenses': td,
            'depenses': dp, 'total_entrees': te, 'total_sorties': ts, 'solde': te - ts,
            'banks': banks, 'inv_list': inv_list}
    return render_template('bilan_comptable.html', page='bilan', tab=tab, data=data, exercice=exercice)

@app.route('/comptabilite/ecritures/add', methods=['POST'])
@permission_required('comptabilite_edit')
def ecriture_add():
    conn = _gdb()
    conn.execute("""INSERT INTO ecritures_comptables (date, journal, piece, compte_debit, compte_credit, libelle, montant, created_by)
        VALUES (?,?,?,?,?,?,?,?)""",
        (request.form.get('date',''), request.form.get('journal','OD'), request.form.get('piece',''),
         request.form.get('compte_debit',''), request.form.get('compte_credit',''),
         request.form.get('libelle',''), float(request.form.get('montant',0) or 0), session['user_id']))
    conn.commit(); conn.close()
    flash("Écriture enregistrée","success")
    year = request.form.get('date','')[:4] or datetime.now().strftime('%Y')
    return redirect(f'/comptabilite/bilan?tab=ecritures&exercice={year}')

@app.route('/comptabilite/plan-comptable/add', methods=['POST'])
@permission_required('comptabilite_edit')
def plan_comptable_add():
    conn = _gdb()
    try:
        conn.execute("INSERT INTO plan_comptable (numero, libelle, type, categorie, classe) VALUES (?,?,?,?,?)",
            (request.form.get('numero',''), request.form.get('libelle',''), request.form.get('type','actif'),
             request.form.get('categorie',''), request.form.get('classe','')))
        conn.commit(); flash("Compte ajouté","success")
    except: flash("Ce numéro de compte existe déjà","error")
    conn.close()
    return redirect('/comptabilite/bilan?tab=plan')

@app.route('/comptabilite/bilan/generer', methods=['POST'])
@permission_required('comptabilite_edit')
def bilan_generer():
    exercice = request.form.get('exercice', datetime.now().strftime('%Y'))
    conn = _gdb()
    rows = conn.execute("""SELECT pc.type, pc.categorie,
        COALESCE(SUM(CASE WHEN ec.compte_debit=pc.numero THEN ec.montant ELSE 0 END),0) as td,
        COALESCE(SUM(CASE WHEN ec.compte_credit=pc.numero THEN ec.montant ELSE 0 END),0) as tc
        FROM plan_comptable pc
        LEFT JOIN ecritures_comptables ec ON (ec.compte_debit=pc.numero OR ec.compte_credit=pc.numero) AND strftime('%Y',ec.date)=?
        WHERE pc.categorie NOT IN ('charges','produits')
        GROUP BY pc.numero""", (exercice,)).fetchall()
    
    t_actif = sum(r['td'] - r['tc'] for r in rows if r['type'] == 'actif')
    t_passif = sum(r['tc'] - r['td'] for r in rows if r['type'] == 'passif')
    
    # Résultat
    ch = conn.execute("""SELECT COALESCE(SUM(CASE WHEN ec.compte_debit=pc.numero THEN ec.montant ELSE 0 END),0)
        FROM plan_comptable pc LEFT JOIN ecritures_comptables ec ON ec.compte_debit=pc.numero AND strftime('%Y',ec.date)=?
        WHERE pc.categorie='charges'""", (exercice,)).fetchone()[0]
    pr = conn.execute("""SELECT COALESCE(SUM(CASE WHEN ec.compte_credit=pc.numero THEN ec.montant ELSE 0 END),0)
        FROM plan_comptable pc LEFT JOIN ecritures_comptables ec ON ec.compte_credit=pc.numero AND strftime('%Y',ec.date)=?
        WHERE pc.categorie='produits'""", (exercice,)).fetchone()[0]
    resultat = pr - ch
    
    bilan_data = {'actif': t_actif, 'passif': t_passif, 'resultat': resultat, 'charges': ch, 'produits': pr}
    conn.execute("""INSERT INTO bilans (exercice, date_cloture, total_actif, total_passif, resultat, data_json, status, created_by)
        VALUES (?,?,?,?,?,?,?,?)""",
        (exercice, datetime.now().strftime('%Y-12-31'), t_actif, t_passif + resultat, resultat,
         json.dumps(bilan_data), 'brouillon', session['user_id']))
    conn.commit(); conn.close()
    flash(f"Bilan exercice {exercice} généré","success")
    return redirect(f'/comptabilite/bilan?tab=bilan_officiel&exercice={exercice}')

@app.route('/comptabilite/bilan/pdf')
@permission_required('comptabilite')
def bilan_pdf():
    exercice = request.args.get('exercice', datetime.now().strftime('%Y'))
    month = request.args.get('month', '')
    tab = request.args.get('tab', 'mensuel')
    
    from reportlab.lib.pagesizes import A4
    from reportlab.lib.units import mm
    from reportlab.lib.colors import HexColor, white
    from reportlab.platypus import SimpleDocTemplate, Paragraph, Spacer, Table, TableStyle
    from reportlab.lib.styles import ParagraphStyle
    from reportlab.lib.enums import TA_CENTER, TA_RIGHT, TA_LEFT
    
    output = os.path.join(app.config['UPLOAD_FOLDER'], f'bilan_{exercice}.pdf')
    os.makedirs(os.path.dirname(output), exist_ok=True)
    doc = SimpleDocTemplate(output, pagesize=A4, leftMargin=12*mm, rightMargin=12*mm, topMargin=10*mm, bottomMargin=10*mm)
    
    BG = HexColor('#44546A'); BL = HexColor('#4472C4'); GR = HexColor('#2e7d32'); RD = HexColor('#c53030')
    hw = ParagraphStyle('hw', fontName='Helvetica-Bold', fontSize=8, textColor=white, alignment=TA_CENTER)
    tc = ParagraphStyle('tc', fontSize=8, alignment=TA_CENTER)
    tl = ParagraphStyle('tl', fontSize=8)
    tr_s = ParagraphStyle('tr', fontSize=8, alignment=TA_RIGHT)
    tb = ParagraphStyle('tb', fontName='Helvetica-Bold', fontSize=9, alignment=TA_RIGHT)
    pw = 186*mm
    fmt = lambda x: f"{x:,.0f}"
    
    conn = _gdb()
    story = []
    
    story.append(Paragraph("RAMYA TECHNOLOGIE &amp; INNOVATION", ParagraphStyle('co', fontSize=9, textColor=HexColor('#999'), alignment=TA_CENTER, spaceAfter=3*mm)))
    story.append(Paragraph(f"<b>BILAN COMPTABLE — Exercice {exercice}</b>", ParagraphStyle('t', fontName='Helvetica-Bold', fontSize=16, textColor=BG, alignment=TA_CENTER, spaceAfter=2*mm)))
    story.append(Paragraph(f"Document conforme au SYSCOHADA révisé — Généré le {datetime.now().strftime('%d/%m/%Y')}", ParagraphStyle('sub', fontSize=7, textColor=HexColor('#999'), alignment=TA_CENTER, spaceAfter=6*mm)))
    
    # Get bilan data
    rows = conn.execute("""SELECT pc.numero, pc.libelle, pc.type, pc.categorie, pc.classe,
        COALESCE(SUM(CASE WHEN ec.compte_debit=pc.numero THEN ec.montant ELSE 0 END),0) as total_debit,
        COALESCE(SUM(CASE WHEN ec.compte_credit=pc.numero THEN ec.montant ELSE 0 END),0) as total_credit
        FROM plan_comptable pc
        LEFT JOIN ecritures_comptables ec ON (ec.compte_debit=pc.numero OR ec.compte_credit=pc.numero) AND strftime('%Y',ec.date)=?
        GROUP BY pc.numero ORDER BY pc.numero""", (exercice,)).fetchall()
    conn.close()
    
    actif_rows = [(dict(r)['numero'], dict(r)['libelle'], abs(dict(r)['total_debit'] - dict(r)['total_credit'])) 
                  for r in rows if dict(r)['type'] == 'actif' and dict(r)['categorie'] not in ('charges','produits')]
    passif_rows = [(dict(r)['numero'], dict(r)['libelle'], abs(dict(r)['total_credit'] - dict(r)['total_debit']))
                   for r in rows if dict(r)['type'] == 'passif' and dict(r)['categorie'] not in ('charges','produits')]
    
    t_actif = sum(r[2] for r in actif_rows)
    t_passif = sum(r[2] for r in passif_rows)
    
    # ACTIF table
    story.append(Paragraph("<b>ACTIF</b>", ParagraphStyle('ha', fontName='Helvetica-Bold', fontSize=11, textColor=BL, spaceAfter=2*mm)))
    ad = [[Paragraph(h, hw) for h in ['N° Compte', 'Libellé', 'Montant (FCFA)']]]
    for num, lib, val in actif_rows:
        if val > 0: ad.append([Paragraph(num, tc), Paragraph(lib, tl), Paragraph(fmt(val), tr_s)])
    ad.append([Paragraph('', tc), Paragraph('<b>TOTAL ACTIF</b>', ParagraphStyle('ta', fontName='Helvetica-Bold', fontSize=9)),
               Paragraph(f'<b>{fmt(t_actif)}</b>', tb)])
    at = Table(ad, colWidths=[20*mm, pw-50*mm, 30*mm])
    at.setStyle(TableStyle([('BACKGROUND',(0,0),(-1,0),BL),('BOX',(0,0),(-1,-1),0.5,HexColor('#8EAADB')),
        ('INNERGRID',(0,0),(-1,-1),0.3,HexColor('#D6E4F0')),('BACKGROUND',(0,-1),(-1,-1),HexColor('#E2EFDA')),
        ('TOPPADDING',(0,0),(-1,-1),3),('BOTTOMPADDING',(0,0),(-1,-1),3)]))
    story.extend([at, Spacer(1, 5*mm)])
    
    # PASSIF table
    story.append(Paragraph("<b>PASSIF</b>", ParagraphStyle('hp', fontName='Helvetica-Bold', fontSize=11, textColor=RD, spaceAfter=2*mm)))
    pd = [[Paragraph(h, hw) for h in ['N° Compte', 'Libellé', 'Montant (FCFA)']]]
    for num, lib, val in passif_rows:
        if val > 0: pd.append([Paragraph(num, tc), Paragraph(lib, tl), Paragraph(fmt(val), tr_s)])
    pd.append([Paragraph('', tc), Paragraph('<b>TOTAL PASSIF</b>', ParagraphStyle('tp', fontName='Helvetica-Bold', fontSize=9)),
               Paragraph(f'<b>{fmt(t_passif)}</b>', tb)])
    pt = Table(pd, colWidths=[20*mm, pw-50*mm, 30*mm])
    pt.setStyle(TableStyle([('BACKGROUND',(0,0),(-1,0),RD),('BOX',(0,0),(-1,-1),0.5,HexColor('#EF9A9A')),
        ('INNERGRID',(0,0),(-1,-1),0.3,HexColor('#FFCDD2')),('BACKGROUND',(0,-1),(-1,-1),HexColor('#FDE8E8')),
        ('TOPPADDING',(0,0),(-1,-1),3),('BOTTOMPADDING',(0,0),(-1,-1),3)]))
    story.extend([pt, Spacer(1, 5*mm)])
    
    # Equilibre
    equil = "EQUILIBRE ✓" if abs(t_actif - t_passif) < 1 else f"ECART: {fmt(abs(t_actif - t_passif))}"
    story.append(Paragraph(f"<b>{equil}</b> — Total Actif: {fmt(t_actif)} F | Total Passif: {fmt(t_passif)} F",
        ParagraphStyle('eq', fontName='Helvetica-Bold', fontSize=10, textColor=BG, alignment=TA_CENTER, spaceAfter=6*mm)))
    
    story.append(Paragraph("Signature du Comptable : ___________________          Signature DG : ___________________",
        ParagraphStyle('sig', fontSize=8, textColor=HexColor('#999'), spaceAfter=4*mm)))
    story.append(Paragraph("Ce bilan est établi conformément au Plan Comptable SYSCOHADA révisé applicable en Côte d'Ivoire.",
        ParagraphStyle('legal', fontSize=7, textColor=HexColor('#bbb'), alignment=TA_CENTER)))
    
    doc.build(story)
    return send_file(output, as_attachment=True, download_name=f"Bilan_SYSCOHADA_{exercice}.pdf")

# ======================== RAPPORT DE CAISSE HEBDOMADAIRE ========================

@app.route('/comptabilite/rapport-caisse')
@permission_required('comptabilite')
def rapport_caisse():
    conn = _gdb()
    reports = [dict(r) for r in conn.execute(
        "SELECT * FROM weekly_cash_reports ORDER BY created_at DESC LIMIT 50").fetchall()]
    conn.close()
    return render_template('rapport_caisse.html', page='comptabilite', reports=reports)

@app.route('/comptabilite/rapport-caisse/new', methods=['GET', 'POST'])
@permission_required('comptabilite_edit')
def rapport_caisse_new():
    if request.method == 'POST':
        items = []
        i = 1
        while request.form.get(f'date_{i}'):
            items.append({
                'n': i, 'date': request.form.get(f'date_{i}',''),
                'description': request.form.get(f'desc_{i}',''),
                'credit': float(request.form.get(f'credit_{i}',0) or 0),
                'pc': request.form.get(f'pc_{i}',''),
                'debit': float(request.form.get(f'debit_{i}',0) or 0),
            })
            i += 1
        total_c = sum(it['credit'] for it in items)
        total_d = sum(it['debit'] for it in items)
        
        conn = _gdb()
        conn.execute("""INSERT INTO weekly_cash_reports 
            (agent_name, matricule, report_number, week_start, week_end, items_json,
             total_credit, total_debit, reste_caisse, deposit_date, created_by)
            VALUES (?,?,?,?,?,?,?,?,?,?,?)""",
            (request.form.get('agent_name',''), request.form.get('matricule',''),
             request.form.get('report_number',''), request.form.get('week_start',''),
             request.form.get('week_end',''), json.dumps(items),
             total_c, total_d, total_c - total_d,
             request.form.get('deposit_date',''), session['user_id']))
        conn.commit(); conn.close()
        flash("Rapport de caisse créé", "success")
        return redirect('/comptabilite/rapport-caisse')
    return render_template('rapport_caisse_new.html', page='comptabilite')

@app.route('/comptabilite/rapport-caisse/import', methods=['POST'])
@permission_required('comptabilite_edit')
def rapport_caisse_import():
    if 'file' not in request.files:
        flash("Fichier requis", "error"); return redirect('/comptabilite/rapport-caisse')
    f = request.files['file']
    
    import openpyxl
    wb = openpyxl.load_workbook(f, data_only=True)
    ws = wb.active
    
    agent = str(ws.cell(2,1).value or '').replace("Nom et Prénom De L'agent:", '').replace("Nom et Prénom De L\u2019agent:", '').strip().lstrip(': ')
    matricule = str(ws.cell(3,1).value or '').replace('N⁰ Matricule:', '').strip()
    report_num = str(ws.cell(4,1).value or '').replace('N⁰ Rapport:', '').strip()
    report_date = str(ws.cell(5,1).value or '').replace('Rapport Du:', '').strip()
    
    items = []
    for row in ws.iter_rows(min_row=7, max_row=ws.max_row, values_only=True):
        if not row[0] or str(row[0]).strip() == '': continue
        try: int(str(row[0]).strip())
        except: continue
        desc = str(row[2] or '').strip()
        credit = float(row[3] or 0) if row[3] and str(row[3]).replace('.','').replace('-','').isdigit() else 0
        pc = str(row[4] or '').strip()
        debit = float(row[5] or 0) if row[5] and str(row[5]).replace('.','').replace('-','').isdigit() else 0
        date_val = str(row[1] or '')[:10]
        if desc or credit > 0 or debit > 0:
            items.append({'n': len(items)+1, 'date': date_val, 'description': desc,
                         'credit': credit, 'pc': pc, 'debit': debit})
    
    total_c = sum(it['credit'] for it in items)
    total_d = sum(it['debit'] for it in items)
    
    conn = _gdb()
    conn.execute("""INSERT INTO weekly_cash_reports 
        (agent_name, matricule, report_number, week_start, items_json,
         total_credit, total_debit, reste_caisse, created_by)
        VALUES (?,?,?,?,?,?,?,?,?)""",
        (agent, matricule, report_num, report_date, json.dumps(items),
         total_c, total_d, total_c - total_d, session['user_id']))
    conn.commit(); conn.close()
    
    flash(f"Rapport importé — {agent} — {len(items)} lignes, Crédit: {total_c:,.0f}, Débit: {total_d:,.0f}", "success")
    return redirect('/comptabilite/rapport-caisse')

@app.route('/comptabilite/rapport-caisse/view/<int:rid>')
@permission_required('comptabilite')
def rapport_caisse_view(rid):
    conn = _gdb()
    r = conn.execute("SELECT * FROM weekly_cash_reports WHERE id=?", (rid,)).fetchone()
    conn.close()
    if not r: flash("Non trouvé","error"); return redirect('/comptabilite/rapport-caisse')
    report = dict(r)
    report_items = json.loads(report.get('items_json','[]') or '[]')
    return render_template('rapport_caisse_view.html', page='comptabilite', report=report, report_items=report_items)

@app.route('/comptabilite/rapport-caisse/edit/<int:rid>', methods=['GET', 'POST'])
@permission_required('comptabilite_edit')
def rapport_caisse_edit(rid):
    conn = _gdb()
    r = conn.execute("SELECT * FROM weekly_cash_reports WHERE id=?", (rid,)).fetchone()
    if not r: flash("Non trouvé","error"); return redirect('/comptabilite/rapport-caisse')
    report = dict(r)
    
    if request.method == 'POST':
        items = []
        i = 1
        while request.form.get(f'date_{i}'):
            items.append({
                'n': i, 'date': request.form.get(f'date_{i}',''),
                'description': request.form.get(f'desc_{i}',''),
                'credit': float(request.form.get(f'credit_{i}',0) or 0),
                'pc': request.form.get(f'pc_{i}',''),
                'debit': float(request.form.get(f'debit_{i}',0) or 0),
            })
            i += 1
        total_c = sum(it['credit'] for it in items)
        total_d = sum(it['debit'] for it in items)
        
        conn.execute("""UPDATE weekly_cash_reports SET agent_name=?, matricule=?, report_number=?,
            week_start=?, items_json=?, total_credit=?, total_debit=?, reste_caisse=?, deposit_date=?
            WHERE id=?""",
            (request.form.get('agent_name',''), request.form.get('matricule',''),
             request.form.get('report_number',''), request.form.get('week_start',''),
             json.dumps(items), total_c, total_d, total_c - total_d,
             request.form.get('deposit_date',''), rid))
        conn.commit(); conn.close()
        flash("Rapport modifié","success")
        return redirect(f'/comptabilite/rapport-caisse/view/{rid}')
    
    conn.close()
    report['items'] = json.loads(report.get('items_json','[]') or '[]')
    return render_template('rapport_caisse_edit.html', page='comptabilite', report=report, report_items=report['items'])

@app.route('/comptabilite/rapport-caisse/pdf/<int:rid>')
@permission_required('comptabilite')
def rapport_caisse_pdf(rid):
    conn = _gdb()
    r = conn.execute("SELECT * FROM weekly_cash_reports WHERE id=?", (rid,)).fetchone()
    conn.close()
    if not r: flash("Non trouvé","error"); return redirect('/comptabilite/rapport-caisse')
    report = dict(r)
    items = json.loads(report.get('items_json','[]') or '[]')
    
    from reportlab.lib.pagesizes import A4
    from reportlab.lib.units import mm
    from reportlab.lib.colors import HexColor, white
    from reportlab.platypus import SimpleDocTemplate, Paragraph, Spacer, Table, TableStyle
    from reportlab.lib.styles import ParagraphStyle
    from reportlab.lib.enums import TA_CENTER, TA_RIGHT, TA_LEFT
    
    output = os.path.join(app.config['UPLOAD_FOLDER'], f'rapport_caisse_{rid}.pdf')
    os.makedirs(os.path.dirname(output), exist_ok=True)
    doc = SimpleDocTemplate(output, pagesize=A4, leftMargin=12*mm, rightMargin=12*mm, topMargin=10*mm, bottomMargin=10*mm)
    
    BG = HexColor('#44546A'); BL = HexColor('#4472C4'); WH = white; BK = HexColor('#222')
    hw = ParagraphStyle('hw', fontName='Helvetica-Bold', fontSize=8, textColor=WH, alignment=TA_CENTER)
    tc = ParagraphStyle('tc', fontSize=8, alignment=TA_CENTER, textColor=BK)
    tr = ParagraphStyle('tr', fontSize=8, alignment=TA_RIGHT, textColor=BK)
    tb = ParagraphStyle('tb', fontName='Helvetica-Bold', fontSize=8, alignment=TA_RIGHT, textColor=BK)
    
    story = []
    pw = 186*mm
    
    story.append(Paragraph("<b>RAPPORT CAISSE HEBDOMADAIRE DES DEPENSES</b>",
        ParagraphStyle('t', fontName='Helvetica-Bold', fontSize=14, textColor=BG, alignment=TA_CENTER, spaceAfter=6*mm)))
    
    info = f"Agent : <b>{report['agent_name']}</b>  |  Matricule : {report['matricule']}  |  {report['report_number']}  |  {report['week_start']}"
    story.append(Paragraph(info, ParagraphStyle('i', fontSize=9, textColor=BK, alignment=TA_LEFT, spaceAfter=4*mm)))
    
    hdrs = ['N°', 'DATE', 'DESCRIPTION', 'CREDITE', 'N° P.C', 'DEBITE']
    cw = [8*mm, 20*mm, 68*mm, 24*mm, 30*mm, 24*mm]
    td = [[Paragraph(h, hw) for h in hdrs]]
    
    for it in items:
        td.append([
            Paragraph(str(it['n']), tc), Paragraph(str(it.get('date',''))[:10], tc),
            Paragraph(it.get('description',''), ParagraphStyle('d', fontSize=7, textColor=BK)),
            Paragraph(f"{it['credit']:,.0f}" if it['credit'] else '', tr),
            Paragraph(it.get('pc',''), ParagraphStyle('pc', fontSize=6, textColor=BK, alignment=TA_CENTER)),
            Paragraph(f"{it['debit']:,.0f}" if it['debit'] else '', tr),
        ])
    
    # Totals row
    td.append([Paragraph('', tc), Paragraph('Reste En Caisse:', ParagraphStyle('rc', fontName='Helvetica-Bold', fontSize=8, textColor=BK)),
        Paragraph(f"{report['reste_caisse']:,.0f}", tb),
        Paragraph(f"{report['total_credit']:,.0f}", tb),
        Paragraph('<b>TOTAL</b>', ParagraphStyle('tt', fontName='Helvetica-Bold', fontSize=8, textColor=BK, alignment=TA_CENTER)),
        Paragraph(f"{report['total_debit']:,.0f}", tb)])
    
    t = Table(td, colWidths=cw, repeatRows=1)
    sc = [('BACKGROUND', (0,0), (-1,0), BL), ('TEXTCOLOR', (0,0), (-1,0), WH),
          ('BOX', (0,0), (-1,-1), 0.5, HexColor('#8EAADB')),
          ('INNERGRID', (0,0), (-1,-1), 0.3, HexColor('#B4C6E7')),
          ('ALIGN', (0,0), (-1,-1), 'CENTER'), ('VALIGN', (0,0), (-1,-1), 'MIDDLE'),
          ('TOPPADDING', (0,0), (-1,-1), 3), ('BOTTOMPADDING', (0,0), (-1,-1), 3),
          ('BACKGROUND', (0,-1), (-1,-1), HexColor('#E2EFDA'))]
    for i in range(2, len(td)-1, 2):
        sc.append(('BACKGROUND', (0,i), (-1,i), HexColor('#F2F2F2')))
    t.setStyle(TableStyle(sc))
    story.append(t)
    
    story.append(Spacer(1, 6*mm))
    story.append(Paragraph(f"Date de Dépôt Rapport Comptabilité : {report.get('deposit_date','')}", ParagraphStyle('ft', fontSize=8, textColor=HexColor('#888'))))
    story.append(Spacer(1, 4*mm))
    story.append(Paragraph("Signature Et Cachet Caisse :", ParagraphStyle('ft2', fontSize=8, textColor=HexColor('#888'))))
    
    doc.build(story)
    return send_file(output, as_attachment=True, download_name=f"Rapport_Caisse_{report['agent_name'].replace(' ','_')}.pdf")

@app.route('/comptabilite/rapport-caisse/excel/<int:rid>')
@permission_required('comptabilite')
def rapport_caisse_excel(rid):
    conn = _gdb()
    r = conn.execute("SELECT * FROM weekly_cash_reports WHERE id=?", (rid,)).fetchone()
    conn.close()
    if not r: flash("Non trouvé","error"); return redirect('/comptabilite/rapport-caisse')
    report = dict(r)
    items = json.loads(report.get('items_json','[]') or '[]')
    
    import openpyxl
    from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Rapport Caisse"
    
    ws.merge_cells('A1:F1'); ws['A1'] = 'RAPPORT CAISSE HEBDOMADAIRE DES DEPENSES'
    ws['A1'].font = Font(bold=True, size=14); ws['A1'].alignment = Alignment(horizontal='center')
    ws['A2'] = f"Nom et Prénom De L'agent: {report['agent_name']}"
    ws['A3'] = f"N° Matricule: {report['matricule']}"
    ws['A4'] = f"N° Rapport: {report['report_number']}"
    ws['A5'] = f"Rapport Du: {report['week_start']}"
    
    hdrs = ['N°', 'DATE', 'DESCRIPTION', 'CREDITE', 'N° P.C', 'DEBITE']
    hfill = PatternFill(start_color='4472C4', end_color='4472C4', fill_type='solid')
    hfont = Font(bold=True, color='FFFFFF')
    for i, h in enumerate(hdrs, 1):
        c = ws.cell(6, i, h); c.font = hfont; c.fill = hfill; c.alignment = Alignment(horizontal='center')
    
    for idx, it in enumerate(items):
        row = 7 + idx
        ws.cell(row, 1, it['n'])
        ws.cell(row, 2, it.get('date','')[:10])
        ws.cell(row, 3, it.get('description',''))
        if it['credit']: ws.cell(row, 4, it['credit'])
        ws.cell(row, 5, it.get('pc',''))
        if it['debit']: ws.cell(row, 6, it['debit'])
    
    tr = 7 + len(items)
    ws.cell(tr, 2, 'Reste En Caisse:').font = Font(bold=True)
    ws.cell(tr, 3, report['reste_caisse']).font = Font(bold=True)
    ws.cell(tr, 4, report['total_credit']).font = Font(bold=True)
    ws.cell(tr, 5, 'TOTAL').font = Font(bold=True)
    ws.cell(tr, 6, report['total_debit']).font = Font(bold=True)
    
    ws.column_dimensions['A'].width = 5; ws.column_dimensions['B'].width = 14
    ws.column_dimensions['C'].width = 45; ws.column_dimensions['D'].width = 14
    ws.column_dimensions['E'].width = 22; ws.column_dimensions['F'].width = 14
    
    output = os.path.join(app.config['UPLOAD_FOLDER'], f'rapport_caisse_{rid}.xlsx')
    os.makedirs(os.path.dirname(output), exist_ok=True)
    wb.save(output)
    return send_file(output, as_attachment=True, download_name=f"Rapport_Caisse_{report['agent_name'].replace(' ','_')}.xlsx")


# ======================== RAPPORTS DE VISITE ========================

@app.route('/visites')
@permission_required('visites')
def visites_page():
    tab = request.args.get('tab', 'en_attente')
    visits = get_visit_reports(tab if tab != 'all' else None)
    v_stats = get_visit_stats()
    return render_template('visites.html', page='visites', tab=tab, visits=visits, v_stats=v_stats)

@app.route('/visites/new', methods=['GET', 'POST'])
@permission_required('visites')
def visites_new():
    if request.method == 'POST':
        client_id = request.form.get('client_id', '')
        client_id = int(client_id) if client_id else None
        client_name = request.form.get('client_name', '').strip()
        
        if client_id and not client_name:
            c = get_client_by_id(client_id)
            if c: client_name = c['name']
        
        create_visit_report(
            client_id, client_name,
            request.form.get('site_name', ''),
            request.form.get('site_address', ''),
            request.form.get('site_location', ''),
            request.form.get('contact_name', ''),
            request.form.get('contact_tel', ''),
            request.form.get('visit_date', ''),
            request.form.get('needs', ''),
            request.form.get('observations', ''),
            request.form.get('equipment', ''),
            session['user_id']
        )
        user = get_user_by_id(session['user_id'])
        log_activity(session['user_id'], user['full_name'] if user else '?',
                    'Visite', f"Rapport de visite créé — {client_name}", request.remote_addr)
        flash("Rapport de visite créé — En attente de proforma", "success")
        return redirect(url_for('visites_page'))
    
    clients = get_all_clients()
    return render_template('visite_new.html', page='visites', clients=clients)

@app.route('/visites/<int:vid>')
@login_required
def visite_detail(vid):
    visit = get_visit_by_id(vid)
    if not visit:
        flash("Rapport non trouvé", "error")
        return redirect(url_for('visites_page'))
    return render_template('visite_detail.html', page='visites', visit=visit)

@app.route('/visites/proforma/<int:vid>', methods=['POST'])
@permission_required('proforma')
def visites_proforma(vid):
    ref = request.form.get('proforma_ref', '').strip()
    amount = float(request.form.get('proforma_amount', 0) or 0)
    update_visit_proforma(vid, ref, amount, session['user_id'])
    user = get_user_by_id(session['user_id'])
    log_activity(session['user_id'], user['full_name'] if user else '?',
                'Proforma', f"Proforma {ref} envoyé pour visite #{vid}", request.remote_addr)
    flash(f"Proforma {ref} envoyé", "success")
    return redirect(url_for('visites_page'))


# ======================== LANGUE ========================

@app.route('/lang/<lang>')
def set_language(lang):
    if lang in ('fr', 'en'):
        session['lang'] = lang
    return redirect(request.referrer or url_for('dashboard'))


# ======================== EXPORT STATS ========================

@app.route('/export/stats')
@permission_required('admin')
def export_stats():
    """Exporte les statistiques en PDF avec graphiques."""
    from rapport_core import _generate_chart_image
    
    stats = get_dashboard_stats()
    monthly = get_client_monthly_stats()
    
    # Générer un rapport simple avec reportlab
    from reportlab.lib.pagesizes import A4
    from reportlab.lib.units import mm
    from reportlab.platypus import SimpleDocTemplate, Paragraph, Spacer, Table, TableStyle
    from reportlab.lib.styles import ParagraphStyle
    from reportlab.lib.enums import TA_CENTER
    from reportlab.lib.colors import HexColor
    
    export_dir = os.path.join(app.config['UPLOAD_FOLDER'], 'export')
    os.makedirs(export_dir, exist_ok=True)
    output = os.path.join(export_dir, 'statistiques_ramya.pdf')
    
    doc = SimpleDocTemplate(output, pagesize=A4)
    story = []
    
    title_style = ParagraphStyle('title', fontSize=18, alignment=TA_CENTER,
                                  textColor=HexColor('#1a3a5c'), spaceAfter=20)
    h2_style = ParagraphStyle('h2', fontSize=14, textColor=HexColor('#1a3a5c'),
                               spaceAfter=10, spaceBefore=20)
    normal = ParagraphStyle('normal', fontSize=11, spaceAfter=6)
    
    story.append(Paragraph("RAMYA TECHNOLOGIE & INNOVATION", title_style))
    story.append(Paragraph("Rapport Statistique", title_style))
    story.append(Spacer(1, 10*mm))
    
    story.append(Paragraph("Vue d'ensemble", h2_style))
    story.append(Paragraph(f"Rapports traités : {stats['total_jobs']}", normal))
    story.append(Paragraph(f"En attente d'envoi : {stats['pending_jobs']}", normal))
    story.append(Paragraph(f"Envoyés : {stats['sent_jobs']}", normal))
    story.append(Paragraph(f"Clients : {stats['total_clients']}", normal))
    story.append(Paragraph(f"Utilisateurs : {stats['total_users']}", normal))
    
    if monthly:
        story.append(Paragraph("Détail par client et par mois", h2_style))
        for client, months in monthly.items():
            story.append(Paragraph(f"<b>{client}</b>", normal))
            data = [['Mois', 'Rapports', 'Employés', 'Envoyés', 'En attente']]
            for month, m_stats in sorted(months.items()):
                data.append([month, m_stats['count'], m_stats['employees'],
                           m_stats['sent'], m_stats['pending']])
            t = Table(data, colWidths=[35*mm, 25*mm, 25*mm, 25*mm, 25*mm])
            t.setStyle(TableStyle([
                ('BACKGROUND', (0,0), (-1,0), HexColor('#1a3a5c')),
                ('TEXTCOLOR', (0,0), (-1,0), HexColor('#ffffff')),
                ('FONTSIZE', (0,0), (-1,-1), 9),
                ('GRID', (0,0), (-1,-1), 0.5, HexColor('#cccccc')),
                ('TOPPADDING', (0,0), (-1,-1), 4),
                ('BOTTOMPADDING', (0,0), (-1,-1), 4),
            ]))
            story.append(t)
            story.append(Spacer(1, 5*mm))
    
    story.append(Spacer(1, 10*mm))
    story.append(Paragraph(f"Généré le {datetime.now().strftime('%d/%m/%Y à %H:%M')}", normal))
    
    doc.build(story)
    
    user = get_user_by_id(session['user_id'])
    log_activity(session['user_id'], user['full_name'] if user else '?',
                'Export', 'Export statistiques PDF', request.remote_addr)
    
    return send_file(output, as_attachment=True,
                    download_name=f"stats_ramya_{datetime.now().strftime('%Y%m')}.pdf")


# ======================== DEVIS / PROFORMA ========================

@app.route('/devis')
@permission_required('proforma')
def devis_page():
    tab = request.args.get('tab', 'all')
    d_stats = get_devis_stats()
    devis_list = get_all_devis(tab if tab in ('devis', 'proforma') else None)
    return render_template('devis.html', page='devis', tab=tab, devis_list=devis_list, d_stats=d_stats)

@app.route('/devis/new', methods=['GET', 'POST'])
@permission_required('proforma_edit')
def devis_new():
    if request.method == 'POST':
        items = []
        i = 1
        while request.form.get(f'item_{i}_designation'):
            items.append({
                'num': i,
                'designation': request.form.get(f'item_{i}_designation', ''),
                'detail': request.form.get(f'item_{i}_detail', ''),
                'qty': int(request.form.get(f'item_{i}_qty', 1) or 1),
                'prix': float(request.form.get(f'item_{i}_prix', 0) or 0),
                'remise': float(request.form.get(f'item_{i}_remise', 0) or 0),
            })
            i += 1
        
        total_ht = sum(it['qty'] * it['prix'] - it['remise'] for it in items)
        main_oeuvre = float(request.form.get('main_oeuvre', 0) or 0)
        petites_fourn = float(request.form.get('petites_fournitures', 0) or 0)
        remise_glob = float(request.form.get('remise', 0) or 0)
        total_ttc = total_ht + petites_fourn - remise_glob
        
        client_id = request.form.get('client_id', '')
        client_id = int(client_id) if client_id else None
        
        did, ref = create_devis(
            client_id, request.form.get('client_name', ''),
            request.form.get('client_code', ''),
            request.form.get('contact_commercial', ''),
            request.form.get('objet', ''),
            json.dumps(items), total_ht, petites_fourn, total_ttc,
            main_oeuvre, remise_glob,
            request.form.get('notes', ''),
            session['user_id'],
            request.form.get('doc_type', 'devis')
        )
        
        user = get_user_by_id(session['user_id'])
        log_activity(session['user_id'], user['full_name'] if user else '?',
                    'Devis', f"{ref} créé pour {request.form.get('client_name', '')}", request.remote_addr)
        flash(f"Devis {ref} créé — {total_ttc:,.0f} FCFA TTC", "success")
        return redirect(url_for('devis_page'))
    
    clients = get_all_clients()
    from models import db_get_all
    stock_items = db_get_all('stock_items', order='name ASC')
    return render_template('devis_new.html', page='devis', clients=clients, stock_items=stock_items)

@app.route('/devis/pdf/<int:did>')
@login_required
def devis_pdf(did):
    devis = get_devis_by_id(did)
    if not devis:
        flash("Devis non trouvé", "error")
        return redirect(url_for('devis_page'))
    
    export_dir = os.path.join(app.config['UPLOAD_FOLDER'], 'devis')
    os.makedirs(export_dir, exist_ok=True)
    output = os.path.join(export_dir, f"{devis['reference']}.pdf")
    
    devis['date'] = devis.get('created_at', '')[:10]
    generate_devis_pdf(devis, output)
    
    return send_file(output, as_attachment=True, download_name=f"{devis['reference']}.pdf")

@app.route('/devis/status/<int:did>/<status>')
@permission_required('proforma_edit')
def devis_status(did, status):
    if status in ('brouillon', 'envoye', 'accepte', 'refuse'):
        update_devis_status(did, status)
        # Stock deduction when devis is accepted
        if status == 'accepte':
            devis = get_devis_by_id(did)
            if devis:
                items = json.loads(devis.get('items_json', '[]')) if isinstance(devis.get('items_json'), str) else devis.get('items_json', [])
                conn = _gdb()
                for item in items:
                    designation = item.get('designation', '')
                    qty = int(item.get('qty', 0) or 0)
                    if designation and qty > 0:
                        stock_item = conn.execute("SELECT id, quantity FROM stock_items WHERE name=? OR reference=?",
                            (designation, designation)).fetchone()
                        if stock_item:
                            new_qty = max(0, stock_item['quantity'] - qty)
                            conn.execute("UPDATE stock_items SET quantity=? WHERE id=?", (new_qty, stock_item['id']))
                            conn.execute("""INSERT INTO stock_movements (item_id, movement_type, quantity, reference, notes, created_by)
                                VALUES (?, 'sortie', ?, ?, ?, ?)""",
                                (stock_item['id'], qty, devis.get('reference',''), f"Devis {devis['reference']} accepté", session.get('user_id')))
                conn.commit(); conn.close()
                flash(f"Articles déduits du stock", "info")
        flash(f"Statut mis à jour : {status}", "success")
    return redirect(url_for('devis_page'))

@app.route('/devis/edit/<int:did>', methods=['GET', 'POST'])
@permission_required('proforma_edit')
def devis_edit(did):
    devis = get_devis_by_id(did)
    if not devis: flash("Devis non trouvé", "error"); return redirect(url_for('devis_page'))
    if request.method == 'POST':
        items = []
        i = 1
        while request.form.get(f'item_{i}_designation'):
            items.append({
                'num': i, 'designation': request.form.get(f'item_{i}_designation', ''),
                'detail': request.form.get(f'item_{i}_detail', ''),
                'qty': int(request.form.get(f'item_{i}_qty', 1) or 1),
                'prix': float(request.form.get(f'item_{i}_prix', 0) or 0),
                'remise': float(request.form.get(f'item_{i}_remise', 0) or 0),
            })
            i += 1
        total_ht = sum(it['qty'] * it['prix'] - it['remise'] for it in items)
        main_oeuvre = float(request.form.get('main_oeuvre', 0) or 0)
        petites_fourn = float(request.form.get('petites_fournitures', 0) or 0)
        remise_glob = float(request.form.get('remise', 0) or 0)
        total_ttc = total_ht + petites_fourn - remise_glob
        
        conn = _gdb()
        conn.execute("""UPDATE devis SET client_name=?, client_code=?, contact_commercial=?, objet=?,
            items_json=?, total_ht=?, petites_fournitures=?, total_ttc=?, main_oeuvre=?, remise=?, notes=?
            WHERE id=?""", (request.form.get('client_name',''), request.form.get('client_code',''),
            request.form.get('contact_commercial',''), request.form.get('objet',''),
            json.dumps(items), total_ht, petites_fourn, total_ttc, main_oeuvre, remise_glob,
            request.form.get('notes',''), did))
        conn.commit(); conn.close()
        flash("Devis modifié", "success"); return redirect(url_for('devis_page'))
    
    items = json.loads(devis.get('items_json', '[]')) if isinstance(devis.get('items_json'), str) else []
    clients = get_all_clients()
    from models import db_get_all
    stock_items = db_get_all('stock_items', order='name ASC')
    return render_template('devis_edit.html', page='devis', devis=devis, items=items, clients=clients, stock_items=stock_items)

@app.route('/devis/delete/<int:did>')
@permission_required('proforma_edit')
def devis_delete(did):
    conn = _gdb()
    conn.execute("DELETE FROM devis WHERE id=?", (did,))
    conn.commit(); conn.close()
    flash("Devis supprimé", "success"); return redirect(url_for('devis_page'))

@app.route('/devis/duplicate/<int:did>')
@permission_required('proforma_edit')
def devis_duplicate(did):
    devis = get_devis_by_id(did)
    if devis:
        from models import create_devis
        new_id, new_ref = create_devis(
            devis.get('client_id'), devis['client_name'], devis.get('client_code',''),
            devis.get('contact_commercial',''), devis.get('objet',''),
            devis.get('items_json','[]'), devis['total_ht'], devis.get('petites_fournitures',0),
            devis['total_ttc'], devis.get('main_oeuvre',0), devis.get('remise',0),
            devis.get('notes',''), session['user_id'], devis.get('doc_type','devis'))
        flash(f"Devis dupliqué → {new_ref}", "success")
    return redirect(url_for('devis_page'))

@app.route('/devis/preview/<int:did>')
@login_required
def devis_preview(did):
    devis = get_devis_by_id(did)
    if not devis: return "Non trouvé", 404
    export_dir = os.path.join(app.config['UPLOAD_FOLDER'], 'devis')
    os.makedirs(export_dir, exist_ok=True)
    output = os.path.join(export_dir, f"preview_{devis['reference']}.pdf")
    devis['date'] = devis.get('created_at', '')[:10]
    generate_devis_pdf(devis, output)
    return send_file(output, as_attachment=False, download_name=f"{devis['reference']}.pdf")


# ======================== RH MODULE ========================

@app.route('/rh')
@permission_required('fichiers')
def rh_dashboard():
    emp_stats = get_employee_stats()
    return render_template('rh_dashboard.html', page='rh', emp_stats=emp_stats)

@app.route('/rh/personnel')
@permission_required('fichiers')
def rh_personnel():
    employees = get_all_employees()
    return render_template('rh_personnel.html', page='personnel', employees=employees)

@app.route('/rh/personnel/add', methods=['GET', 'POST'])
@permission_required('fichiers')
def rh_personnel_add():
    if request.method == 'POST':
        from models import get_db
        fields = {}
        for key in ['first_name','last_name','matricule','email','tel','position','department',
                     'hire_date','contract_type','insurance','insurance_number',
                     'emergency_contact','emergency_tel','code_rh','birth_date','gender','blood_type',
                     'birth_place','birth_city','civil_status','nationality','religion',
                     'id_type','id_expiry','id_place','resident','address','education_level',
                     'work_location','bank_account','bank_name_emp','bank_holder',
                     'fiscal_code','hourly_rate','facebook','linkedin','skype',
                     'direction','email_signature','other_info','status']:
            fields[key] = request.form.get(key, '')
        fields['salary'] = float(request.form.get('salary', 0) or 0)
        try:
            create_employee(**fields)
        except Exception as e:
            flash(f"Erreur : {str(e)}", "error")
            return redirect('/rh/personnel/add')
        # Get new employee ID for photo
        conn = get_db()
        new_emp = conn.execute("SELECT id FROM employees ORDER BY id DESC LIMIT 1").fetchone()
        conn.close()
        if new_emp and 'photo' in request.files:
            f = request.files['photo']
            if f and f.filename:
                from werkzeug.utils import secure_filename
                ext = os.path.splitext(f.filename)[1].lower()
                if ext in ('.jpg','.jpeg','.png','.webp'):
                    photo_dir = os.path.join(app.config['UPLOAD_FOLDER'], 'photos')
                    os.makedirs(photo_dir, exist_ok=True)
                    fname = f"emp_{new_emp['id']}{ext}"
                    f.save(os.path.join(photo_dir, fname))
                    update_employee(new_emp['id'], photo=fname)
        user = get_user_by_id(session['user_id'])
        log_activity(session['user_id'], user['full_name'] if user else '?', 'RH',
                    f"Employé ajouté: {request.form.get('first_name','')} {request.form.get('last_name','')}", request.remote_addr)
        flash("Employé ajouté", "success")
        return redirect(url_for('rh_personnel'))
    return render_template('rh_personnel_add.html', page='personnel')

@app.route('/rh/personnel/edit/<int:eid>', methods=['GET', 'POST'])
@permission_required('fichiers')
def rh_personnel_edit(eid):
    emp = get_employee_by_id(eid)
    if not emp:
        flash("Employé non trouvé", "error")
        return redirect(url_for('rh_personnel'))
    if request.method == 'POST':
        fields = {}
        for key in ['first_name','last_name','matricule','email','tel','position','department',
                     'hire_date','contract_type','insurance','insurance_number',
                     'emergency_contact','emergency_tel','code_rh','birth_date','gender','blood_type',
                     'birth_place','birth_city','civil_status','nationality','religion',
                     'id_type','id_expiry','id_place','resident','address','education_level',
                     'work_location','bank_account','bank_name_emp','bank_holder',
                     'fiscal_code','hourly_rate','facebook','linkedin','skype',
                     'direction','email_signature','other_info','status']:
            fields[key] = request.form.get(key, '')
        fields['salary'] = float(request.form.get('salary', 0) or 0)
        update_employee(eid, **fields)
        # Photo
        if 'photo' in request.files and request.files['photo'].filename:
            photo = request.files['photo']
            fname = f"emp_{eid}_{secure_filename(photo.filename)}"
            photo_dir = os.path.join(app.config['UPLOAD_FOLDER'], 'photos')
            os.makedirs(photo_dir, exist_ok=True)
            photo.save(os.path.join(photo_dir, fname))
            update_employee(eid, photo=fname)
        flash("Employé modifié", "success")
        return redirect(url_for('rh_personnel'))
    return render_template('rh_personnel_edit.html', page='personnel', emp=emp)

@app.route('/rh/conges')
@permission_required('fichiers')
def rh_conges():
    tab = request.args.get('tab', 'en_attente')
    leaves = get_leaves(tab if tab != 'all' else None)
    employees = get_all_employees()
    return render_template('rh_conges.html', page='conges', tab=tab, leaves=leaves, employees=employees)

@app.route('/rh/conges/add', methods=['POST'])
@permission_required('fichiers')
def rh_conges_add():
    create_leave(
        int(request.form['employee_id']),
        request.form.get('leave_type', 'conge_annuel'),
        request.form['start_date'],
        request.form['end_date'],
        int(request.form.get('days', 0) or 0),
        request.form.get('reason', '')
    )
    flash("Demande de congé créée", "success")
    return redirect(url_for('rh_conges'))

@app.route('/rh/conges/approve/<int:lid>/<status>')
@permission_required('fichiers')
def rh_conges_status(lid, status):
    if status in ('approuve', 'refuse'):
        update_leave_status(lid, status, session.get('user_id'))
        flash(f"Congé {'approuvé' if status == 'approuve' else 'refusé'}", "success")
    return redirect(url_for('rh_conges'))

@app.route('/rh/paie')
@permission_required('fichiers')
def rh_paie():
    period = request.args.get('period', '')
    payslips = get_payslips(period if period else None)
    employees = get_all_employees()
    return render_template('rh_paie.html', page='paie', payslips=payslips, employees=employees, period=period)

@app.route('/rh/paie/add', methods=['POST'])
@permission_required('fichiers')
def rh_paie_add():
    f = lambda k: float(request.form.get(k, 0) or 0)
    base = f('base_salary')
    primes = f('bonus') + f('prime_transport') + f('prime_anciennete') + f('prime_logement') + f('prime_rendement') + f('avantages_nature')
    brut = base + f('overtime_amount') + primes + f('commission')
    retenues = f('cnps_employee') + f('insurance_amount') + f('its') + f('deductions') + f('autres_retenues') + f('avances')
    net = brut - retenues
    
    create_payslip(
        employee_id=int(request.form['employee_id']), period=request.form['period'],
        base_salary=base, bonus=f('bonus'), commission=f('commission'),
        overtime_amount=f('overtime_amount'), deductions=f('deductions'),
        insurance_amount=f('insurance_amount'), tax_amount=f('its'), net_salary=net,
        prime_transport=f('prime_transport'), prime_anciennete=f('prime_anciennete'),
        prime_logement=f('prime_logement'), prime_rendement=f('prime_rendement'),
        avantages_nature=f('avantages_nature'), cnps_employee=f('cnps_employee'),
        its=f('its'), autres_retenues=f('autres_retenues'), avances=f('avances'),
        jours_travailles=int(request.form.get('jours_travailles', 26) or 26),
        heures_travaillees=f('heures_travaillees'),
        conges_payes=int(request.form.get('conges_payes', 0) or 0),
        jours_absence=int(request.form.get('jours_absence', 0) or 0),
        mode_paiement=request.form.get('mode_paiement', 'virement'),
        cnps_employer=f('cnps_employer'),
    )
    flash(f"Bulletin créé — Net: {net:,.0f} FCFA", "success")
    return redirect(url_for('rh_paie'))

@app.route('/rh/paie/<int:pid>/edit', methods=['GET','POST'])
@permission_required('fichiers')
def rh_paie_edit(pid):
    p = get_payslip_detail_v2(pid)
    if not p: flash("Non trouvé","error"); return redirect(url_for('rh_paie'))
    if request.method == 'POST':
        f = lambda k: float(request.form.get(k, 0) or 0)
        base = f('base_salary')
        primes = f('bonus') + f('prime_transport') + f('prime_anciennete') + f('prime_logement') + f('prime_rendement') + f('avantages_nature')
        brut = base + f('overtime_amount') + primes + f('commission')
        retenues = f('cnps_employee') + f('insurance_amount') + f('its') + f('deductions') + f('autres_retenues') + f('avances')
        net = brut - retenues
        update_payslip(pid, base_salary=base, bonus=f('bonus'), commission=f('commission'),
            overtime_amount=f('overtime_amount'), deductions=f('deductions'),
            insurance_amount=f('insurance_amount'), tax_amount=f('its'), net_salary=net,
            prime_transport=f('prime_transport'), prime_anciennete=f('prime_anciennete'),
            prime_logement=f('prime_logement'), prime_rendement=f('prime_rendement'),
            avantages_nature=f('avantages_nature'), cnps_employee=f('cnps_employee'),
            its=f('its'), autres_retenues=f('autres_retenues'), avances=f('avances'),
            jours_travailles=int(request.form.get('jours_travailles', 26) or 26),
            heures_travaillees=f('heures_travaillees'),
            conges_payes=int(request.form.get('conges_payes', 0) or 0),
            jours_absence=int(request.form.get('jours_absence', 0) or 0),
            mode_paiement=request.form.get('mode_paiement', 'virement'),
            cnps_employer=f('cnps_employer'))
        flash(f"Bulletin modifié — Net: {net:,.0f} FCFA", "success")
        return redirect(url_for('rh_paie_view', pid=pid))
    employees = get_all_employees()
    return render_template('rh_paie_edit.html', page='paie', p=p, employees=employees)

@app.route('/rh/paie/<int:pid>/pdf')
@permission_required('fichiers')
def rh_paie_pdf(pid):
    """Génère le bulletin de paie PDF format CI."""
    p = get_payslip_detail_v2(pid)
    if not p: flash("Non trouvé","error"); return redirect(url_for('rh_paie'))
    
    from reportlab.lib.pagesizes import A4
    from reportlab.lib.units import mm
    from reportlab.lib.colors import HexColor
    from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer, HRFlowable
    from reportlab.lib.styles import ParagraphStyle
    from reportlab.lib.enums import TA_CENTER, TA_RIGHT, TA_LEFT
    
    output = os.path.join(app.config['UPLOAD_FOLDER'], f'bulletin_{p["period"]}_{p["employee_name"].replace(" ","_")}.pdf')
    os.makedirs(os.path.dirname(output), exist_ok=True)
    doc = SimpleDocTemplate(output, pagesize=A4, leftMargin=15*mm, rightMargin=15*mm, topMargin=12*mm, bottomMargin=12*mm)
    
    DARK = HexColor('#222222'); GREY = HexColor('#888'); LIGHT = HexColor('#f5f5f5')
    s_t = ParagraphStyle('t', fontSize=14, fontName='Helvetica-Bold', textColor=DARK, alignment=TA_CENTER)
    s_s = ParagraphStyle('s', fontSize=9, alignment=TA_CENTER, textColor=GREY)
    s_n = ParagraphStyle('n', fontSize=9, leading=12, textColor=DARK)
    s_b = ParagraphStyle('b', fontSize=9, fontName='Helvetica-Bold', textColor=DARK)
    s_h = ParagraphStyle('h', fontSize=8, fontName='Helvetica-Bold', textColor=DARK)
    s_c = ParagraphStyle('c', fontSize=8, leading=10, textColor=DARK)
    s_r = ParagraphStyle('r', fontSize=8, alignment=TA_RIGHT, textColor=DARK)
    s_rb = ParagraphStyle('rb', fontSize=8, fontName='Helvetica-Bold', alignment=TA_RIGHT, textColor=DARK)
    fmt = lambda x: f"{(x or 0):,.0f}"
    
    story = []
    story.append(Paragraph("BULLETIN DE PAIE", s_t))
    story.append(Spacer(1, 2*mm))
    story.append(HRFlowable(width="100%", thickness=1, color=DARK))
    story.append(Spacer(1, 4*mm))
    
    # === HEADER ===
    header = [[
        Paragraph("<b>EMPLOYEUR</b><br/>WannyGest<br/>Abidjan, Côte d'Ivoire<br/>RC: CI-ABJ-XXXX<br/>N° CNPS: XXXX", s_c),
        Paragraph(f"<b>EMPLOYÉ</b><br/>{p['employee_name']}<br/>Matricule: {p.get('matricule','') or '-'}<br/>N° CNPS: {p.get('insurance_number','') or '-'}<br/>Poste: {p.get('position','') or '-'}<br/>Embauche: {p.get('hire_date','') or '-'}", s_c),
        Paragraph(f"<b>PÉRIODE</b><br/>{p['period']}<br/>Jours: {p.get('jours_travailles',26)}<br/>Heures: {fmt(p.get('heures_travaillees',0))}<br/>Congés: {p.get('conges_payes',0)}<br/>Absences: {p.get('jours_absence',0)}", s_c),
    ]]
    ht = Table(header, colWidths=[60*mm, 60*mm, 55*mm])
    ht.setStyle(TableStyle([('GRID',(0,0),(-1,-1),0.5,HexColor('#ccc')),('BACKGROUND',(0,0),(-1,-1),LIGHT),
        ('TOPPADDING',(0,0),(-1,-1),6),('BOTTOMPADDING',(0,0),(-1,-1),6),('LEFTPADDING',(0,0),(-1,-1),8)]))
    story.append(ht)
    story.append(Spacer(1, 5*mm))
    
    # === GAINS ===
    g_rows = [
        [Paragraph(h, s_h) for h in ['GAINS', 'Base', 'Taux', 'Montant']],
        [Paragraph('Salaire de base', s_c), Paragraph(fmt(p.get('base_salary',0)), s_r), Paragraph('', s_r), Paragraph(fmt(p.get('base_salary',0)), s_rb)],
        [Paragraph('Heures supplémentaires', s_c), Paragraph('', s_r), Paragraph('', s_r), Paragraph(fmt(p.get('overtime_amount',0)), s_r)],
        [Paragraph('Prime de transport', s_c), Paragraph('', s_r), Paragraph('', s_r), Paragraph(fmt(p.get('prime_transport',0)), s_r)],
        [Paragraph("Prime d'ancienneté", s_c), Paragraph('', s_r), Paragraph('', s_r), Paragraph(fmt(p.get('prime_anciennete',0)), s_r)],
        [Paragraph('Prime de logement', s_c), Paragraph('', s_r), Paragraph('', s_r), Paragraph(fmt(p.get('prime_logement',0)), s_r)],
        [Paragraph('Prime de rendement / KPI', s_c), Paragraph('', s_r), Paragraph('', s_r), Paragraph(fmt(p.get('prime_rendement',0) + (p.get('bonus',0) or 0)), s_r)],
        [Paragraph('Avantages en nature', s_c), Paragraph('', s_r), Paragraph('', s_r), Paragraph(fmt(p.get('avantages_nature',0)), s_r)],
        [Paragraph('Commissions', s_c), Paragraph('', s_r), Paragraph('', s_r), Paragraph(fmt(p.get('commission',0)), s_r)],
        [Paragraph('<b>SALAIRE BRUT</b>', s_c), Paragraph('', s_r), Paragraph('', s_r), Paragraph(f"<b>{fmt(p['salaire_brut'])}</b>", s_rb)],
    ]
    cw = [65*mm, 30*mm, 30*mm, 40*mm]
    gt = Table(g_rows, colWidths=cw)
    gt.setStyle(TableStyle([('BACKGROUND',(0,0),(-1,0),LIGHT),('GRID',(0,0),(-1,-1),0.5,HexColor('#ddd')),
        ('TOPPADDING',(0,0),(-1,-1),3),('BOTTOMPADDING',(0,0),(-1,-1),3),
        ('LINEABOVE',(0,9),(-1,9),1,DARK)]))
    story.append(gt)
    story.append(Spacer(1, 3*mm))
    
    # === RETENUES ===
    r_rows = [
        [Paragraph(h, s_h) for h in ['RETENUES', 'Base', 'Taux', 'Montant']],
        [Paragraph('Cotisation CNPS (salarié)', s_c), Paragraph(fmt(p['salaire_brut']), s_r), Paragraph('6.3%', s_r), Paragraph(fmt(p.get('cnps_employee',0)), s_r)],
        [Paragraph('Assurance maladie', s_c), Paragraph('', s_r), Paragraph('', s_r), Paragraph(fmt(p.get('insurance_amount',0)), s_r)],
        [Paragraph('Impôt sur salaire (ITS)', s_c), Paragraph('', s_r), Paragraph('', s_r), Paragraph(fmt(p.get('its',0)), s_r)],
        [Paragraph('Autres déductions', s_c), Paragraph('', s_r), Paragraph('', s_r), Paragraph(fmt(p.get('deductions',0)), s_r)],
        [Paragraph('Avances / prêts', s_c), Paragraph('', s_r), Paragraph('', s_r), Paragraph(fmt(p.get('avances',0)), s_r)],
        [Paragraph('Autres retenues', s_c), Paragraph('', s_r), Paragraph('', s_r), Paragraph(fmt(p.get('autres_retenues',0)), s_r)],
        [Paragraph('<b>TOTAL RETENUES</b>', s_c), Paragraph('', s_r), Paragraph('', s_r), Paragraph(f"<b>{fmt(p['total_retenues'])}</b>", s_rb)],
    ]
    rt = Table(r_rows, colWidths=cw)
    rt.setStyle(TableStyle([('BACKGROUND',(0,0),(-1,0),LIGHT),('GRID',(0,0),(-1,-1),0.5,HexColor('#ddd')),
        ('TOPPADDING',(0,0),(-1,-1),3),('BOTTOMPADDING',(0,0),(-1,-1),3),
        ('LINEABOVE',(0,7),(-1,7),1,DARK)]))
    story.append(rt)
    story.append(Spacer(1, 4*mm))
    
    # === NET À PAYER ===
    story.append(HRFlowable(width="100%", thickness=2, color=DARK))
    net_row = [[Paragraph('<b>NET À PAYER</b>', ParagraphStyle('net', fontSize=13, fontName='Helvetica-Bold', textColor=DARK)),
                Paragraph(f'<b>{fmt(p["net_salary"])} FCFA</b>', ParagraphStyle('nv', fontSize=13, fontName='Helvetica-Bold', textColor=DARK, alignment=TA_RIGHT))]]
    nt = Table(net_row, colWidths=[100*mm, 70*mm])
    nt.setStyle(TableStyle([('TOPPADDING',(0,0),(-1,-1),8),('BOTTOMPADDING',(0,0),(-1,-1),8)]))
    story.append(nt)
    story.append(HRFlowable(width="100%", thickness=2, color=DARK))
    story.append(Spacer(1, 4*mm))
    
    # === Infos complémentaires ===
    story.append(Paragraph(f"Mode de paiement : {p.get('mode_paiement','virement')}", s_n))
    if p.get('cnps_employer',0):
        story.append(Paragraph(f"Cotisation CNPS patronale : {fmt(p['cnps_employer'])} FCFA", s_n))
    story.append(Spacer(1, 12*mm))
    
    # Signatures
    sig = [[Paragraph("<b>L'Employeur</b><br/><br/><br/>Signature et cachet", s_c),
            Paragraph(f"<b>L'Employé</b><br/><br/><br/>Lu et approuvé<br/>{p['employee_name']}", s_c)]]
    st = Table(sig, colWidths=[85*mm, 85*mm])
    story.append(st)
    story.append(Spacer(1, 8*mm))
    story.append(Paragraph("Ce bulletin de paie doit être conservé sans limitation de durée (Art. 32.4 du Code du Travail)", ParagraphStyle('f', fontSize=7, alignment=TA_CENTER, textColor=GREY)))
    
    doc.build(story)
    return send_file(output, as_attachment=True, download_name=f"Bulletin_{p['period']}_{p['employee_name']}.pdf")

@app.route('/rh/paie/<int:pid>/view')
@permission_required('fichiers')
def rh_paie_view(pid):
    p = get_payslip_detail_v2(pid)
    if not p: flash("Non trouvé","error"); return redirect(url_for('rh_paie'))
    return render_template('rh_paie_view.html', page='paie', p=p)

@app.route('/rh/paie/<int:pid>/status/<status>')
@permission_required('fichiers')
def rh_paie_status(pid, status):
    if status in ('brouillon','valide','envoye'):
        update_payslip(pid, status=status)
        if status == 'envoye':
            update_payslip(pid, sent_at=datetime.now().isoformat())
        flash(f"Statut → {status}", "success")
    return redirect(url_for('rh_paie'))

# ======================== EMPLOYEE PHOTO ========================

@app.route('/rh/personnel/photo/<int:eid>', methods=['POST'])
@permission_required('fichiers')
def rh_personnel_photo(eid):
    if 'photo' in request.files:
        f = request.files['photo']
        if f and f.filename:
            from werkzeug.utils import secure_filename
            ext = os.path.splitext(f.filename)[1].lower()
            if ext in ('.jpg','.jpeg','.png','.webp'):
                photo_dir = os.path.join(app.config['UPLOAD_FOLDER'], 'photos')
                os.makedirs(photo_dir, exist_ok=True)
                fname = f"emp_{eid}{ext}"
                f.save(os.path.join(photo_dir, fname))
                update_employee(eid, photo=fname)
                flash("Photo mise à jour", "success")
    return redirect(url_for('rh_personnel'))

@app.route('/uploads/photos/<path:filename>')
def employee_photo(filename):
    return send_from_directory(os.path.join(app.config['UPLOAD_FOLDER'], 'photos'), filename)

# ======================== IMPORT CLIENTS EXCEL ========================

@app.route('/clients/import', methods=['POST'])
@permission_required('clients')
def clients_import():
    if 'file' not in request.files:
        flash("Aucun fichier", "error"); return redirect(url_for('clients_page'))
    f = request.files['file']
    if f.filename.endswith(('.xlsx','.xls')):
        import openpyxl
        tmp = os.path.join(app.config['UPLOAD_FOLDER'], 'import_tmp.xlsx')
        f.save(tmp)
        wb = openpyxl.load_workbook(tmp, data_only=True)
        ws = wb.active
        count = 0
        for row in ws.iter_rows(min_row=2, values_only=True):
            if row[0]:
                try:
                    create_client(str(row[0]), str(row[1] or ''), str(row[2] or ''), str(row[3] or ''))
                    count += 1
                except: pass
        flash(f"{count} clients importés depuis Excel", "success")
        os.remove(tmp)
    else:
        flash("Format non supporté (xlsx uniquement)", "error")
    return redirect(url_for('clients_page'))

# ======================== CENTRE TECHNIQUE ========================

@app.route('/centre-technique')
@login_required
def tech_center():
    from models import db_get_all, db_insert, db_update
    systems = db_get_all('tech_center', order='next_maintenance ASC')
    due = get_maintenance_due()
    clients = get_all_clients()
    return render_template('tech_center.html', page='tech_center', systems=systems, due=due, clients=clients)

@app.route('/centre-technique/add', methods=['POST'])
@login_required
def tech_center_add():
    from models import db_insert
    db_insert('tech_center', client_name=request.form.get('client_name',''),
        client_id=int(request.form['client_id']) if request.form.get('client_id') else None,
        system_type=request.form.get('system_type',''),
        installation_date=request.form.get('installation_date',''),
        next_maintenance=request.form.get('next_maintenance',''),
        maintenance_interval=int(request.form.get('maintenance_interval', 90) or 90),
        notes=request.form.get('notes',''), created_by=session['user_id'])
    flash("Système ajouté au centre technique", "success")
    return redirect(url_for('tech_center'))

@app.route('/centre-technique/view/<int:sid>')
@login_required
def tech_center_view(sid):
    from models import db_get_by_id
    system = db_get_by_id('tech_center', sid)
    if not system: flash("Non trouvé","error"); return redirect(url_for('tech_center'))
    conn = _gdb()
    devis_list = [dict(r) for r in conn.execute("SELECT * FROM devis WHERE client_name=? ORDER BY created_at DESC LIMIT 10",
        (system.get('client_name',''),)).fetchall()]
    conn.close()
    return render_template('tech_center_view.html', page='tech_center', system=system, devis_list=devis_list)

@app.route('/centre-technique/edit/<int:sid>', methods=['POST'])
@login_required
def tech_center_edit(sid):
    from models import db_update
    db_update('tech_center', sid, client_name=request.form.get('client_name',''),
        code=request.form.get('code',''), contact_name=request.form.get('contact_name',''),
        tel=request.form.get('tel',''), email=request.form.get('email',''),
        system_type=request.form.get('system_type',''), category=request.form.get('category',''),
        next_maintenance=request.form.get('next_maintenance',''),
        maintenance_interval=int(request.form.get('maintenance_interval',90) or 90),
        notes=request.form.get('notes',''), status=request.form.get('status','actif'))
    flash("Modifié","success"); return redirect(f'/centre-technique/view/{sid}')

@app.route('/centre-technique/delete/<int:sid>')
@login_required
def tech_center_delete(sid):
    conn = _gdb(); conn.execute("DELETE FROM tech_center WHERE id=?", (sid,)); conn.commit(); conn.close()
    flash("Supprimé","success"); return redirect(url_for('tech_center'))

@app.route('/prospects/view/<int:pid>')
@login_required
def prospect_view(pid):
    from models import db_get_by_id
    p = db_get_by_id('prospects', pid)
    if not p: flash("Non trouvé","error"); return redirect('/prospects')
    conn = _gdb()
    notes = [dict(r) for r in conn.execute("SELECT * FROM prospect_notes WHERE prospect_id=? ORDER BY created_at DESC", (pid,)).fetchall()]
    tasks = [dict(r) for r in conn.execute("SELECT * FROM prospect_tasks WHERE prospect_id=? ORDER BY created_at DESC", (pid,)).fetchall()]
    offers = [dict(r) for r in conn.execute("SELECT * FROM prospect_offers WHERE prospect_id=? ORDER BY created_at DESC", (pid,)).fetchall()]
    reminders = [dict(r) for r in conn.execute("SELECT * FROM prospect_reminders WHERE prospect_id=? ORDER BY reminder_date ASC", (pid,)).fetchall()]
    files = [dict(r) for r in conn.execute("SELECT * FROM prospect_files WHERE prospect_id=? ORDER BY created_at DESC", (pid,)).fetchall()]
    conn.close()
    tab = request.args.get('tab', 'profil')
    from datetime import datetime as dt2
    now_str = dt2.now().strftime('%Y-%m-%d')
    return render_template('prospect_view.html', page='prospects', prospect=p, tab=tab,
        notes=notes, tasks=tasks, offers=offers, reminders=reminders, files=files, now_str=now_str)

@app.route('/prospects/view/<int:pid>/note/add', methods=['POST'])
@login_required
def prospect_note_add(pid):
    from models import db_insert
    db_insert('prospect_notes', prospect_id=pid, content=request.form.get('content',''), created_by=session['user_id'])
    flash("Note ajoutée","success"); return redirect(f'/prospects/view/{pid}?tab=notes')

@app.route('/prospects/view/<int:pid>/task/add', methods=['POST'])
@login_required
def prospect_task_add(pid):
    from models import db_insert
    db_insert('prospect_tasks', prospect_id=pid, title=request.form.get('title',''),
        status=request.form.get('status','a_faire'), priority=request.form.get('priority','normale'),
        due_date=request.form.get('due_date',''), assigned_to=request.form.get('assigned_to',''),
        created_by=session['user_id'])
    flash("Tâche ajoutée","success"); return redirect(f'/prospects/view/{pid}?tab=taches')

@app.route('/prospects/view/<int:pid>/offer/add', methods=['POST'])
@login_required
def prospect_offer_add(pid):
    from models import db_insert
    db_insert('prospect_offers', prospect_id=pid, title=request.form.get('title',''),
        amount=float(request.form.get('amount',0) or 0), status=request.form.get('status','brouillon'),
        description=request.form.get('description',''), created_by=session['user_id'])
    flash("Offre ajoutée","success"); return redirect(f'/prospects/view/{pid}?tab=offres')

@app.route('/prospects/view/<int:pid>/reminder/add', methods=['POST'])
@login_required
def prospect_reminder_add(pid):
    from models import db_insert
    db_insert('prospect_reminders', prospect_id=pid, title=request.form.get('title',''),
        reminder_date=request.form.get('reminder_date',''), notes=request.form.get('notes',''),
        created_by=session['user_id'])
    flash("Rappel ajouté","success"); return redirect(f'/prospects/view/{pid}?tab=rappels')

@app.route('/prospects/view/<int:pid>/file/add', methods=['POST'])
@login_required
def prospect_file_add(pid):
    from models import db_insert
    if 'file' in request.files and request.files['file'].filename:
        f = request.files['file']
        fname = f"prospect_{pid}_{secure_filename(f.filename)}"
        fdir = os.path.join(app.config['UPLOAD_FOLDER'], 'prospect_files')
        os.makedirs(fdir, exist_ok=True)
        f.save(os.path.join(fdir, fname))
        db_insert('prospect_files', prospect_id=pid, filename=fname,
            original_name=f.filename, created_by=session['user_id'])
        flash("Fichier ajouté","success")
    return redirect(f'/prospects/view/{pid}?tab=pieces')

@app.route('/prospects/view/<int:pid>/delete/<table>/<int:item_id>')
@login_required
def prospect_item_delete(pid, table, item_id):
    allowed = {'note': 'prospect_notes', 'task': 'prospect_tasks', 'offer': 'prospect_offers',
               'reminder': 'prospect_reminders', 'file': 'prospect_files'}
    if table in allowed:
        conn = _gdb(); conn.execute(f"DELETE FROM {allowed[table]} WHERE id=? AND prospect_id=?", (item_id, pid))
        conn.commit(); conn.close()
        flash("Supprimé","success")
    return redirect(f'/prospects/view/{pid}')

# ======================== RESPONSABLE PROJET ========================

@app.route('/resp-projet')
@permission_required('resp_projet')
def resp_projet_dashboard():
    conn = _gdb()
    projects = [dict(r) for r in conn.execute("SELECT * FROM projects ORDER BY created_at DESC").fetchall()]
    tasks = [dict(r) for r in conn.execute("SELECT * FROM tasks ORDER BY created_at DESC").fetchall()]
    
    # KPIs
    total_p = len(projects)
    en_cours = len([p for p in projects if p['status'] == 'en_cours'])
    termines = len([p for p in projects if p['status'] == 'termine'])
    en_retard = len([p for p in projects if p.get('end_date') and p['end_date'] < datetime.now().strftime('%Y-%m-%d') and p['status'] != 'termine'])
    
    tasks_total = len(tasks)
    tasks_done = len([t for t in tasks if t['status'] == 'termine'])
    tasks_pending = len([t for t in tasks if t['status'] == 'a_faire'])
    tasks_progress = len([t for t in tasks if t['status'] == 'en_cours'])
    
    budget_total = sum(float(p.get('budget',0) or 0) for p in projects)
    budget_consumed = sum(float(p.get('budget_consumed',0) or 0) for p in projects)
    
    # Recent tasks
    recent = [dict(r) for r in conn.execute("""SELECT t.*, p.name as project_name, u.full_name as assignee 
        FROM tasks t LEFT JOIN projects p ON t.project_id=p.id LEFT JOIN users u ON t.assigned_to=u.id
        ORDER BY t.created_at DESC LIMIT 10""").fetchall()]
    
    # Deadlines this week
    from datetime import timedelta
    today = datetime.now().date()
    week_end = (today + timedelta(days=6)).strftime('%Y-%m-%d')
    deadlines = [dict(r) for r in conn.execute("""SELECT t.*, p.name as project_name 
        FROM tasks t LEFT JOIN projects p ON t.project_id=p.id 
        WHERE t.due_date <= ? AND t.status != 'termine' ORDER BY t.due_date""",
        (week_end,)).fetchall()]
    
    conn.close()
    return render_template('resp_projet.html', page='resp_projet',
        projects=projects, total_p=total_p, en_cours=en_cours, termines=termines, en_retard=en_retard,
        tasks_total=tasks_total, tasks_done=tasks_done, tasks_pending=tasks_pending, tasks_progress=tasks_progress,
        budget_total=budget_total, budget_consumed=budget_consumed,
        recent=recent, deadlines=deadlines, today=datetime.now().strftime('%Y-%m-%d'))

@app.route('/resp-projet/projets')
@permission_required('resp_projet')
def resp_projet_list():
    conn = _gdb()
    projects = [dict(r) for r in conn.execute("""SELECT p.*, u.full_name as manager_name,
        (SELECT COUNT(*) FROM tasks WHERE project_id=p.id) as task_count,
        (SELECT COUNT(*) FROM tasks WHERE project_id=p.id AND status='termine') as task_done
        FROM projects p LEFT JOIN users u ON p.manager_id=u.id ORDER BY p.created_at DESC""").fetchall()]
    users = get_all_users()
    clients = get_all_clients()
    conn.close()
    return render_template('resp_projet_list.html', page='resp_projets', projects=projects, users=users, clients=clients)

@app.route('/resp-projet/projet/add', methods=['POST'])
@permission_required('resp_projet')
def resp_projet_add():
    from models import db_insert
    db_insert('projects', name=request.form.get('name',''), description=request.form.get('description',''),
        objectives=request.form.get('objectives',''), client=request.form.get('client',''),
        status=request.form.get('status','non_commence'), priority=request.form.get('priority','moyenne'),
        start_date=request.form.get('start_date',''), end_date=request.form.get('end_date',''),
        budget=float(request.form.get('budget',0) or 0),
        manager_id=int(request.form.get('manager_id',0) or 0) or None,
        client_id=int(request.form.get('client_id',0) or 0) or None,
        created_by=session['user_id'])
    flash("Projet créé","success"); return redirect('/resp-projet/projets')

@app.route('/resp-projet/projet/<int:pid>')
@permission_required('resp_projet')
def resp_projet_view(pid):
    from models import db_get_by_id
    project = db_get_by_id('projects', pid)
    if not project: flash("Non trouvé","error"); return redirect('/resp-projet/projets')
    conn = _gdb()
    tasks = [dict(r) for r in conn.execute("""SELECT t.*, u.full_name as assignee FROM tasks t 
        LEFT JOIN users u ON t.assigned_to=u.id WHERE t.project_id=? ORDER BY t.priority DESC, t.due_date""",
        (pid,)).fetchall()]
    comments = [dict(r) for r in conn.execute("""SELECT tc.*, u.full_name FROM task_comments tc
        LEFT JOIN users u ON tc.user_id=u.id WHERE tc.task_id IN (SELECT id FROM tasks WHERE project_id=?)
        ORDER BY tc.created_at DESC LIMIT 20""", (pid,)).fetchall()]
    users = get_all_users()
    conn.close()
    
    done = len([t for t in tasks if t['status'] == 'termine'])
    progress = round(done / max(len(tasks), 1) * 100)
    
    return render_template('resp_projet_view.html', page='resp_projets', project=project,
        tasks=tasks, comments=comments, users=users, progress=progress, today=datetime.now().strftime('%Y-%m-%d'))

@app.route('/resp-projet/projet/<int:pid>/task/add', methods=['POST'])
@permission_required('resp_projet')
def resp_projet_task_add(pid):
    from models import db_insert
    db_insert('tasks', project_id=pid, title=request.form.get('title',''),
        description=request.form.get('description',''), priority=request.form.get('priority','moyenne'),
        status='a_faire', due_date=request.form.get('due_date',''),
        assigned_to=int(request.form.get('assigned_to',0) or 0) or None,
        created_by=session['user_id'])
    flash("Tâche ajoutée","success"); return redirect(f'/resp-projet/projet/{pid}')

@app.route('/resp-projet/task/<int:tid>/status/<status>')
@permission_required('resp_projet')
def resp_projet_task_status(tid, status):
    from models import db_get_by_id as _gbi, db_update
    if status in ('a_faire','en_cours','termine'):
        db_update('tasks', tid, status=status)
        flash("Statut mis à jour","success")
    t = _gbi('tasks', tid)
    return redirect(f'/resp-projet/projet/{t["project_id"]}' if t else '/resp-projet')

@app.route('/resp-projet/task/<int:tid>/comment', methods=['POST'])
@permission_required('resp_projet')
def resp_projet_comment(tid):
    from models import db_insert, db_get_by_id as _gbi2
    db_insert('task_comments', task_id=tid, user_id=session['user_id'], content=request.form.get('content',''))
    t = _gbi2('tasks', tid)
    flash("Commentaire ajouté","success")
    return redirect(f'/resp-projet/projet/{t["project_id"]}' if t else '/resp-projet')

@app.route('/resp-projet/planning')
@permission_required('resp_projet')
def resp_projet_planning():
    conn = _gdb()
    projects = [dict(r) for r in conn.execute("SELECT * FROM projects WHERE start_date != '' ORDER BY start_date").fetchall()]
    tasks = [dict(r) for r in conn.execute("""SELECT t.*, p.name as project_name FROM tasks t 
        LEFT JOIN projects p ON t.project_id=p.id WHERE t.due_date != '' ORDER BY t.due_date""").fetchall()]
    conn.close()
    return render_template('resp_projet_planning.html', page='resp_planning', projects=projects, tasks=tasks, today=datetime.now().strftime('%Y-%m-%d'))

# ======================== CONTRATS RH ========================

@app.route('/rh/contrats-rh')
@permission_required('fichiers')
def rh_contracts():
    from models import db_get_all
    contracts = db_get_all('rh_contracts', order='end_date ASC')
    # Enrich with employee names
    employees = get_all_employees(status=None)
    emp_map = {e['id']: f"{e['first_name']} {e['last_name']}" for e in employees}
    for c in contracts:
        c['employee_name'] = emp_map.get(c.get('employee_id'), '-')
    return render_template('rh_contrats.html', page='contrats_rh', contracts=contracts, employees=employees)

@app.route('/rh/contrats-rh/add', methods=['POST'])
@permission_required('fichiers')
def rh_contracts_add():
    from models import db_insert
    db_insert('rh_contracts', code=request.form.get('code',''),
        employee_id=int(request.form['employee_id']),
        contract_type=request.form.get('contract_type','CDI'),
        start_date=request.form.get('start_date',''),
        end_date=request.form.get('end_date',''),
        salary=float(request.form.get('salary',0) or 0),
        notes=request.form.get('notes',''))
    flash("Contrat RH ajouté", "success")
    return redirect(url_for('rh_contracts'))


# ======================== ORGANIGRAMME ========================

@app.route('/rh/organigramme')
@permission_required('fichiers')
def rh_organigramme():
    employees = get_all_employees()
    depts = {}
    for e in employees:
        d = e.get('department') or 'Non assigné'
        if d not in depts: depts[d] = []
        depts[d].append(e)
    
    # Build hierarchy levels based on position keywords
    hierarchy = {0: [], 1: [], 2: [], 3: []}
    level_names = {0: 'Direction Générale', 1: 'Directeurs & Chefs', 2: 'Responsables', 3: 'Équipes'}
    for e in employees:
        pos = (e.get('position') or '').lower()
        if any(k in pos for k in ['dg', 'directeur général', 'pdg', 'ceo', 'gérant']):
            hierarchy[0].append(e)
        elif any(k in pos for k in ['directeur', 'director', 'chef de département', 'daf', 'drh', 'dsi']):
            hierarchy[1].append(e)
        elif any(k in pos for k in ['responsable', 'chef', 'superviseur', 'manager', 'coordinat']):
            hierarchy[2].append(e)
        else:
            hierarchy[3].append(e)
    
    return render_template('rh_organigramme.html', page='organigramme', depts=depts,
                          hierarchy=hierarchy, level_names=level_names)


# ======================== CHAT ========================

@app.route('/chat')
@login_required
def chat_page():
    channel = request.args.get('channel', 'general')
    dm_user = request.args.get('dm')
    users = get_all_users()
    if dm_user:
        msgs = get_direct_messages(session['user_id'], int(dm_user))
        target = get_user_by_id(int(dm_user))
        mark_chat_read(session['user_id'], '_dm_all')
        return render_template('chat.html', page='chat', messages=msgs, users=users,
                              channel=f'dm_{dm_user}', dm_target=target)
    msgs = get_messages(channel)
    mark_chat_read(session['user_id'], channel)
    return render_template('chat.html', page='chat', messages=msgs, users=users,
                          channel=channel, dm_target=None)

@app.route('/chat/send', methods=['POST'])
@login_required
def chat_send():
    content = request.form.get('content', '').strip()
    channel = request.form.get('channel', 'general')
    dm_id = None
    if channel.startswith('dm_'):
        dm_id = int(channel.split('_')[1])
    if content:
        if dm_id:
            send_message(session['user_id'], content, 'direct', dm_id)
        else:
            send_message(session['user_id'], content, channel, None)
    if dm_id:
        return redirect(f'/chat?dm={dm_id}')
    return redirect(f'/chat?channel={channel}')

@app.route('/chat/unread')
@login_required
def chat_unread_api():
    """API: nombre de messages non lus."""
    count = get_unread_count(session['user_id'])
    return jsonify({'unread': count})

@app.route('/chat/api')
@login_required
def chat_api():
    """API pour rafraîchir les messages (polling)."""
    channel = request.args.get('channel', 'general')
    dm_user = request.args.get('dm')
    if dm_user:
        msgs = get_direct_messages(session['user_id'], int(dm_user))
    else:
        msgs = get_messages(channel)
    return jsonify([{'id': m['id'], 'sender': m['sender_name'], 'sender_id': m['sender_id'],
                     'content': m['content'], 'time': m['created_at'][11:16]} for m in msgs])


# ======================== APPELS AUDIO/VIDEO ========================

@app.route('/call/start', methods=['POST'])
@login_required
def call_start():
    """Initier un appel."""
    target_id = int(request.form.get('target_id', 0))
    call_type = request.form.get('type', 'audio')
    room = request.form.get('room', '') or f"wannygest-{session['user_id']}-{target_id}-{int(datetime.now().timestamp())}"
    
    conn = _gdb()
    # Clear old ringing calls from this caller
    conn.execute("DELETE FROM calls WHERE caller_id=? AND status='ringing'", (session['user_id'],))
    
    if target_id > 0:
        # Direct call to specific user
        conn.execute("""INSERT INTO calls (caller_id, callee_id, room, call_type, status)
            VALUES (?, ?, ?, ?, 'ringing')""",
            (session['user_id'], target_id, room, call_type))
    else:
        # Channel call — ring all other users
        users = conn.execute("SELECT id FROM users WHERE id!=?", (session['user_id'],)).fetchall()
        for u in users:
            conn.execute("""INSERT INTO calls (caller_id, callee_id, room, call_type, status)
                VALUES (?, ?, ?, ?, 'ringing')""",
                (session['user_id'], u['id'], room, call_type))
    
    conn.commit(); conn.close()
    return jsonify({'room': room, 'url': f'https://meet.jit.si/{room}'})

@app.route('/call/check')
@login_required
def call_check():
    """Vérifie si un appel entrant est en cours (polling)."""
    conn = _gdb()
    # Auto-expire ringing calls older than 60s
    conn.execute("""UPDATE calls SET status='missed' 
        WHERE status='ringing' AND created_at < datetime('now', '-60 seconds')""")
    conn.commit()
    
    call = conn.execute("""SELECT c.*, u.full_name as caller_name FROM calls c
        JOIN users u ON c.caller_id=u.id
        WHERE c.callee_id=? AND c.status='ringing'
        ORDER BY c.created_at DESC LIMIT 1""", (session['user_id'],)).fetchone()
    conn.close()
    if call:
        return jsonify({
            'incoming': True,
            'call_id': call['id'],
            'caller': call['caller_name'],
            'type': call['call_type'],
            'room': call['room'],
            'url': f"https://meet.jit.si/{call['room']}"
        })
    return jsonify({'incoming': False})

@app.route('/call/<int:cid>/accept')
@login_required
def call_accept(cid):
    conn = _gdb()
    call = conn.execute("SELECT * FROM calls WHERE id=? AND callee_id=?", (cid, session['user_id'])).fetchone()
    if call:
        conn.execute("UPDATE calls SET status='active' WHERE id=?", (cid,))
        conn.commit()
        conn.close()
        return jsonify({'url': f"https://meet.jit.si/{call['room']}", 'room': call['room']})
    conn.close()
    return jsonify({'error': 'Appel non trouvé'}), 404

@app.route('/call/<int:cid>/reject')
@login_required
def call_reject(cid):
    conn = _gdb()
    conn.execute("UPDATE calls SET status='rejected' WHERE id=? AND callee_id=?", (cid, session['user_id']))
    conn.commit(); conn.close()
    return jsonify({'ok': True})

@app.route('/call/<int:cid>/end')
@login_required
def call_end(cid):
    conn = _gdb()
    conn.execute("UPDATE calls SET status='ended', ended_at=CURRENT_TIMESTAMP WHERE id=?", (cid,))
    conn.commit(); conn.close()
    return jsonify({'ok': True})


# ======================== RH EXPANDED ========================

@app.route('/rh/postes')
@permission_required('fichiers')
def rh_postes():
    from models import db_get_all
    postes = db_get_all('rh_job_descriptions', order='title ASC')
    return render_template('rh_postes.html', page='postes', postes=postes)

@app.route('/rh/postes/add', methods=['POST'])
@permission_required('fichiers')
def rh_postes_add():
    from models import db_insert
    db_insert('rh_job_descriptions', title=request.form['title'],
        department=request.form.get('department',''), description=request.form.get('description',''),
        requirements=request.form.get('requirements',''), responsibilities=request.form.get('responsibilities',''),
        salary_range=request.form.get('salary_range',''))
    flash("Poste ajouté", "success"); return redirect(url_for('rh_postes'))

@app.route('/rh/formations')
@permission_required('fichiers')
def rh_formations():
    from models import db_get_all
    u = dict(get_user_by_id(session["user_id"]) or {})
    all_trainings = db_get_all('rh_trainings', order='date DESC')
    # Filter: show only trainings for user's department or "tous"
    if u['role'] in ('admin', 'dg', 'rh'):
        trainings = all_trainings  # Admin/DG/RH see all
    else:
        dept = u.get('department', '') or u.get('role', '')
        trainings = [t for t in all_trainings if not t.get('target') or t.get('target') == 'tous' or t.get('target','').lower() == dept.lower()]
    return render_template('rh_formations.html', page='formations', trainings=trainings)

@app.route('/rh/formations/add', methods=['POST'])
@permission_required('fichiers')
def rh_formations_add():
    from models import db_insert
    target = request.form.get('target', 'tous')
    db_insert('rh_trainings', title=request.form['title'], description=request.form.get('description',''),
        trainer=request.form.get('trainer',''), date=request.form.get('date',''),
        duration=request.form.get('duration',''),
        department=request.form.get('department',''),
        cost=request.form.get('cost','0'),
        target=target)
    flash("Formation ajoutée", "success"); return redirect(url_for('rh_formations'))

@app.route('/rh/formations/edit/<int:fid>', methods=['POST'])
@permission_required('fichiers')
def rh_formations_edit(fid):
    from models import db_update
    db_update('rh_trainings', fid, title=request.form['title'], description=request.form.get('description',''),
        trainer=request.form.get('trainer',''), date=request.form.get('date',''),
        duration=request.form.get('duration',''), department=request.form.get('department',''),
        cost=request.form.get('cost','0'), target=request.form.get('target','tous'),
        status=request.form.get('status','planifie'))
    flash("Formation modifiée", "success"); return redirect(url_for('rh_formations'))

@app.route('/rh/formations/delete/<int:fid>')
@permission_required('fichiers')
def rh_formations_delete(fid):
    conn = _gdb()
    conn.execute("DELETE FROM rh_trainings WHERE id=?", (fid,))
    conn.commit(); conn.close()
    flash("Formation supprimée", "success"); return redirect(url_for('rh_formations'))

@app.route('/rh/annonces')
@permission_required('fichiers')
def rh_annonces():
    from models import db_get_all
    annonces = db_get_all('rh_announcements', order='created_at DESC')
    return render_template('rh_annonces.html', page='annonces', annonces=annonces)

@app.route('/rh/annonces/add', methods=['POST'])
@permission_required('fichiers')
def rh_annonces_add():
    from models import db_insert
    db_insert('rh_announcements', title=request.form['title'], content=request.form.get('content',''),
        priority=request.form.get('priority','normale'), created_by=session['user_id'])
    flash("Annonce publiée", "success"); return redirect(url_for('rh_annonces'))

@app.route('/rh/annonces/edit/<int:aid>', methods=['POST'])
@permission_required('fichiers')
def rh_annonces_edit(aid):
    conn = _gdb()
    conn.execute("UPDATE rh_announcements SET title=?, content=?, priority=? WHERE id=?",
        (request.form['title'], request.form.get('content',''), request.form.get('priority','normale'), aid))
    conn.commit(); conn.close()
    flash("Annonce modifiée", "success"); return redirect(url_for('rh_annonces'))

@app.route('/rh/annonces/delete/<int:aid>')
@permission_required('fichiers')
def rh_annonces_delete(aid):
    conn = _gdb()
    conn.execute("DELETE FROM rh_announcements WHERE id=?", (aid,))
    conn.commit(); conn.close()
    flash("Annonce supprimée", "success"); return redirect(url_for('rh_annonces'))


# ======================== MODULE TRACKING GPS ========================

@app.route('/tracking')
@permission_required('tracking')
def tracking_dashboard():
    conn = _gdb()
    vehicles = [dict(r) for r in conn.execute("SELECT * FROM tracking_vehicles ORDER BY status, immatriculation").fetchall()]
    actifs = len([v for v in vehicles if v['status'] == 'actif'])
    en_mouvement = len([v for v in vehicles if v.get('last_speed',0) and float(v.get('last_speed',0) or 0) > 2])
    alerts_new = conn.execute("SELECT COUNT(*) FROM tracking_alerts WHERE acknowledged=0").fetchone()[0]
    recent_alerts = [dict(r) for r in conn.execute("""SELECT a.*, v.immatriculation, v.marque 
        FROM tracking_alerts a LEFT JOIN tracking_vehicles v ON a.vehicle_id=v.id
        ORDER BY a.created_at DESC LIMIT 10""").fetchall()]
    conn.close()
    return render_template('tracking_dashboard.html', page='tracking', vehicles=vehicles,
        actifs=actifs, en_mouvement=en_mouvement, alerts_new=alerts_new, recent_alerts=recent_alerts)

@app.route('/tracking/vehicules')
@permission_required('tracking')
def tracking_vehicules():
    conn = _gdb()
    vehicles = [dict(r) for r in conn.execute("""SELECT v.*, u.full_name as tech_name 
        FROM tracking_vehicles v LEFT JOIN users u ON v.created_by=u.id
        ORDER BY v.created_at DESC""").fetchall()]
    conn.close()
    return render_template('tracking_vehicules.html', page='tracking_vehicules', vehicles=vehicles)

@app.route('/tracking/vehicules/add', methods=['POST'])
@permission_required('tracking')
def tracking_vehicule_add():
    conn = _gdb()
    conn.execute("""INSERT INTO tracking_vehicles (immatriculation, marque, modele, type, couleur, annee,
        proprietaire, tel_proprietaire, gps_device_id, gps_brand, gps_model, gps_sim, gps_imei,
        installation_date, installation_tech, notes, created_by)
        VALUES (?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?)""",
        (request.form.get('immatriculation',''), request.form.get('marque',''),
         request.form.get('modele',''), request.form.get('type','voiture'),
         request.form.get('couleur',''), request.form.get('annee',''),
         request.form.get('proprietaire',''), request.form.get('tel_proprietaire',''),
         request.form.get('gps_device_id',''), request.form.get('gps_brand','Concox'),
         request.form.get('gps_model',''), request.form.get('gps_sim',''),
         request.form.get('gps_imei',''), request.form.get('installation_date',''),
         request.form.get('installation_tech',''), request.form.get('notes',''),
         session['user_id']))
    conn.commit(); conn.close()
    flash("Véhicule ajouté","success"); return redirect('/tracking/vehicules')

@app.route('/tracking/vehicules/edit/<int:vid>', methods=['GET','POST'])
@permission_required('tracking')
def tracking_vehicule_edit(vid):
    conn = _gdb()
    v = conn.execute("SELECT * FROM tracking_vehicles WHERE id=?", (vid,)).fetchone()
    if not v: flash("Non trouvé","error"); return redirect('/tracking/vehicules')
    if request.method == 'POST':
        for col in ['immatriculation','marque','modele','type','couleur','annee',
                     'proprietaire','tel_proprietaire','gps_device_id','gps_brand',
                     'gps_model','gps_sim','gps_imei','installation_date',
                     'installation_tech','status','notes']:
            val = request.form.get(col, '')
            conn.execute(f"UPDATE tracking_vehicles SET {col}=? WHERE id=?", (val, vid))
        conn.commit(); conn.close()
        flash("Véhicule modifié","success"); return redirect('/tracking/vehicules')
    conn.close()
    return render_template('tracking_vehicule_edit.html', page='tracking_vehicules', vehicle=dict(v))

@app.route('/tracking/vehicules/view/<int:vid>')
@permission_required('tracking')
def tracking_vehicule_view(vid):
    conn = _gdb()
    v = conn.execute("SELECT * FROM tracking_vehicles WHERE id=?", (vid,)).fetchone()
    if not v: flash("Non trouvé","error"); return redirect('/tracking/vehicules')
    history = [dict(r) for r in conn.execute(
        "SELECT * FROM tracking_history WHERE vehicle_id=? ORDER BY created_at DESC LIMIT 50", (vid,)).fetchall()]
    alerts = [dict(r) for r in conn.execute(
        "SELECT * FROM tracking_alerts WHERE vehicle_id=? ORDER BY created_at DESC LIMIT 20", (vid,)).fetchall()]
    conn.close()
    return render_template('tracking_vehicule_view.html', page='tracking_vehicules', vehicle=dict(v), history=history, alerts=alerts)

@app.route('/tracking/carte')
@permission_required('tracking')
def tracking_carte():
    conn = _gdb()
    vehicles = [dict(r) for r in conn.execute(
        "SELECT * FROM tracking_vehicles WHERE status='actif' AND last_lat IS NOT NULL ORDER BY immatriculation").fetchall()]
    conn.close()
    return render_template('tracking_carte.html', page='tracking_carte', vehicles=vehicles)

@app.route('/tracking/alertes')
@permission_required('tracking')
def tracking_alertes():
    conn = _gdb()
    alerts = [dict(r) for r in conn.execute("""SELECT a.*, v.immatriculation, v.marque, v.modele
        FROM tracking_alerts a LEFT JOIN tracking_vehicles v ON a.vehicle_id=v.id
        ORDER BY a.created_at DESC LIMIT 100""").fetchall()]
    conn.close()
    return render_template('tracking_alertes.html', page='tracking_alertes', alerts=alerts)

@app.route('/tracking/alertes/<int:aid>/ack')
@permission_required('tracking')
def tracking_alert_ack(aid):
    conn = _gdb()
    conn.execute("UPDATE tracking_alerts SET acknowledged=1 WHERE id=?", (aid,))
    conn.commit(); conn.close()
    flash("Alerte acquittée","success"); return redirect('/tracking/alertes')

@app.route('/tracking/position/update', methods=['POST'])
def tracking_position_update():
    """API endpoint pour recevoir les positions GPS (depuis les balises ou un serveur intermédiaire)."""
    data = request.get_json(silent=True) or {}
    device_id = data.get('device_id','')
    if not device_id: return jsonify({'error': 'device_id requis'}), 400
    conn = _gdb()
    v = conn.execute("SELECT id FROM tracking_vehicles WHERE gps_device_id=?", (device_id,)).fetchone()
    if not v: conn.close(); return jsonify({'error': 'device inconnu'}), 404
    lat = float(data.get('lat',0)); lng = float(data.get('lng',0)); speed = float(data.get('speed',0))
    addr = data.get('address','')
    conn.execute("UPDATE tracking_vehicles SET last_lat=?, last_lng=?, last_speed=?, last_address=?, last_update=? WHERE id=?",
        (lat, lng, speed, addr, datetime.now().strftime('%Y-%m-%d %H:%M:%S'), v['id']))
    conn.execute("INSERT INTO tracking_history (vehicle_id, lat, lng, speed, address) VALUES (?,?,?,?,?)",
        (v['id'], lat, lng, speed, addr))
    # Alert: excès de vitesse
    if speed > 120:
        conn.execute("INSERT INTO tracking_alerts (vehicle_id, alert_type, message, lat, lng) VALUES (?,?,?,?,?)",
            (v['id'], 'vitesse', f'Excès de vitesse: {speed:.0f} km/h', lat, lng))
    conn.commit(); conn.close()
    return jsonify({'ok': True})

# ======================== MODULE IT ========================

@app.route('/it')
@permission_required('informatique')
def it_dashboard():
    conn = _gdb()
    equip_total = conn.execute("SELECT COUNT(*) FROM it_equipment").fetchone()[0]
    equip_actif = conn.execute("SELECT COUNT(*) FROM it_equipment WHERE status='actif'").fetchone()[0]
    tickets_open = conn.execute("SELECT COUNT(*) FROM it_tickets WHERE status IN ('ouvert','en_cours')").fetchone()[0]
    tickets_resolved = conn.execute("SELECT COUNT(*) FROM it_tickets WHERE status='resolu'").fetchone()[0]
    tickets_recent = [dict(r) for r in conn.execute("""SELECT t.*, u.full_name as requester, u2.full_name as assignee
        FROM it_tickets t LEFT JOIN users u ON t.requester_id=u.id LEFT JOIN users u2 ON t.assigned_to=u2.id
        ORDER BY t.created_at DESC LIMIT 10""").fetchall()]
    logs_recent = [dict(r) for r in conn.execute("SELECT * FROM it_logs ORDER BY created_at DESC LIMIT 10").fetchall()]
    conn.close()
    return render_template('it_dashboard.html', page='it', equip_total=equip_total, equip_actif=equip_actif,
        tickets_open=tickets_open, tickets_resolved=tickets_resolved, tickets_recent=tickets_recent, logs_recent=logs_recent)

@app.route('/it/parc')
@permission_required('informatique')
def it_parc():
    conn = _gdb()
    equipment = [dict(r) for r in conn.execute("""SELECT e.*, u.full_name as user_name FROM it_equipment e
        LEFT JOIN users u ON e.assigned_to=u.id ORDER BY e.name""").fetchall()]
    users = get_all_users()
    conn.close()
    return render_template('it_parc.html', page='it_parc', equipment=equipment, users=users)

@app.route('/it/parc/add', methods=['POST'])
@permission_required('informatique')
def it_parc_add():
    conn = _gdb()
    conn.execute("""INSERT INTO it_equipment (name, type, brand, model, serial_number, assigned_to, location,
        status, purchase_date, purchase_price, warranty_end, notes) VALUES (?,?,?,?,?,?,?,?,?,?,?,?)""",
        (request.form.get('name',''), request.form.get('type',''), request.form.get('brand',''),
         request.form.get('model',''), request.form.get('serial_number',''),
         int(request.form.get('assigned_to',0) or 0) or None, request.form.get('location',''),
         'actif', request.form.get('purchase_date',''),
         float(request.form.get('purchase_price',0) or 0), request.form.get('warranty_end',''),
         request.form.get('notes','')))
    conn.commit(); conn.close()
    flash("Équipement ajouté","success"); return redirect('/it/parc')

@app.route('/it/tickets')
@permission_required('informatique')
def it_tickets():
    tab = request.args.get('tab', 'ouvert')
    conn = _gdb()
    if tab == 'all':
        tickets = [dict(r) for r in conn.execute("""SELECT t.*, u.full_name as requester, u2.full_name as assignee
            FROM it_tickets t LEFT JOIN users u ON t.requester_id=u.id LEFT JOIN users u2 ON t.assigned_to=u2.id
            ORDER BY t.created_at DESC""").fetchall()]
    else:
        tickets = [dict(r) for r in conn.execute("""SELECT t.*, u.full_name as requester, u2.full_name as assignee
            FROM it_tickets t LEFT JOIN users u ON t.requester_id=u.id LEFT JOIN users u2 ON t.assigned_to=u2.id
            WHERE t.status=? ORDER BY t.created_at DESC""", (tab,)).fetchall()]
    stats = {
        'ouvert': conn.execute("SELECT COUNT(*) FROM it_tickets WHERE status='ouvert'").fetchone()[0],
        'en_cours': conn.execute("SELECT COUNT(*) FROM it_tickets WHERE status='en_cours'").fetchone()[0],
        'resolu': conn.execute("SELECT COUNT(*) FROM it_tickets WHERE status='resolu'").fetchone()[0],
    }
    users = get_all_users()
    equipment = [dict(r) for r in conn.execute("SELECT id, name FROM it_equipment ORDER BY name").fetchall()]
    conn.close()
    return render_template('it_tickets.html', page='it_tickets', tickets=tickets, tab=tab, stats=stats, users=users, equipment=equipment)

@app.route('/it/tickets/add', methods=['POST'])
@permission_required('informatique')
def it_ticket_add():
    conn = _gdb()
    conn.execute("""INSERT INTO it_tickets (title, description, category, priority, status, requester_id, assigned_to, equipment_id)
        VALUES (?,?,?,?,?,?,?,?)""",
        (request.form.get('title',''), request.form.get('description',''),
         request.form.get('category','incident'), request.form.get('priority','normal'),
         'ouvert', session['user_id'],
         int(request.form.get('assigned_to',0) or 0) or None,
         int(request.form.get('equipment_id',0) or 0) or None))
    conn.commit(); conn.close()
    flash("Ticket créé","success"); return redirect('/it/tickets')

@app.route('/it/tickets/<int:tid>/status/<status>')
@permission_required('informatique')
def it_ticket_status(tid, status):
    conn = _gdb()
    if status in ('ouvert','en_cours','resolu','ferme'):
        updates = {'status': status}
        if status == 'resolu': updates['resolved_at'] = datetime.now().strftime('%Y-%m-%d %H:%M')
        conn.execute(f"UPDATE it_tickets SET status=?, resolved_at=? WHERE id=?",
            (status, updates.get('resolved_at',''), tid))
        conn.commit()
    conn.close()
    flash("Ticket mis à jour","success"); return redirect('/it/tickets')

@app.route('/it/securite')
@permission_required('informatique')
def it_securite():
    conn = _gdb()
    logs = [dict(r) for r in conn.execute("SELECT * FROM it_logs ORDER BY created_at DESC LIMIT 50").fetchall()]
    login_logs = [dict(r) for r in conn.execute("SELECT * FROM activity_logs WHERE action LIKE '%Connexion%' ORDER BY created_at DESC LIMIT 20").fetchall()]
    conn.close()
    return render_template('it_securite.html', page='it_securite', logs=logs, login_logs=login_logs)

# ======================== KANBAN ========================

@app.route('/kanban')
@login_required
def kanban():
    from models import db_get_all, get_all_users, get_db
    user = get_user_by_id(session['user_id'])
    if user and user['role'] == 'admin':
        tasks = db_get_all('tasks')
    else:
        tasks = db_get_all('tasks', where={'assigned_to': session['user_id']})
    # Enrich with project names
    conn = get_db()
    projects = {}
    try:
        for p in conn.execute("SELECT id, name FROM projects").fetchall():
            projects[p['id']] = p['name']
    except: pass
    conn.close()
    for t in tasks:
        t['project_name'] = projects.get(t.get('project_id'), '')
    by_status = {'a_faire': [], 'en_cours': [], 'en_revue': [], 'termine': []}
    for t in tasks:
        s = t.get('status', 'a_faire')
        if s in by_status: by_status[s].append(t)
        else: by_status['a_faire'].append(t)
    users = get_all_users()
    user_map = {u['id']: u['full_name'] for u in users}
    return render_template('kanban.html', page='kanban', by_status=by_status, users=users, user_map=user_map)

@app.route('/taches')
@login_required
def taches_list():
    return redirect(url_for('kanban'))

@app.route('/taches/new')
@login_required
def taches_new():
    return redirect('/resp-projet/projets')

@app.route('/kanban/move/<int:tid>/<status>')
@login_required
def kanban_move(tid, status):
    from models import db_update
    if status in ('a_faire','en_cours','en_revue','termine'):
        db_update('tasks', tid, status=status)
        user = get_user_by_id(session['user_id'])
        log_audit(session['user_id'], user['full_name'] if user else '?', 'tasks', tid, 'status_change', 'status', '', status)
    return redirect(url_for('kanban'))


# ======================== HISTORIQUE DES MODIFICATIONS ========================

@app.route('/historique')
@permission_required('admin')
def historique():
    table = request.args.get('table', '')
    trail = get_audit_trail(table_name=table if table else None, limit=100)
    return render_template('historique.html', page='historique', trail=trail, filter_table=table)


# ======================== TABLEAU DE BORD EXÉCUTIF ========================

@app.route('/executif')
@permission_required('admin')
def executif():
    stats = get_executive_stats()
    return render_template('executif.html', page='executif', s=stats)

@app.route('/executif/pdf')
@permission_required('admin')
def executif_pdf():
    """Génère le rapport PDF mensuel exécutif."""
    from reportlab.lib.pagesizes import A4
    from reportlab.lib.units import mm
    from reportlab.lib.colors import HexColor
    from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer, HRFlowable
    from reportlab.lib.styles import ParagraphStyle
    from reportlab.lib.enums import TA_CENTER, TA_RIGHT

    s = get_executive_stats()
    output = os.path.join(app.config['UPLOAD_FOLDER'], 'rapport_executif.pdf')
    os.makedirs(os.path.dirname(output), exist_ok=True)
    doc = SimpleDocTemplate(output, pagesize=A4, leftMargin=20*mm, rightMargin=20*mm, topMargin=15*mm, bottomMargin=15*mm)

    NAVY = HexColor('#1a3a5c'); ORANGE = HexColor('#e8672a')
    s_t = ParagraphStyle('t', fontSize=20, fontName='Helvetica-Bold', textColor=NAVY, alignment=TA_CENTER)
    s_sub = ParagraphStyle('sub', fontSize=11, alignment=TA_CENTER, textColor=HexColor('#888'))
    s_h = ParagraphStyle('h', fontSize=14, fontName='Helvetica-Bold', textColor=NAVY, spaceBefore=12, spaceAfter=6)
    s_n = ParagraphStyle('n', fontSize=10)
    s_r = ParagraphStyle('r', fontSize=10, alignment=TA_RIGHT)
    s_hd = ParagraphStyle('hd', fontSize=9, fontName='Helvetica-Bold', textColor=HexColor('#fff'))
    s_c = ParagraphStyle('c', fontSize=9)
    fmt = lambda x: f"{x:,.0f}"

    story = []
    story.append(Paragraph("RAPPORT EXÉCUTIF MENSUEL", s_t))
    from datetime import datetime
    story.append(Paragraph(f"WannyGest — {datetime.now().strftime('%B %Y')}", s_sub))
    story.append(Spacer(1, 3*mm))
    story.append(HRFlowable(width="100%", thickness=2, color=NAVY))
    story.append(Spacer(1, 6*mm))

    # KPIs table
    story.append(Paragraph("Indicateurs Clés", s_h))
    kpis = [
        [Paragraph(h, s_hd) for h in ['Indicateur', 'Valeur']],
        [Paragraph('Clients actifs', s_c), Paragraph(str(s['clients']), s_r)],
        [Paragraph('Employés', s_c), Paragraph(str(s['employes']), s_r)],
        [Paragraph('Rapports traités', s_c), Paragraph(str(s['rapports']), s_r)],
        [Paragraph('Prospects totaux', s_c), Paragraph(str(s['prospects']), s_r)],
        [Paragraph('Prospects gagnés', s_c), Paragraph(str(s['prospects_gagnes']), s_r)],
    ]
    t = Table(kpis, colWidths=[100*mm, 60*mm])
    t.setStyle(TableStyle([('BACKGROUND',(0,0),(-1,0),NAVY),('GRID',(0,0),(-1,-1),0.5,HexColor('#ddd')),
        ('TOPPADDING',(0,0),(-1,-1),5),('BOTTOMPADDING',(0,0),(-1,-1),5)]))
    story.append(t)
    story.append(Spacer(1, 6*mm))

    # Finance
    story.append(Paragraph("Finance", s_h))
    fin = [
        [Paragraph(h, s_hd) for h in ['Rubrique', 'Montant (FCFA)']],
        [Paragraph('CA facturé', s_c), Paragraph(fmt(s['montant_facture']), s_r)],
        [Paragraph('CA encaissé', s_c), Paragraph(fmt(s['montant_paye']), s_r)],
        [Paragraph('Impayé', s_c), Paragraph(fmt(s['montant_impaye']), s_r)],
        [Paragraph('CA devis acceptés', s_c), Paragraph(fmt(s['ca_devis']), s_r)],
        [Paragraph('Recettes trésorerie', s_c), Paragraph(fmt(s['recettes']), s_r)],
        [Paragraph('Dépenses', s_c), Paragraph(fmt(s['depenses']), s_r)],
        [Paragraph('<b>Solde</b>', s_c), Paragraph(f"<b>{fmt(s['solde'])}</b>", s_r)],
    ]
    t2 = Table(fin, colWidths=[100*mm, 60*mm])
    t2.setStyle(TableStyle([('BACKGROUND',(0,0),(-1,0),ORANGE),('GRID',(0,0),(-1,-1),0.5,HexColor('#ddd')),
        ('TOPPADDING',(0,0),(-1,-1),5),('BOTTOMPADDING',(0,0),(-1,-1),5),
        ('BACKGROUND',(0,7),(-1,7),HexColor('#e8f0f5'))]))
    story.append(t2)
    story.append(Spacer(1, 6*mm))

    # Commercial
    story.append(Paragraph("Commercial", s_h))
    com = [
        [Paragraph(h, s_hd) for h in ['Métrique', 'Valeur']],
        [Paragraph('Devis émis', s_c), Paragraph(str(s['devis_total']), s_r)],
        [Paragraph('Devis acceptés', s_c), Paragraph(str(s['devis_acceptes']), s_r)],
        [Paragraph('Taux conversion', s_c), Paragraph(f"{s['devis_acceptes']*100//max(s['devis_total'],1)}%", s_r)],
        [Paragraph('Factures émises', s_c), Paragraph(str(s['factures_total']), s_r)],
        [Paragraph('Factures payées', s_c), Paragraph(str(s['factures_payees']), s_r)],
    ]
    t3 = Table(com, colWidths=[100*mm, 60*mm])
    t3.setStyle(TableStyle([('BACKGROUND',(0,0),(-1,0),HexColor('#2e7d32')),('GRID',(0,0),(-1,-1),0.5,HexColor('#ddd')),
        ('TOPPADDING',(0,0),(-1,-1),5),('BOTTOMPADDING',(0,0),(-1,-1),5)]))
    story.append(t3)
    story.append(Spacer(1, 6*mm))

    # RH
    story.append(Paragraph("Ressources Humaines", s_h))
    rh = [
        [Paragraph(h, s_hd) for h in ['Indicateur', 'Valeur']],
        [Paragraph('Effectif actif', s_c), Paragraph(str(s['employes']), s_r)],
        [Paragraph('Masse salariale mensuelle', s_c), Paragraph(fmt(s['masse_salariale']), s_r)],
        [Paragraph('Congés en attente', s_c), Paragraph(str(s['conges_pending']), s_r)],
        [Paragraph('Formations planifiées', s_c), Paragraph(str(s['formations']), s_r)],
    ]
    t4 = Table(rh, colWidths=[100*mm, 60*mm])
    t4.setStyle(TableStyle([('BACKGROUND',(0,0),(-1,0),HexColor('#7b1fa2')),('GRID',(0,0),(-1,-1),0.5,HexColor('#ddd')),
        ('TOPPADDING',(0,0),(-1,-1),5),('BOTTOMPADDING',(0,0),(-1,-1),5)]))
    story.append(t4)
    story.append(Spacer(1, 6*mm))

    # KPIs performance
    story.append(Paragraph("Indicateurs de Performance", s_h))
    taux_conv = s['devis_acceptes']*100//max(s['devis_total'],1)
    taux_recouv = s['factures_payees']*100//max(s['factures_total'],1)
    ca_client = s['montant_facture'] // max(s['clients'],1)
    perf = [
        [Paragraph(h, s_hd) for h in ['KPI', 'Valeur', 'Cible']],
        [Paragraph('Taux conversion devis', s_c), Paragraph(f"{taux_conv}%", s_r), Paragraph('> 20%', s_r)],
        [Paragraph('Taux recouvrement', s_c), Paragraph(f"{taux_recouv}%", s_r), Paragraph('> 80%', s_r)],
        [Paragraph('CA moyen / client', s_c), Paragraph(f"{fmt(ca_client)} F", s_r), Paragraph('> 500 000 F', s_r)],
    ]
    t5 = Table(perf, colWidths=[80*mm, 40*mm, 40*mm])
    t5.setStyle(TableStyle([('BACKGROUND',(0,0),(-1,0),NAVY),('GRID',(0,0),(-1,-1),0.5,HexColor('#ddd')),
        ('TOPPADDING',(0,0),(-1,-1),5),('BOTTOMPADDING',(0,0),(-1,-1),5)]))
    story.append(t5)

    story.append(Spacer(1, 15*mm))
    story.append(Paragraph(f"Document généré le {datetime.now().strftime('%d/%m/%Y à %H:%M')} — WannyGest", ParagraphStyle('f', fontSize=8, alignment=TA_CENTER, textColor=HexColor('#999'))))

    doc.build(story)
    return send_file(output, as_attachment=True, download_name=f"Rapport_Executif_{datetime.now().strftime('%Y-%m')}.pdf")


# ======================== MODÈLES DE DEVIS ========================

@app.route('/devis/templates')
@permission_required('proforma')
def devis_templates_page():
    templates = get_devis_templates()
    return render_template('devis_templates.html', page='devis', templates=templates)

@app.route('/devis/templates/add', methods=['POST'])
@permission_required('proforma_edit')
def devis_templates_add():
    from models import db_insert
    items = []
    designations = request.form.getlist('item_designation[]')
    quantities = request.form.getlist('item_qty[]')
    prices = request.form.getlist('item_price[]')
    for d, q, p in zip(designations, quantities, prices):
        if d.strip():
            items.append({'designation': d, 'quantity': float(q or 1), 'unit_price': float(p or 0)})
    db_insert('devis_templates', name=request.form['name'],
        category=request.form.get('category', ''),
        description=request.form.get('description', ''),
        items_json=json.dumps(items),
        notes=request.form.get('notes', ''),
        created_by=session['user_id'])
    flash("Modèle de devis créé", "success")
    return redirect(url_for('devis_templates_page'))

@app.route('/devis/from-template/<int:tid>')
@permission_required('proforma_edit')
def devis_from_template(tid):
    tpl = get_devis_template(tid)
    if not tpl:
        flash("Modèle non trouvé", "error"); return redirect(url_for('devis_templates_page'))
    clients = get_all_clients()
    return render_template('devis_from_template.html', page='devis', tpl=tpl, clients=clients,
                          items=json.loads(tpl['items_json']) if tpl.get('items_json') else [])


# ======================== PIÈCE DE CAISSE SORTIE ========================

@app.route('/caisse-sortie')
@login_required
def caisse_sortie():
    month = request.args.get('month', datetime.now().strftime('%Y-%m'))
    tab = request.args.get('tab', 'sorties')
    sorties = get_caisse_sorties(month=month)
    stats = get_caisse_stats(month=month)
    conn = _gdb()
    entrees = [dict(r) for r in conn.execute("SELECT * FROM caisse_entrees WHERE strftime('%Y-%m',date)=? ORDER BY date DESC", (month,)).fetchall()]
    total_entrees = conn.execute("SELECT COALESCE(SUM(montant),0) FROM caisse_entrees WHERE strftime('%Y-%m',date)=?", (month,)).fetchone()[0]
    conn.close()
    return render_template('caisse_sortie.html', page='caisse_sortie', sorties=sorties, stats=stats, month=month,
        tab=tab, entrees=entrees, total_entrees=total_entrees)

@app.route('/caisse-sortie/demande', methods=['GET','POST'])
@login_required
def caisse_demande():
    """Tout le personnel peut faire une demande."""
    if request.method == 'POST':
        user = get_user_by_id(session['user_id'])
        ref = gen_caisse_ref()
        from models import get_db
        conn = get_db()
        conn.execute("""INSERT INTO caisse_sorties (reference, date, beneficiaire, type_beneficiaire,
            montant, nature, motif, demandeur_id, demandeur_name, sig_beneficiaire) VALUES (?,?,?,?,?,?,?,?,?,?)""",
            (ref, request.form.get('date', datetime.now().strftime('%Y-%m-%d')),
             request.form['beneficiaire'], request.form.get('type_beneficiaire', 'particulier'),
             float(request.form.get('montant', 0) or 0),
             request.form.get('nature', 'espece'), request.form.get('motif', ''),
             session['user_id'], user['full_name'] if user else '?',
             request.form.get('sig_beneficiaire', '')))
        conn.commit(); conn.close()
        log_activity(session['user_id'], user['full_name'] if user else '?',
                    'Caisse', f"Demande sortie {ref} — {float(request.form.get('montant',0)):,.0f} F", request.remote_addr)
        flash(f"Demande de sortie de caisse {ref} envoyée au DG pour validation", "success")
        return redirect(url_for('caisse_sortie'))
    return render_template('caisse_demande.html', page='caisse_sortie')

@app.route('/caisse-sortie/<int:sid>/valider')
@login_required
def caisse_valider(sid):
    from models import get_db
    """DG ou admin valide la demande."""
    user = get_user_by_id(session['user_id'])
    if not user or user['role'] not in ('admin', 'dg', 'directeur'):
        flash("Seul le DG peut valider les sorties de caisse", "error")
        return redirect(url_for('caisse_sortie'))
    conn = get_db()
    conn.execute("UPDATE caisse_sorties SET status='valide', valideur_id=?, valideur_name=?, validated_at=? WHERE id=? AND status='en_attente'",
                 (session['user_id'], user['full_name'], datetime.now().isoformat(), sid))
    conn.commit()
    s = conn.execute("SELECT * FROM caisse_sorties WHERE id=?", (sid,)).fetchone()
    conn.close()
    if s:
        log_activity(session['user_id'], user['full_name'], 'Caisse',
                    f"Sortie {s['reference']} validée — {s['montant']:,.0f} F", request.remote_addr)
    flash("Sortie de caisse validée → transmise à la comptabilité", "success")
    return redirect(url_for('caisse_sortie'))

@app.route('/caisse-sortie/<int:sid>/refuser')
@login_required
def caisse_refuser(sid):
    from models import get_db
    user = get_user_by_id(session['user_id'])
    if not user or user['role'] not in ('admin', 'dg', 'directeur'):
        flash("Seul le DG peut refuser", "error")
        return redirect(url_for('caisse_sortie'))
    conn = get_db()
    conn.execute("UPDATE caisse_sorties SET status='refuse', valideur_id=?, valideur_name=?, validated_at=? WHERE id=? AND status='en_attente'",
                 (session['user_id'], user['full_name'], datetime.now().isoformat(), sid))
    conn.commit(); conn.close()
    flash("Demande refusée", "info")
    return redirect(url_for('caisse_sortie'))

@app.route('/caisse-sortie/<int:sid>/comptabiliser')
@login_required
def caisse_comptabiliser(sid):
    from models import get_db
    """La comptabilité enregistre le décaissement."""
    conn = get_db()
    conn.execute("UPDATE caisse_sorties SET comptabilise=1, comptabilise_at=? WHERE id=? AND status='valide'",
                 (datetime.now().isoformat(), sid))
    # Ajouter dans la trésorerie comme dépense
    s = conn.execute("SELECT * FROM caisse_sorties WHERE id=?", (sid,)).fetchone()
    if s:
        try:
            conn.execute("INSERT INTO treasury (type, amount, description, category, created_by, created_at) VALUES (?,?,?,?,?,?)",
                         ('depense', s['montant'], f"Sortie caisse {s['reference']} — {s['beneficiaire']} — {s['motif']}",
                          'sortie_caisse', session.get('user_id'), datetime.now().isoformat()))
        except: pass
    conn.commit(); conn.close()
    flash("Décaissement comptabilisé", "success")
    return redirect(url_for('caisse_sortie'))

@app.route('/caisse-sortie/<int:sid>/edit', methods=['GET','POST'])
@login_required
def caisse_edit(sid):
    from models import get_db
    user = get_user_by_id(session['user_id'])
    if not user or user['role'] not in ('admin', 'dg', 'directeur'):
        flash("Seul l'admin ou le DG peut modifier", "error"); return redirect(url_for('caisse_sortie'))
    conn = get_db()
    s = conn.execute("SELECT * FROM caisse_sorties WHERE id=?", (sid,)).fetchone()
    conn.close()
    if not s: flash("Non trouvé","error"); return redirect(url_for('caisse_sortie'))
    if request.method == 'POST':
        conn = get_db()
        conn.execute("""UPDATE caisse_sorties SET beneficiaire=?, type_beneficiaire=?, montant=?,
            nature=?, motif=?, date=? WHERE id=?""",
            (request.form['beneficiaire'], request.form.get('type_beneficiaire','particulier'),
             float(request.form.get('montant',0) or 0), request.form.get('nature','espece'),
             request.form.get('motif',''), request.form.get('date',''), sid))
        conn.commit(); conn.close()
        flash("Sortie de caisse modifiée", "success")
        return redirect(url_for('caisse_sortie'))
    return render_template('caisse_edit.html', page='caisse_sortie', s=dict(s))

@app.route('/caisse-sortie/<int:sid>/preview')
@login_required
def caisse_preview(sid):
    from models import get_db
    conn = get_db()
    s = conn.execute("SELECT * FROM caisse_sorties WHERE id=?", (sid,)).fetchone()
    conn.close()
    if not s: flash("Non trouvé","error"); return redirect(url_for('caisse_sortie'))
    return render_template('caisse_preview.html', page='caisse_sortie', s=dict(s))

@app.route('/caisse-sortie/<int:sid>/signer', methods=['POST'])
@login_required
def caisse_signer(sid):
    """Enregistre la signature (base64 canvas) dans une colonne dédiée."""
    from models import get_db
    sig_type = request.form.get('type', 'beneficiaire')
    sig_data = request.form.get('signature', '')
    if sig_data and sig_type in ('beneficiaire', 'caisse', 'autorisation'):
        conn = get_db()
        conn.execute(f"UPDATE caisse_sorties SET sig_{sig_type}=? WHERE id=?", (sig_data, sid))
        conn.commit(); conn.close()
        flash(f"Signature {sig_type} enregistrée ✓", "success")
    return redirect(f'/caisse-sortie/{sid}/preview')

@app.route('/caisse-sortie/<int:sid>/delete')
@login_required
def caisse_delete(sid):
    """Admin supprime une demande de caisse."""
    user = get_user_by_id(session['user_id'])
    if not user or user['role'] not in ('admin', 'dg', 'directeur'):
        flash("Non autorisé", "error"); return redirect(url_for('caisse_sortie'))
    delete_caisse(sid)
    log_activity(session['user_id'], user['full_name'], 'Caisse', f"Sortie #{sid} supprimée", request.remote_addr)
    flash("Demande supprimée", "success")
    return redirect(url_for('caisse_sortie'))

@app.route('/caisse-sortie/<int:sid>/pdf')
@login_required
def caisse_pdf(sid):
    from models import get_db
    """Génère le PDF de la pièce de caisse sortie."""
    conn = get_db()
    s = conn.execute("SELECT * FROM caisse_sorties WHERE id=?", (sid,)).fetchone()
    conn.close()
    if not s: flash("Non trouvé","error"); return redirect(url_for('caisse_sortie'))
    s = dict(s)
    
    from reportlab.lib.pagesizes import A4
    from reportlab.lib.units import mm
    from reportlab.lib.colors import HexColor
    from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer, HRFlowable
    from reportlab.lib.styles import ParagraphStyle
    from reportlab.lib.enums import TA_CENTER, TA_RIGHT, TA_LEFT
    
    output = os.path.join(app.config['UPLOAD_FOLDER'], f'caisse_{s["reference"]}.pdf')
    os.makedirs(os.path.dirname(output), exist_ok=True)
    doc = SimpleDocTemplate(output, pagesize=A4, leftMargin=20*mm, rightMargin=20*mm, topMargin=15*mm, bottomMargin=15*mm)
    
    NAVY = HexColor('#1a3a5c'); RED = HexColor('#c53030')
    s_t = ParagraphStyle('t', fontSize=18, fontName='Helvetica-Bold', textColor=NAVY, alignment=TA_CENTER)
    s_ref = ParagraphStyle('ref', fontSize=12, fontName='Helvetica-Bold', alignment=TA_RIGHT, textColor=RED)
    s_n = ParagraphStyle('n', fontSize=11, leading=14)
    s_b = ParagraphStyle('b', fontSize=11, fontName='Helvetica-Bold')
    s_c = ParagraphStyle('c', fontSize=10)
    s_r = ParagraphStyle('r', fontSize=10, alignment=TA_RIGHT)
    s_h = ParagraphStyle('h', fontSize=10, fontName='Helvetica-Bold', textColor=HexColor('#fff'))
    
    story = []
    # Header
    story.append(Paragraph("PIÈCE DE CAISSE SORTIE", s_t))
    story.append(Paragraph(f"N° {s['reference']}", s_ref))
    story.append(Spacer(1, 3*mm))
    story.append(HRFlowable(width="100%", thickness=2, color=NAVY))
    story.append(Spacer(1, 5*mm))
    
    # Date + lieu
    story.append(Paragraph(f"ABIDJAN le {s.get('date','') or ''}", ParagraphStyle('d', fontSize=11, alignment=TA_CENTER)))
    story.append(Spacer(1, 6*mm))
    
    # Info table
    typ = '☑' if s['type_beneficiaire']=='entreprise' else '☐'
    typ2 = '☑' if s['type_beneficiaire']=='particulier' else '☐'
    typ3 = '☑' if s['type_beneficiaire'] not in ('entreprise','particulier') else '☐'
    
    info = [
        [Paragraph("<b>A l'ordre de :</b>", s_b), Paragraph(f"<b>{s['beneficiaire']}</b>", s_b)],
        [Paragraph(f"Entreprise: {typ}     Particulier: {typ2}     Autres: {typ3}", s_c), Paragraph('', s_c)],
    ]
    it = Table(info, colWidths=[85*mm, 85*mm])
    it.setStyle(TableStyle([('GRID',(0,0),(-1,-1),0.5,HexColor('#ccc')),
        ('TOPPADDING',(0,0),(-1,-1),6),('BOTTOMPADDING',(0,0),(-1,-1),6),('LEFTPADDING',(0,0),(-1,-1),8)]))
    story.append(it)
    story.append(Spacer(1, 4*mm))
    
    # Montant + Nature + Motif
    nature = s.get('nature','espece')
    ch = '...........' if nature!='cheque' else f" N°{s.get('notes','')}"
    vi = '...........' if nature!='virement' else f" ✓"
    es = '...........' if nature!='espece' else f" ✓"
    
    details = [
        [Paragraph(h, s_h) for h in ['Rubrique', 'Détail']],
        [Paragraph('<b>MONTANT</b>', s_c), Paragraph(f"<b>{s['montant']:,.0f} FCFA</b>", ParagraphStyle('m', fontSize=14, fontName='Helvetica-Bold', textColor=RED))],
        [Paragraph('<b>NATURE</b>', s_c), Paragraph(f"Chèque {ch}     Virement {vi}     Espèce {es}", s_c)],
        [Paragraph('<b>MOTIF</b>', s_c), Paragraph(s.get('motif','') or '-', s_n)],
    ]
    dt = Table(details, colWidths=[40*mm, 130*mm])
    dt.setStyle(TableStyle([('BACKGROUND',(0,0),(-1,0),NAVY),('GRID',(0,0),(-1,-1),0.5,HexColor('#ccc')),
        ('TOPPADDING',(0,0),(-1,-1),8),('BOTTOMPADDING',(0,0),(-1,-1),8),('LEFTPADDING',(0,0),(-1,-1),8)]))
    story.append(dt)
    story.append(Spacer(1, 8*mm))
    
    # Status
    if s['status'] == 'valide':
        story.append(Paragraph(f"✅ Validé par {s.get('valideur_name','')} le {(s.get('validated_at','') or '')[:10]}", ParagraphStyle('st', fontSize=10, textColor=HexColor('#2e7d32'))))
    elif s['status'] == 'refuse':
        story.append(Paragraph(f"❌ Refusé par {s.get('valideur_name','')}", ParagraphStyle('st', fontSize=10, textColor=RED)))
    story.append(Spacer(1, 10*mm))
    
    # Signatures
    sig = [[Paragraph("<b>Bénéficiaire</b>", ParagraphStyle('s1', fontSize=10, alignment=TA_CENTER)),
            Paragraph("<b>Caisse</b>", ParagraphStyle('s2', fontSize=10, alignment=TA_CENTER)),
            Paragraph("<b>Autorisation</b>", ParagraphStyle('s3', fontSize=10, alignment=TA_CENTER))]]
    st = Table(sig, colWidths=[57*mm, 57*mm, 57*mm])
    st.setStyle(TableStyle([('LINEABOVE',(0,0),(-1,0),1,HexColor('#000')),
        ('TOPPADDING',(0,0),(-1,-1),6)]))
    story.append(st)
    story.append(Spacer(1, 20*mm))
    story.append(Paragraph("WannyGest — Pièce de caisse générée automatiquement", ParagraphStyle('f', fontSize=7, alignment=TA_CENTER, textColor=HexColor('#999'))))
    
    doc.build(story)
    return send_file(output, as_attachment=True, download_name=f"Caisse_{s['reference']}.pdf")

@app.route('/caisse-sortie/rapport')
@login_required
def caisse_rapport():
    """Rapport mensuel Excel + PDF."""
    month = request.args.get('month', datetime.now().strftime('%Y-%m'))
    fmt = request.args.get('format', 'excel')
    sorties = get_caisse_sorties(status='valide', month=month)
    stats = get_caisse_stats(month=month)
    
    if fmt == 'pdf':
        from reportlab.lib.pagesizes import A4
        from reportlab.lib.units import mm
        from reportlab.lib.colors import HexColor
        from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer, HRFlowable
        from reportlab.lib.styles import ParagraphStyle
        from reportlab.lib.enums import TA_CENTER, TA_RIGHT
        
        output = os.path.join(app.config['UPLOAD_FOLDER'], f'rapport_caisse_{month}.pdf')
        os.makedirs(os.path.dirname(output), exist_ok=True)
        doc = SimpleDocTemplate(output, pagesize=A4, leftMargin=15*mm, rightMargin=15*mm, topMargin=12*mm, bottomMargin=12*mm)
        NAVY = HexColor('#1a3a5c')
        s_t = ParagraphStyle('t', fontSize=16, fontName='Helvetica-Bold', textColor=NAVY, alignment=TA_CENTER)
        s_s = ParagraphStyle('s', fontSize=10, alignment=TA_CENTER, textColor=HexColor('#888'))
        s_h = ParagraphStyle('h', fontSize=8, fontName='Helvetica-Bold', textColor=HexColor('#fff'))
        s_c = ParagraphStyle('c', fontSize=8)
        s_r = ParagraphStyle('r', fontSize=8, alignment=TA_RIGHT)
        s_rb = ParagraphStyle('rb', fontSize=9, fontName='Helvetica-Bold', alignment=TA_RIGHT)
        fmt_n = lambda x: f"{x:,.0f}"
        
        story = []
        story.append(Paragraph(f"RAPPORT DES SORTIES DE CAISSE", s_t))
        story.append(Paragraph(f"Période : {month}", s_s))
        story.append(Spacer(1, 3*mm))
        story.append(HRFlowable(width="100%", thickness=2, color=NAVY))
        story.append(Spacer(1, 5*mm))
        
        # Summary
        summary = [
            [Paragraph(h, s_h) for h in ['Total validé', 'Espèces', 'Chèques', 'Virements', 'Nb opérations']],
            [Paragraph(f"<b>{fmt_n(stats['montant_total'])}</b>", s_rb),
             Paragraph(fmt_n(stats['montant_espece']), s_r),
             Paragraph(fmt_n(stats['montant_cheque']), s_r),
             Paragraph(fmt_n(stats['montant_virement']), s_r),
             Paragraph(str(stats['valide']), s_r)],
        ]
        st = Table(summary, colWidths=[36*mm]*5)
        st.setStyle(TableStyle([('BACKGROUND',(0,0),(-1,0),NAVY),('GRID',(0,0),(-1,-1),0.5,HexColor('#ddd')),
            ('TOPPADDING',(0,0),(-1,-1),5),('BOTTOMPADDING',(0,0),(-1,-1),5)]))
        story.append(st)
        story.append(Spacer(1, 6*mm))
        
        # Detail
        rows = [[Paragraph(h, s_h) for h in ['N°', 'Date', 'Bénéficiaire', 'Motif', 'Nature', 'Montant']]]
        for s_item in sorties:
            rows.append([Paragraph(s_item['reference'], s_c), Paragraph(s_item.get('date','')[:10], s_c),
                Paragraph(s_item['beneficiaire'], s_c), Paragraph((s_item.get('motif','') or '')[:30], s_c),
                Paragraph(s_item.get('nature',''), s_c), Paragraph(fmt_n(s_item['montant']), s_r)])
        rows.append([Paragraph('', s_c)]*4 + [Paragraph('<b>TOTAL</b>', s_c), Paragraph(f"<b>{fmt_n(stats['montant_total'])} FCFA</b>", s_rb)])
        
        dt = Table(rows, colWidths=[28*mm, 20*mm, 40*mm, 35*mm, 20*mm, 27*mm])
        dt.setStyle(TableStyle([('BACKGROUND',(0,0),(-1,0),NAVY),('GRID',(0,0),(-1,-1),0.5,HexColor('#ddd')),
            ('TOPPADDING',(0,0),(-1,-1),3),('BOTTOMPADDING',(0,0),(-1,-1),3),
            ('BACKGROUND',(0,-1),(-1,-1),HexColor('#f0f4f8'))]))
        story.append(dt)
        
        story.append(Spacer(1, 15*mm))
        story.append(Paragraph(f"Généré le {datetime.now().strftime('%d/%m/%Y')} — WannyGest", ParagraphStyle('f', fontSize=7, alignment=TA_CENTER, textColor=HexColor('#999'))))
        doc.build(story)
        return send_file(output, as_attachment=True, download_name=f"Rapport_Caisse_{month}.pdf")
    
    else:  # Excel
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = f"Sorties Caisse {month}"
        
        # Header
        ws.merge_cells('A1:F1')
        ws['A1'] = f"RAPPORT DES SORTIES DE CAISSE — {month}"
        ws['A1'].font = Font(bold=True, size=14, color="1A3A5C")
        ws['A1'].alignment = Alignment(horizontal='center')
        
        # Summary
        ws['A3'] = 'Total validé'; ws['B3'] = stats['montant_total']
        ws['C3'] = 'Espèces'; ws['D3'] = stats['montant_espece']
        ws['E3'] = 'Opérations'; ws['F3'] = stats['valide']
        for cell in ws[3]:
            cell.font = Font(bold=True)
        
        # Headers
        headers = ['Référence', 'Date', 'Bénéficiaire', 'Motif', 'Nature', 'Montant (FCFA)']
        for i, h in enumerate(headers, 1):
            cell = ws.cell(row=5, column=i, value=h)
            cell.font = Font(bold=True, color="FFFFFF")
            cell.fill = PatternFill(start_color="1A3A5C", end_color="1A3A5C", patternType="solid")
        
        # Data
        for row_idx, s_item in enumerate(sorties, 6):
            ws.cell(row=row_idx, column=1, value=s_item['reference'])
            ws.cell(row=row_idx, column=2, value=s_item.get('date',''))
            ws.cell(row=row_idx, column=3, value=s_item['beneficiaire'])
            ws.cell(row=row_idx, column=4, value=s_item.get('motif',''))
            ws.cell(row=row_idx, column=5, value=s_item.get('nature',''))
            ws.cell(row=row_idx, column=6, value=s_item['montant'])
        
        # Total row
        total_row = len(sorties) + 6
        ws.cell(row=total_row, column=5, value='TOTAL').font = Font(bold=True)
        ws.cell(row=total_row, column=6, value=stats['montant_total']).font = Font(bold=True, color="C53030")
        
        for col in range(1, 7):
            ws.column_dimensions[chr(64+col)].width = 18
        
        output = os.path.join(app.config['UPLOAD_FOLDER'], f'rapport_caisse_{month}.xlsx')
        os.makedirs(os.path.dirname(output), exist_ok=True)
        wb.save(output)
        return send_file(output, as_attachment=True, download_name=f"Rapport_Caisse_{month}.xlsx")


# ======================== UTILS ========================

def _cleanup_old(upload_dir, max_age_hours=2):
    import time
    now = time.time()
    try:
        for name in os.listdir(upload_dir):
            path = os.path.join(upload_dir, name)
            if os.path.isdir(path) and (now - os.path.getmtime(path)) > max_age_hours * 3600:
                shutil.rmtree(path, ignore_errors=True)
    except:
        pass


# ======================== RAPPORTS JOURNALIERS ========================
from models import get_db as _gdb, db_insert as _dbi

@app.route('/rapports-journaliers')
@permission_required('rapports_j')
def rapports_journaliers():
    u = dict(get_user_by_id(session['user_id']))
    conn = _gdb()
    if u['role'] in ('admin', 'dg'):
        rapports = [dict(r) for r in conn.execute("""SELECT rj.*, u.full_name FROM rapports_journaliers rj 
            LEFT JOIN users u ON rj.user_id=u.id ORDER BY rj.date DESC, rj.created_at DESC LIMIT 100""").fetchall()]
    else:
        rapports = [dict(r) for r in conn.execute("""SELECT rj.*, u.full_name FROM rapports_journaliers rj 
            LEFT JOIN users u ON rj.user_id=u.id WHERE rj.user_id=? ORDER BY rj.date DESC LIMIT 50""",
            (session['user_id'],)).fetchall()]
    
    # My counter this week
    from datetime import timedelta
    today = datetime.now().date()
    week_start = today - timedelta(days=today.weekday())
    my_week = conn.execute("SELECT COUNT(DISTINCT date) FROM rapports_journaliers WHERE user_id=? AND date>=?",
        (session['user_id'], week_start.strftime('%Y-%m-%d'))).fetchone()[0]
    my_total = conn.execute("SELECT COUNT(*) FROM rapports_journaliers WHERE user_id=?",
        (session['user_id'],)).fetchone()[0]
    conn.close()
    return render_template('rapports_journaliers.html', page='rapports_j', rapports=rapports, user=u,
        my_week=my_week, my_total=my_total)

@app.route('/rapports-journaliers/add', methods=['POST'])
@login_required
def rapports_journaliers_add():
    u = dict(get_user_by_id(session['user_id']))
    _dbi('rapports_journaliers', user_id=session['user_id'],
        date=request.form.get('date', datetime.now().strftime('%Y-%m-%d')),
        tasks_done=request.form.get('tasks_done', ''),
        tasks_planned=request.form.get('tasks_planned', ''),
        issues=request.form.get('issues', ''),
        achievements=request.form.get('achievements', ''),
        completion_pct=int(request.form.get('completion_pct', 0) or 0),
        department=u.get('department', '') or u.get('role', ''))
    flash("Rapport journalier soumis", "success")
    return redirect('/rapports-journaliers')

@app.route('/rapports-journaliers/<int:rid>/validate', methods=['POST'])
@login_required
def rapports_journaliers_validate(rid):
    u = dict(get_user_by_id(session['user_id']))
    if u['role'] not in ('admin', 'dg'):
        flash("Seuls les responsables peuvent valider", "error")
        return redirect('/rapports-journaliers')
    conn = _gdb()
    conn.execute("UPDATE rapports_journaliers SET status=?, validated_by=?, comments=? WHERE id=?",
        (request.form.get('status', 'valide'), session['user_id'], request.form.get('comments', ''), rid))
    conn.commit(); conn.close()
    flash("Rapport validé", "success")
    return redirect('/rapports-journaliers')

@app.route('/rapports-journaliers/assiduite')
@permission_required('rapports_j')
def rapports_assiduite():
    """Tableau d'assiduité — suivi des dépôts de rapports."""
    from datetime import datetime as dt2, timedelta
    conn = _gdb()
    u = dict(get_user_by_id(session['user_id']))
    
    # Current week boundaries (Mon→Sun)
    today = dt2.now().date()
    week_start = today - timedelta(days=today.weekday())
    week_end = week_start + timedelta(days=6)
    
    # Previous weeks for history
    prev_week_start = week_start - timedelta(days=7)
    prev_week_end = week_start - timedelta(days=1)
    
    # Period filter
    period = request.args.get('period', 'semaine')
    if period == 'mois':
        p_start = today.replace(day=1).strftime('%Y-%m-%d')
        p_end = today.strftime('%Y-%m-%d')
        p_label = f"Mois de {today.strftime('%B %Y')}"
        expected_days = sum(1 for d in range((today - today.replace(day=1)).days + 1)
                          if (today.replace(day=1) + timedelta(days=d)).weekday() < 6)
    elif period == 'semaine_prec':
        p_start = prev_week_start.strftime('%Y-%m-%d')
        p_end = prev_week_end.strftime('%Y-%m-%d')
        p_label = f"Semaine du {prev_week_start.strftime('%d/%m')} au {prev_week_end.strftime('%d/%m/%Y')}"
        expected_days = 6
    else:
        p_start = week_start.strftime('%Y-%m-%d')
        p_end = week_end.strftime('%Y-%m-%d')
        p_label = f"Semaine du {week_start.strftime('%d/%m')} au {week_end.strftime('%d/%m/%Y')}"
        expected_days = min(6, (today - week_start).days + 1)  # Only count past days
    
    # Get all reports in period grouped by user
    rows = conn.execute("""
        SELECT rj.user_id, u.full_name, u.role, COUNT(DISTINCT rj.date) as nb_rapports,
               AVG(rj.completion_pct) as avg_completion,
               GROUP_CONCAT(DISTINCT rj.date) as dates
        FROM rapports_journaliers rj
        LEFT JOIN users u ON rj.user_id=u.id
        WHERE rj.date >= ? AND rj.date <= ?
        GROUP BY rj.user_id
        ORDER BY nb_rapports DESC, avg_completion DESC
    """, (p_start, p_end)).fetchall()
    
    ranking = []
    for r in rows:
        d = dict(r)
        d['expected'] = expected_days
        d['taux'] = round((d['nb_rapports'] / max(expected_days, 1)) * 100)
        d['avg_completion'] = round(d['avg_completion'] or 0)
        # Badges
        if d['taux'] >= 100:
            d['badge'] = '🏆 Assidu'
            d['badge_color'] = '#B8860B'
        elif d['taux'] >= 80:
            d['badge'] = '⭐ Régulier'
            d['badge_color'] = '#2e7d32'
        elif d['taux'] >= 60:
            d['badge'] = '👍 Correct'
            d['badge_color'] = '#1565c0'
        elif d['taux'] >= 40:
            d['badge'] = '⚠️ À améliorer'
            d['badge_color'] = '#e8672a'
        else:
            d['badge'] = '🔴 Insuffisant'
            d['badge_color'] = '#c53030'
        ranking.append(d)
    
    # All users for comparison
    all_users = [dict(r) for r in conn.execute("SELECT id, full_name, role FROM users WHERE is_active=1").fetchall()]
    active_ids = {r['user_id'] for r in ranking}
    for usr in all_users:
        if usr['id'] not in active_ids:
            ranking.append({
                'user_id': usr['id'], 'full_name': usr['full_name'], 'role': usr['role'],
                'nb_rapports': 0, 'expected': expected_days, 'taux': 0, 'avg_completion': 0,
                'dates': '', 'badge': '❌ Aucun rapport', 'badge_color': '#c53030'
            })
    
    # Global stats
    total_reports = sum(r['nb_rapports'] for r in ranking)
    total_users = len([r for r in ranking if r['nb_rapports'] > 0])
    assidus = len([r for r in ranking if r['taux'] >= 100])
    
    # Best employee overall
    best = ranking[0] if ranking and ranking[0]['nb_rapports'] > 0 else None
    
    # Split ranking by category
    tech_roles = ['technicien', 'tech_reseau', 'tech_maintenance', 'installateur', 'stagiaire']
    ranking_tech = sorted([r for r in ranking if r.get('role','') in tech_roles], key=lambda x: -x['nb_rapports'])
    ranking_admin = sorted([r for r in ranking if r.get('role','') not in tech_roles], key=lambda x: -x['nb_rapports'])
    
    best_tech = ranking_tech[0] if ranking_tech and ranking_tech[0]['nb_rapports'] > 0 else None
    best_admin = ranking_admin[0] if ranking_admin and ranking_admin[0]['nb_rapports'] > 0 else None
    
    # All stored champions for attestation download
    champions_list = [dict(r) for r in conn.execute("SELECT * FROM weekly_champion ORDER BY week_end DESC LIMIT 10").fetchall()]
    
    conn.close()
    return render_template('rapports_assiduite.html', page='rapports_j',
        ranking=ranking, ranking_tech=ranking_tech, ranking_admin=ranking_admin,
        period=period, p_label=p_label, expected_days=expected_days,
        total_reports=total_reports, total_users=total_users, assidus=assidus,
        best=best, best_tech=best_tech, best_admin=best_admin,
        user=u, today=today.strftime('%Y-%m-%d'), champions_list=champions_list)

@app.route('/rapports-journaliers/attestation/<int:champ_id>')
@permission_required('rapports_j')
def rapports_attestation(champ_id):
    """Génère l'attestation d'encouragement PDF."""
    from models import get_db as _gdb2
    conn = _gdb2()
    champ = conn.execute("SELECT * FROM weekly_champion WHERE id=?", (champ_id,)).fetchone()
    conn.close()
    if not champ:
        flash("Champion non trouvé", "error")
        return redirect('/rapports-journaliers/assiduite')
    champ = dict(champ)
    
    from reportlab.lib.pagesizes import A4, landscape
    from reportlab.lib.units import mm, cm
    from reportlab.lib.colors import HexColor, white, black
    from reportlab.platypus import SimpleDocTemplate, Paragraph, Spacer, Table, TableStyle, HRFlowable, Frame, PageTemplate
    from reportlab.lib.styles import ParagraphStyle
    from reportlab.lib.enums import TA_CENTER, TA_LEFT, TA_RIGHT
    from reportlab.graphics.shapes import Drawing, Rect, Line, String
    from reportlab.graphics import renderPDF
    
    output = os.path.join(app.config['UPLOAD_FOLDER'], f'attestation_{champ_id}.pdf')
    os.makedirs(os.path.dirname(output), exist_ok=True)
    
    pw, ph = landscape(A4)
    doc = SimpleDocTemplate(output, pagesize=landscape(A4),
        leftMargin=30*mm, rightMargin=30*mm, topMargin=15*mm, bottomMargin=15*mm)
    
    GOLD = HexColor('#B8860B')
    GOLD_LIGHT = HexColor('#DAA520')
    NAVY = HexColor('#1a1a2e')
    GREY = HexColor('#555555')
    LGREY = HexColor('#999999')
    
    usable_w = pw - 60*mm
    
    story = []
    
    # ===== TOP LINE =====
    top_line = Table([['']], colWidths=[usable_w], rowHeights=[1*mm])
    top_line.setStyle(TableStyle([('LINEABOVE', (0,0), (-1,0), 3, GOLD)]))
    story.append(top_line)
    story.append(Spacer(1, 8*mm))
    
    # ===== COMPANY =====
    story.append(Paragraph("RAMYA TECHNOLOGIE &amp; INNOVATION",
        ParagraphStyle('company', fontName='Helvetica', fontSize=9, textColor=LGREY, 
                       alignment=TA_CENTER, spaceAfter=8*mm)))
    
    # ===== TITLE =====
    story.append(Paragraph("ATTESTATION D'ENCOURAGEMENT",
        ParagraphStyle('title', fontName='Helvetica-Bold', fontSize=24, textColor=NAVY, 
                       alignment=TA_CENTER, spaceAfter=12*mm)))
    
    # ===== LINE =====
    story.append(HRFlowable(width="35%", thickness=2, color=GOLD, spaceAfter=12*mm))
    
    # ===== SUBTITLE =====
    story.append(Paragraph("Employ\u00e9 le plus assidu de la semaine",
        ParagraphStyle('subtitle', fontName='Helvetica-Bold', fontSize=12, textColor=GOLD, 
                       alignment=TA_CENTER, spaceAfter=10*mm)))
    
    # ===== INTRO =====
    story.append(Paragraph("Nous avons l'honneur de d\u00e9cerner cette attestation \u00e0 :",
        ParagraphStyle('intro', fontName='Helvetica', fontSize=10, textColor=GREY, 
                       alignment=TA_CENTER, spaceAfter=8*mm)))
    
    # ===== NAME =====
    story.append(Paragraph(f"<b>{champ['full_name']}</b>",
        ParagraphStyle('fullname', fontName='Helvetica-Bold', fontSize=28, textColor=NAVY, 
                       alignment=TA_CENTER, spaceAfter=10*mm)))
    
    # ===== ROLE =====
    role_dept = champ['role']
    if champ.get('department'):
        role_dept += f" - {champ['department']}"
    story.append(Paragraph(role_dept,
        ParagraphStyle('roledept', fontName='Helvetica', fontSize=10, textColor=LGREY, 
                       alignment=TA_CENTER, spaceAfter=12*mm)))
    
    # ===== LINE =====
    story.append(HRFlowable(width="35%", thickness=1, color=GOLD_LIGHT, spaceAfter=8*mm))
    
    # ===== REASON =====
    story.append(Paragraph("Pour son assiduit\u00e9 exemplaire dans le d\u00e9p\u00f4t des rapports journaliers",
        ParagraphStyle('reason', fontName='Helvetica', fontSize=11, textColor=HexColor('#333333'), 
                       alignment=TA_CENTER, spaceAfter=4*mm)))
    
    story.append(Paragraph(f"Semaine du <b>{champ['week_start']}</b> au <b>{champ['week_end']}</b>",
        ParagraphStyle('period', fontName='Helvetica', fontSize=10, textColor=GREY, 
                       alignment=TA_CENTER, spaceAfter=8*mm)))
    
    # ===== STATS =====
    s_label = ParagraphStyle('sl', fontName='Helvetica', fontSize=8, textColor=LGREY, alignment=TA_CENTER)
    stats_data = [[
        Paragraph(f"<b>{champ['nb_rapports']}</b>", ParagraphStyle('sn', fontName='Helvetica-Bold', fontSize=20, textColor=GOLD, alignment=TA_CENTER)),
        Paragraph(f"<b>{int(champ['avg_completion'])}%</b>", ParagraphStyle('sc', fontName='Helvetica-Bold', fontSize=20, textColor=HexColor('#2e7d32'), alignment=TA_CENTER)),
        Paragraph("<b>ASSIDU</b>", ParagraphStyle('sb', fontName='Helvetica-Bold', fontSize=14, textColor=GOLD, alignment=TA_CENTER)),
    ], [
        Paragraph("rapports d\u00e9pos\u00e9s", s_label),
        Paragraph("r\u00e9alisation moyenne", s_label),
        Paragraph("badge d'assiduit\u00e9", s_label),
    ]]
    cw = usable_w / 3
    st = Table(stats_data, colWidths=[cw, cw, cw], rowHeights=[12*mm, 5*mm])
    st.setStyle(TableStyle([
        ('ALIGN', (0,0), (-1,-1), 'CENTER'), ('VALIGN', (0,0), (-1,-1), 'MIDDLE'),
        ('LINEABOVE', (0,0), (-1,0), 0.5, HexColor('#eeeeee')),
        ('LINEBELOW', (0,-1), (-1,-1), 0.5, HexColor('#eeeeee')),
        ('INNERGRID', (0,0), (-1,-1), 0.3, HexColor('#eeeeee')),
    ]))
    story.append(st)
    story.append(Spacer(1, 5*mm))
    
    # ===== FOOTER =====
    now_str = datetime.now().strftime("%d/%m/%Y")
    story.append(Paragraph(f"Fait le {now_str} - RAMYA TECHNOLOGIE &amp; INNOVATION",
        ParagraphStyle('date', fontName='Helvetica', fontSize=8, textColor=LGREY, alignment=TA_CENTER, spaceAfter=3*mm)))
    
    # ===== BOTTOM LINE =====
    bot_line = Table([['']], colWidths=[usable_w], rowHeights=[1*mm])
    bot_line.setStyle(TableStyle([('LINEBELOW', (0,0), (-1,0), 3, GOLD)]))
    story.append(bot_line)
    
    doc.build(story)
    return send_file(output, as_attachment=True, download_name=f"Attestation_{champ['full_name'].replace(' ','_')}.pdf")


# ======================== PIÈCES DE CAISSE / DÉPENSES ========================

@app.route('/comptabilite/pieces')
@permission_required('comptabilite')
def pieces_caisse():
    conn = _gdb()
    pieces = [dict(r) for r in conn.execute("SELECT * FROM pieces_caisse ORDER BY date DESC").fetchall()]
    total = conn.execute("SELECT COALESCE(SUM(amount),0) FROM pieces_caisse").fetchone()[0]
    by_cat = [dict(r) for r in conn.execute("SELECT category, SUM(amount) as total, COUNT(*) as count FROM pieces_caisse GROUP BY category ORDER BY total DESC").fetchall()]
    conn.close()
    return render_template('pieces_caisse.html', page='pieces', pieces=pieces, total=total, by_cat=by_cat)

@app.route('/comptabilite/pieces/add', methods=['POST'])
@permission_required('comptabilite')
def pieces_caisse_add():
    ref = f"PC-{datetime.now().strftime('%Y%m%d%H%M%S')}"
    file_path = ''
    if 'file' in request.files and request.files['file'].filename:
        f = request.files['file']
        ext = os.path.splitext(f.filename)[1].lower()
        if ext in ('.jpg', '.jpeg', '.png', '.pdf', '.webp'):
            fname = f"piece_{ref}{ext}"
            fdir = os.path.join(app.config['UPLOAD_FOLDER'], 'pieces')
            os.makedirs(fdir, exist_ok=True)
            f.save(os.path.join(fdir, fname))
            file_path = fname
    
    amount = float(request.form.get('amount', 0) or 0)
    _dbi('pieces_caisse', reference=ref,
        date=request.form.get('date', datetime.now().strftime('%Y-%m-%d')),
        description=request.form.get('description', ''),
        amount=amount, category=request.form.get('category', 'divers'),
        supplier=request.form.get('supplier', ''),
        file_path=file_path, created_by=session['user_id'])
    flash(f"Pièce {ref} — {amount:,.0f} F enregistrée", "success")
    return redirect('/comptabilite/pieces')

@app.route('/uploads/pieces/<path:filename>')
def piece_file(filename):
    return send_from_directory(os.path.join(app.config['UPLOAD_FOLDER'], 'pieces'), filename)

@app.route('/uploads/stock/<path:filename>')
def stock_image(filename):
    return send_from_directory(os.path.join(app.config['UPLOAD_FOLDER'], 'stock'), filename)


# ======================== MODULE ACHATS ========================

@app.route('/achats')
@permission_required('comptabilite')
def achats_page():
    tab = request.args.get('tab', 'fournisseurs')
    conn = _gdb()
    from models import db_get_all
    
    data = {'tab': tab}
    data['stock_items'] = db_get_all('stock_items', order='name ASC')
    data['fournisseurs'] = [dict(r) for r in conn.execute("SELECT * FROM achats_fournisseurs ORDER BY name").fetchall()]
    data['demandes'] = [dict(r) for r in conn.execute("""SELECT ad.*, u.full_name as requester FROM achats_demandes ad 
        LEFT JOIN users u ON ad.requested_by=u.id ORDER BY ad.created_at DESC LIMIT 50""").fetchall()]
    data['devis_achats'] = [dict(r) for r in conn.execute("""SELECT ad.*, f.name as fournisseur_name FROM achats_devis ad
        LEFT JOIN achats_fournisseurs f ON ad.fournisseur_id=f.id ORDER BY ad.created_at DESC LIMIT 50""").fetchall()]
    data['commandes'] = [dict(r) for r in conn.execute("""SELECT ac.*, f.name as fournisseur_name FROM achats_commandes ac
        LEFT JOIN achats_fournisseurs f ON ac.fournisseur_id=f.id ORDER BY ac.created_at DESC LIMIT 50""").fetchall()]
    data['contrats'] = [dict(r) for r in conn.execute("""SELECT ac.*, f.name as fournisseur_name FROM achats_contrats ac
        LEFT JOIN achats_fournisseurs f ON ac.fournisseur_id=f.id ORDER BY ac.created_at DESC""").fetchall()]
    
    # Stats
    data['total_commandes'] = conn.execute("SELECT COALESCE(SUM(total),0) FROM achats_commandes").fetchone()[0]
    data['pending_demandes'] = conn.execute("SELECT COUNT(*) FROM achats_demandes WHERE status='en_attente'").fetchone()[0]
    conn.close()
    return render_template('achats.html', page='achats', **data)

@app.route('/achats/fournisseur/add', methods=['POST'])
@permission_required('comptabilite_edit')
def achats_fournisseur_add():
    _dbi('achats_fournisseurs', name=request.form['name'], contact_name=request.form.get('contact_name',''),
        tel=request.form.get('tel',''), email=request.form.get('email',''),
        address=request.form.get('address',''), city=request.form.get('city',''),
        sector=request.form.get('sector',''), payment_terms=request.form.get('payment_terms',''))
    flash("Fournisseur ajouté", "success"); return redirect('/achats?tab=fournisseurs')

@app.route('/achats/fournisseur/edit/<int:fid>', methods=['POST'])
@permission_required('comptabilite_edit')
def achats_fournisseur_edit(fid):
    conn = _gdb()
    conn.execute("""UPDATE achats_fournisseurs SET name=?,contact_name=?,tel=?,email=?,address=?,city=?,sector=?,payment_terms=?,notes=? WHERE id=?""",
        (request.form['name'],request.form.get('contact_name',''),request.form.get('tel',''),request.form.get('email',''),
         request.form.get('address',''),request.form.get('city',''),request.form.get('sector',''),
         request.form.get('payment_terms',''),request.form.get('notes',''),fid))
    conn.commit(); conn.close()
    flash("Fournisseur modifié", "success"); return redirect('/achats?tab=fournisseurs')

@app.route('/achats/fournisseur/delete/<int:fid>')
@permission_required('comptabilite_edit')
def achats_fournisseur_delete(fid):
    conn = _gdb(); conn.execute("DELETE FROM achats_fournisseurs WHERE id=?", (fid,)); conn.commit(); conn.close()
    flash("Fournisseur supprimé", "success"); return redirect('/achats?tab=fournisseurs')

@app.route('/achats/demande/add', methods=['POST'])
@permission_required('comptabilite_edit')
def achats_demande_add():
    ref = f"DA-{datetime.now().strftime('%Y%m%d%H%M%S')}"
    _dbi('achats_demandes', reference=ref, date=request.form.get('date', datetime.now().strftime('%Y-%m-%d')),
        department=request.form.get('department',''), requested_by=session['user_id'],
        description=request.form.get('description',''), urgency=request.form.get('urgency','normale'))
    flash(f"Demande {ref} créée", "success"); return redirect('/achats?tab=demandes')

@app.route('/achats/demande/<int:did>/approve')
@permission_required('comptabilite_edit')
def achats_demande_approve(did):
    conn = _gdb()
    conn.execute("UPDATE achats_demandes SET status='approuvee', approved_by=?, approved_at=CURRENT_TIMESTAMP WHERE id=?",
        (session['user_id'], did))
    conn.commit(); conn.close()
    flash("Demande approuvée", "success"); return redirect('/achats?tab=demandes')

@app.route('/achats/demande/<int:did>/reject')
@permission_required('comptabilite_edit')
def achats_demande_reject(did):
    conn = _gdb()
    conn.execute("UPDATE achats_demandes SET status='refusee' WHERE id=?", (did,))
    conn.commit(); conn.close()
    flash("Demande refusée", "success"); return redirect('/achats?tab=demandes')

@app.route('/achats/devis/add', methods=['POST'])
@permission_required('proforma_edit')
def achats_devis_add():
    ref = f"DAC-{datetime.now().strftime('%Y%m%d%H%M%S')}"
    total_ht = float(request.form.get('total_ht', 0) or 0)
    tva = total_ht * 0.18
    _dbi('achats_devis', reference=ref, fournisseur_id=int(request.form.get('fournisseur_id',0) or 0),
        date=request.form.get('date', datetime.now().strftime('%Y-%m-%d')),
        items_json=request.form.get('items_description',''), total_ht=total_ht, tva=tva,
        total_ttc=total_ht + tva, status='en_attente', notes=request.form.get('notes',''),
        created_by=session['user_id'])
    flash(f"Devis fournisseur {ref} enregistré", "success"); return redirect('/achats?tab=devis')

@app.route('/achats/devis/<int:did>/status/<status>')
@permission_required('proforma_edit')
def achats_devis_status(did, status):
    if status in ('en_attente', 'accepte', 'refuse'):
        conn = _gdb(); conn.execute("UPDATE achats_devis SET status=? WHERE id=?", (status, did)); conn.commit(); conn.close()
    flash("Statut mis à jour", "success"); return redirect('/achats?tab=devis')

@app.route('/achats/commande/add', methods=['POST'])
@permission_required('comptabilite_edit')
def achats_commande_add():
    ref = f"BC-{datetime.now().strftime('%Y%m%d%H%M%S')}"
    _dbi('achats_commandes', reference=ref, fournisseur_id=int(request.form.get('fournisseur_id',0) or 0),
        date=request.form.get('date', datetime.now().strftime('%Y-%m-%d')),
        items_json=request.form.get('items_description',''),
        total=float(request.form.get('total',0) or 0),
        delivery_date=request.form.get('delivery_date',''),
        notes=request.form.get('notes',''), created_by=session['user_id'])
    flash(f"Bon de commande {ref} créé", "success"); return redirect('/achats?tab=commandes')

@app.route('/achats/commande/<int:cid>/status/<status>')
@permission_required('comptabilite_edit')
def achats_commande_status(cid, status):
    if status in ('en_cours', 'livree', 'annulee'):
        conn = _gdb(); conn.execute("UPDATE achats_commandes SET status=? WHERE id=?", (status, cid)); conn.commit(); conn.close()
    flash("Statut mis à jour", "success"); return redirect('/achats?tab=commandes')

@app.route('/achats/contrat/add', methods=['POST'])
@permission_required('comptabilite_edit')
def achats_contrat_add():
    ref = f"CTR-{datetime.now().strftime('%Y%m%d%H%M%S')}"
    _dbi('achats_contrats', reference=ref, fournisseur_id=int(request.form.get('fournisseur_id',0) or 0),
        title=request.form.get('title',''), description=request.form.get('description',''),
        start_date=request.form.get('start_date',''), end_date=request.form.get('end_date',''),
        amount=float(request.form.get('amount',0) or 0), created_by=session['user_id'])
    flash("Contrat ajouté", "success"); return redirect('/achats?tab=contrats')


if __name__ == '__main__':
    port = int(os.environ.get('PORT', 5000))
    app.run(host='0.0.0.0', port=port, debug=False)
